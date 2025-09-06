#ui/teachers_form.py
import sys
import os
import csv
import traceback
import subprocess
import platform
from datetime import datetime
from typing import Optional, Dict, Any
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QMessageBox, QFileDialog, QScrollArea, QFrame, QSizePolicy,
    QGroupBox, QGridLayout, QSpacerItem, QComboBox, QFormLayout, 
    QTabWidget, QMenu, QCheckBox, QDateEdit, QTextEdit, QApplication, QLineEdit
)
from PySide6.QtGui import QFont, QPalette, QIcon, QPixmap, QPainter, QAction
from PySide6.QtCore import Qt, Signal, QSize, QDate, QTimer
import mysql.connector
from mysql.connector import Error
from PIL import Image, ImageQt
from utils.permissions import has_permission
from ui.audit_base_form import AuditBaseForm
from ui.departments_form import DepartmentsForm
from utils.pdf_utils import view_pdf
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from models.models import get_db_connection


# Change class definition
class TeachersForm(AuditBaseForm):
    teacher_selected = Signal(int)
    def __init__(self, parent=None, user_session=None):
        super().__init__(parent, user_session)
        self.user_session = user_session
        self.current_teacher_id = None
        self.photo_path = None
        self.teacher_data = {}
        
        # Set up modern styling
        self.setup_styling()
        
        # Database connection
        try:
            self.db_connection = get_db_connection()
            self.cursor = self.db_connection.cursor(buffered=True)
        except Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to connect to database: {e}")
            return
        
        # Ensure photos directory exists
        self.photos_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'photos', 'teachers')
        os.makedirs(self.photos_dir, exist_ok=True)
        
        self.setup_ui()
        self.load_teachers()
        self.load_schools()
        self.load_departments_combo()
        self.apply_button_permissions()

    def setup_ui(self):
        """Set up the user interface"""
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Create tab widget
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # Create tabs - ADD ANALYTICS TAB
        self.teacher_form_tab = QWidget()
        self.teacher_data_tab = QWidget()
        self.analytics_tab = QWidget()  # NEW ANALYTICS TAB
        self.departments_tab = QWidget()
        
        # Add tabs to widget - ADD ANALYTICS TAB
        self.tab_widget.addTab(self.teacher_form_tab, "Staff Form")
        self.tab_widget.addTab(self.teacher_data_tab, "Staff Data")
        self.tab_widget.addTab(self.analytics_tab, "Staff Analytics")  # NEW TAB
        self.departments_tab = DepartmentsForm(self.tab_widget, user_session=self.user_session)
        self.tab_widget.addTab(self.departments_tab, "Departments")
        
        # Setup each tab - ADD ANALYTICS TAB SETUP
        self.setup_teacher_form_tab()
        self.setup_teacher_data_tab()
        self.setup_analytics_tab()  # NEW METHOD - THIS SHOULD NOW WORK
        self.setup_departments_tab()

    def setup_teacher_form_tab(self):
        """Set up the teacher form tab"""
        layout = QHBoxLayout(self.teacher_form_tab)
    
        # Left side - Form fields (70% width)
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setMinimumWidth(800)
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
    
        # Title
        title_label = QLabel("Staff Information")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""...""")
        left_layout.addWidget(title_label)
    
        # Create form sections
        self.create_personal_info_section(left_layout)
        self.create_contact_info_section(left_layout)
        self.create_employment_info_section(left_layout)
    
        # === Status Label ===
        self.status_label = QLabel("Ready")
        self.status_label.setFont(self.fonts['entry'])
        self.status_label.setStyleSheet(f"color: {self.colors['info']}; font-weight: bold;")
    
        status_layout = QHBoxLayout()
        status_layout.addStretch()
        status_layout.addWidget(self.status_label)
        status_layout.setContentsMargins(10, 5, 10, 10)
    
        left_layout.addLayout(status_layout)  # ✅ Correct layout used
    
        left_scroll.setWidget(left_widget)
        layout.addWidget(left_scroll, 6)  # 60% width
    
        # Right side - Photo and actions (30% width)
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_widget.setMaximumWidth(400)
    
        self.create_photo_section(right_layout)
        self.create_action_buttons(right_layout)
    
        layout.addWidget(right_widget, 3)  # 30% width

    def create_personal_info_section(self, parent_layout):
        """Create personal information section"""
        personal_group = QGroupBox("Personal Information")
        personal_layout = QGridLayout(personal_group)
        
        # Row 1: Teacher ID, Salutation, First Name
        row = 0
        personal_layout.addWidget(QLabel("Staff ID Code:"), row, 0)
        self.teacher_id_entry = QLineEdit()
        personal_layout.addWidget(self.teacher_id_entry, row, 1)
        
        personal_layout.addWidget(QLabel("Salutation:"), row, 2)
        self.salutation_combo = QComboBox()
        self.salutation_combo.addItems(["Mr.", "Mrs.", "Ms.", "Dr.", "Prof."])
        personal_layout.addWidget(self.salutation_combo, row, 3)
        
        personal_layout.addWidget(QLabel("First Name:"), row, 4)
        self.first_name_entry = QLineEdit()
        self.first_name_entry.textChanged.connect(self.update_full_name)
        personal_layout.addWidget(self.first_name_entry, row, 5)
        
        # Row 2: Surname, Full Name, Gender
        row += 1
        personal_layout.addWidget(QLabel("Surname:"), row, 0)
        self.surname_entry = QLineEdit()
        self.surname_entry.textChanged.connect(self.update_full_name)
        personal_layout.addWidget(self.surname_entry, row, 1)
        
        personal_layout.addWidget(QLabel("Full Name:"), row, 2)
        self.full_name_entry = QLineEdit()
        self.full_name_entry.setReadOnly(True)
        personal_layout.addWidget(self.full_name_entry, row, 3)
        
        personal_layout.addWidget(QLabel("Gender:"), row, 4)
        self.gender_combo = QComboBox()
        self.gender_combo.addItems(["Male", "Female"])
        personal_layout.addWidget(self.gender_combo, row, 5)
        
        # Row 3: Birth Date, National ID, Next of Kin
        row += 1
        personal_layout.addWidget(QLabel("Birth Date:"), row, 0)
        self.birth_date_edit = QDateEdit()
        self.birth_date_edit.setCalendarPopup(True)
        self.birth_date_edit.setDate(QDate.currentDate())
        personal_layout.addWidget(self.birth_date_edit, row, 1)
        
        personal_layout.addWidget(QLabel("National ID:"), row, 2)
        self.national_id_entry = QLineEdit()
        personal_layout.addWidget(self.national_id_entry, row, 3)
        
        personal_layout.addWidget(QLabel("Next of Kin:"), row, 4)
        self.next_of_kin_entry = QLineEdit()
        personal_layout.addWidget(self.next_of_kin_entry, row, 5)
        
        parent_layout.addWidget(personal_group)

    def create_contact_info_section(self, parent_layout):
        """Create contact information section"""
        contact_group = QGroupBox("Contact & Address Information")
        contact_layout = QGridLayout(contact_group)
        
        # Row 1: Email, Phone Contact 1
        row = 0
        contact_layout.addWidget(QLabel("Email:"), row, 0)
        self.email_entry = QLineEdit()
        contact_layout.addWidget(self.email_entry, row, 1, 1, 2)  # Span 2 columns
        
        contact_layout.addWidget(QLabel("Phone Contact 1:"), row, 3)
        self.phone_contact_1_entry = QLineEdit()
        contact_layout.addWidget(self.phone_contact_1_entry, row, 4)
        
        # Row 2: Day Phone, Home District
        row += 1
        contact_layout.addWidget(QLabel("Day Phone:"), row, 0)
        self.day_phone_entry = QLineEdit()
        contact_layout.addWidget(self.day_phone_entry, row, 1)
        
        contact_layout.addWidget(QLabel("Home District:"), row, 2)
        self.home_district_entry = QLineEdit()
        contact_layout.addWidget(self.home_district_entry, row, 3, 1, 2)
        
        # Row 3: Position, Current Address
        row += 1
        contact_layout.addWidget(QLabel("Position:"), row, 0)
        self.position_combo = QComboBox()
        self.position_combo.setEditable(True)
        self.position_combo.addItems(["Teacher", "Accountant", "Bursar", "HR", "Registrar", "Librarian"])
        contact_layout.addWidget(self.position_combo, row, 1)
        
        contact_layout.addWidget(QLabel("Current Address:"), row, 2)
        self.current_address_entry = QLineEdit()
        contact_layout.addWidget(self.current_address_entry, row, 3, 1, 2)
        
        # Row 4: Emergency Contacts
        row += 1
        contact_layout.addWidget(QLabel("Emergency Contact 1:"), row, 0)
        self.emergency_contact_1_entry = QLineEdit()
        contact_layout.addWidget(self.emergency_contact_1_entry, row, 1, 1, 2)
        
        contact_layout.addWidget(QLabel("Emergency Contact 2:"), row, 3)
        self.emergency_contact_2_entry = QLineEdit()
        contact_layout.addWidget(self.emergency_contact_2_entry, row, 4)
        
        parent_layout.addWidget(contact_group)

    def create_employment_info_section(self, parent_layout):
        """Create employment information section"""
        employment_group = QGroupBox("Employment Information")
        employment_layout = QGridLayout(employment_group)
        
        # Row 1: School, Subject Specialty
        row = 0
        employment_layout.addWidget(QLabel("School:"), row, 0)
        self.school_combo = QComboBox()
        self.school_combo.setMinimumWidth(250)
        employment_layout.addWidget(self.school_combo, row, 1, 1, 2)
        
        employment_layout.addWidget(QLabel("Subject Specialty:"), row, 3)
        self.subject_specialty_entry = QLineEdit()
        employment_layout.addWidget(self.subject_specialty_entry, row, 4)
        
        # Row 2 continuation: Add Department
        # Inside create_employment_info_section(), after subject_specialty
        row += 1
        employment_layout.addWidget(QLabel("Department:"), row, 0)
        self.department_combo = QComboBox()
        self.department_combo.setMinimumWidth(250)
        employment_layout.addWidget(self.department_combo, row, 1, 1, 2)
        
        # Row 2: Qualification, Date Joined, Staff Type
        row += 1
        employment_layout.addWidget(QLabel("Qualification:"), row, 0)
        self.qualification_entry = QLineEdit()
        employment_layout.addWidget(self.qualification_entry, row, 1)
        
        employment_layout.addWidget(QLabel("Date Joined:"), row, 2)
        self.date_joined_edit = QDateEdit()
        self.date_joined_edit.setCalendarPopup(True)
        self.date_joined_edit.setDate(QDate.currentDate())
        employment_layout.addWidget(self.date_joined_edit, row, 3)
        
        employment_layout.addWidget(QLabel("Staff Type:"), row, 4)
        self.staff_type_combo = QComboBox()
        self.staff_type_combo.addItems(["Teaching", "Administrative", "Support", "Intern", "Volunteer", "Contractor"])
        employment_layout.addWidget(self.staff_type_combo, row, 5)
        
        # Row 3: Employment Status, Bank Account, Active Status
        row += 1
        employment_layout.addWidget(QLabel("Employment Status:"), row, 0)
        self.employment_status_combo = QComboBox()
        self.employment_status_combo.addItems(["Full-time", "Part-time", "Contract", "Probation", "Terminated"])
        employment_layout.addWidget(self.employment_status_combo, row, 1)
        
        employment_layout.addWidget(QLabel("Bank Account:"), row, 2)
        self.bank_account_entry = QLineEdit()
        employment_layout.addWidget(self.bank_account_entry, row, 3)
        
        employment_layout.addWidget(QLabel("Monthly Salary:"), row, 4)
        self.monthly_salary_entry = QLineEdit()
        self.monthly_salary_entry.setPlaceholderText("0.00")
        employment_layout.addWidget(self.monthly_salary_entry, row, 5)
        
        # Row 4: Active checkbox
        row += 1
        self.is_active_checkbox = QCheckBox("Active Staff Member")
        self.is_active_checkbox.setChecked(True)
        employment_layout.addWidget(self.is_active_checkbox, row, 0, 1, 2)
        
        parent_layout.addWidget(employment_group)

    def create_photo_section(self, parent_layout):
        """Create photo section"""
        photo_group = QGroupBox("Staff Photo")
        photo_layout = QVBoxLayout(photo_group)
    
        # Photo display
        self.photo_label = QLabel("No Photo\nSelected")
        self.photo_label.setProperty("class", "photo")
        self.photo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.photo_label.setScaledContents(False)
        photo_layout.addWidget(self.photo_label)
    
        # Photo buttons
        photo_btn_layout = QHBoxLayout()
    
        self.select_photo_btn = QPushButton("Select Photo")
        self.select_photo_btn.setProperty("class", "info")
        self.select_photo_btn.setIcon(QIcon("static/icons/upload.png"))
        self.select_photo_btn.setIconSize(QSize(20, 20))
        self.select_photo_btn.clicked.connect(self.select_photo)
        photo_btn_layout.addWidget(self.select_photo_btn)
    
        self.remove_photo_btn = QPushButton("Remove")
        self.remove_photo_btn.setProperty("class", "danger")
        self.remove_photo_btn.setIcon(QIcon("static/icons/remove.png"))
        self.remove_photo_btn.setIconSize(QSize(20, 20))
        self.remove_photo_btn.clicked.connect(self.remove_photo)
        photo_btn_layout.addWidget(self.remove_photo_btn)
    
        photo_layout.addLayout(photo_btn_layout)
        parent_layout.addWidget(photo_group)
    
    
    def create_action_buttons(self, parent_layout):
        """Create action buttons section"""
        actions_group = QGroupBox("Actions")
        actions_layout = QVBoxLayout(actions_group)
    
        # === Primary actions ===
        primary_layout = QHBoxLayout()
    
        self.add_btn = QPushButton("Add Teacher")
        self.add_btn.setProperty("class", "success")
        self.add_btn.setIcon(QIcon("static/icons/add.png"))
        self.add_btn.setIconSize(QSize(20, 20))
        self.add_btn.clicked.connect(self.add_teacher)
        primary_layout.addWidget(self.add_btn)
    
        self.update_btn = QPushButton("Update")
        self.update_btn.setProperty("class", "primary")
        self.update_btn.setIcon(QIcon("static/icons/update.png"))
        self.update_btn.setIconSize(QSize(20, 20))
        self.update_btn.clicked.connect(self.update_teacher)
        primary_layout.addWidget(self.update_btn)
    
        actions_layout.addLayout(primary_layout)
    
        # === Secondary actions ===
        secondary_layout = QHBoxLayout()
    
        self.delete_btn = QPushButton("Delete")
        self.delete_btn.setProperty("class", "danger")
        self.delete_btn.setIcon(QIcon("static/icons/delete.png"))
        self.delete_btn.setIconSize(QSize(20, 20))
        self.delete_btn.clicked.connect(self.delete_teacher)
        secondary_layout.addWidget(self.delete_btn)
    
        self.clear_btn = QPushButton("Clear Fields")
        self.clear_btn.setProperty("class", "warning")
        self.clear_btn.setIcon(QIcon("static/icons/clear.png"))
        self.clear_btn.setIconSize(QSize(20, 20))
        self.clear_btn.clicked.connect(self.clear_fields)
        secondary_layout.addWidget(self.clear_btn)
    
        actions_layout.addLayout(secondary_layout)
    
        # === Utility actions ===
        utility_layout = QHBoxLayout()
    
        self.edit_selected_btn = QPushButton("Edit Selected")
        self.edit_selected_btn.setIcon(QIcon("static/icons/edit.png"))
        self.edit_selected_btn.setIconSize(QSize(20, 20))
        self.edit_selected_btn.clicked.connect(self.edit_selected_teacher)
        utility_layout.addWidget(self.edit_selected_btn)
    
        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.setProperty("class", "secondary")
        self.refresh_btn.setIcon(QIcon("static/icons/refresh.png"))
        self.refresh_btn.setIconSize(QSize(20, 20))
        self.refresh_btn.clicked.connect(self.refresh_data)
        utility_layout.addWidget(self.refresh_btn)
    
        actions_layout.addLayout(utility_layout)
    
        parent_layout.addWidget(actions_group)
        parent_layout.addStretch()  # Push everything to top

    
    def setup_teacher_data_tab(self):
        """Set up the teacher data tab"""
        layout = QVBoxLayout(self.teacher_data_tab)
    
        # === SEARCH SECTION ===
        search_group = QGroupBox("Search Teachers")
        search_layout = QHBoxLayout(search_group)
    
        # Label
        search_layout.addWidget(QLabel("Search:"))
    
        # Input field
        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText("Enter name, ID, or email...")
        self.search_entry.textChanged.connect(self.search_teachers)
        self.search_entry.setFixedWidth(500)
        search_layout.addWidget(self.search_entry)
    
        # Search button
        self.search_btn = QPushButton("Search")
        self.search_btn.setProperty("class", "primary")
        self.search_btn.setIcon(QIcon("static/icons/search.png"))
        self.search_btn.setIconSize(QSize(18, 18))
        self.search_btn.setFixedWidth(110)
        self.search_btn.clicked.connect(self.search_teachers)
        search_layout.addWidget(self.search_btn)
    
        # Clear button
        self.clear_search_btn = QPushButton("Clear")
        self.clear_search_btn.setProperty("class", "warning")
        self.clear_search_btn.setIcon(QIcon("static/icons/clear.png"))
        self.clear_search_btn.setIconSize(QSize(18, 18))
        self.clear_search_btn.setFixedWidth(100)
        self.clear_search_btn.clicked.connect(self.clear_search)
        search_layout.addWidget(self.clear_search_btn)
    
        # Add a little stretch at the end so buttons don’t crowd
        search_layout.addStretch()
    
        layout.addWidget(search_group)

        
        #Data actions
        #actions_group = QGroupBox("Data Actions")
        #actions_layout = QHBoxLayout(actions_group)
        
        #self.export_btn = QPushButton("Export Data")
        #self.export_btn.setProperty("class", "success")
        #self.export_btn.clicked.connect(self.export_teachers_data)
        #actions_layout.addWidget(self.export_btn)

        # Add Import button
        #self.import_btn = QPushButton("Import Data")
        #self.import_btn.setProperty("class", "info")
        #self.import_btn.clicked.connect(self.import_teachers_data)
        #actions_layout.addWidget(self.import_btn)
        
        #self.report_btn = QPushButton("Generate Report")
        #self.report_btn.setProperty("class", "success")
        #self.report_btn.clicked.connect(self.generate_teacher_report)
        #actions_layout.addWidget(self.report_btn)
        
        #self.word_form_btn = QPushButton("Generate Teacher Form")
        #self.word_form_btn.setProperty("class", "info")
        #self.word_form_btn.clicked.connect(self.generate_teacher_profile_pdf)
        #actions_layout.addWidget(self.word_form_btn)
        
        #layout.addWidget(actions_group)
        
        # Teachers table
        self.create_teachers_table(layout)

    def create_teachers_table(self, parent_layout):
        """Create the teachers table"""
        table_group = QGroupBox("Teachers List")
        table_layout = QVBoxLayout(table_group)
        
        # Create table
        self.teachers_table = QTableWidget()
        self.teachers_table.setAlternatingRowColors(True)
        self.teachers_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.teachers_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.teachers_table.setSortingEnabled(True)
        
        # Set headers
        self.teachers_table_headers = [
            "ID", "Teacher ID", "Full Name", "Email", "Phone", 
            "Subject", "Staff Type", "Emp. Status", "Active", "Position",
            "Date Joined", "Qualification", "Gender", "Curr. Address"
        ]
        self.teachers_table.setColumnCount(len(self.teachers_table_headers))
        self.teachers_table.setHorizontalHeaderLabels(self.teachers_table_headers)
        
        # Configure table
        header = self.teachers_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)
        
        # Connect selection signal
        self.teachers_table.itemSelectionChanged.connect(self.on_teacher_select)
        self.teachers_table.itemDoubleClicked.connect(self.edit_selected_teacher)
        
        table_layout.addWidget(self.teachers_table)
        parent_layout.addWidget(table_group)

    def setup_departments_tab(self):
        """Set up departments tab placeholder"""
        layout = QVBoxLayout(self.departments_tab)
        placeholder_label = QLabel("Departments functionality will be implemented here")
        placeholder_label.setAlignment(Qt.AlignCenter)
        placeholder_label.setStyleSheet("font-size: 18px; color: #7f8c8d; padding: 50px;")
        layout.addWidget(placeholder_label)

    def update_full_name(self):
        """Automatically update full name when first name or surname changes"""
        first_name = self.first_name_entry.text().strip()
        surname = self.surname_entry.text().strip()
        full_name = f"{first_name} {surname}".strip()
        self.full_name_entry.setText(full_name)

    def select_photo(self):
        """Select and display teacher photo"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Teacher Photo",
            "",
            "Image files (*.jpg *.jpeg *.png *.gif *.bmp)"
        )
        
        if file_path:
            try:
                # Load and resize image
                image = Image.open(file_path)
                image = image.resize((180, 180), Image.Resampling.LANCZOS)
                
                # Convert to QPixmap
                qimage = ImageQt.ImageQt(image)
                pixmap = QPixmap.fromImage(qimage)
                
                # Update label
                self.photo_label.setPixmap(pixmap)
                self.photo_label.setText("")
                
                # Store path for saving
                self.photo_path = file_path
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error loading image: {str(e)}")

    def remove_photo(self):
        """Remove selected photo"""
        self.reset_photo_display()
        self.photo_path = None
        print("Photo removed")

    def reset_photo_display(self):
        """Reset photo display to default state"""
        self.photo_label.clear()
        self.photo_label.setText("No Photo\nSelected")
        self.photo_label.setProperty("class", "photo")
        self.photo_label.setAlignment(Qt.AlignCenter)

    def load_photo(self, photo_path):
        """Load and display teacher photo"""
        try:
            if not photo_path or not os.path.exists(photo_path):
                self.reset_photo_display()
                return
                
            # Load and process image
            image = Image.open(photo_path)
            image = image.resize((180, 180), Image.Resampling.LANCZOS)
            
            # Convert to QPixmap
            qimage = ImageQt.ImageQt(image)
            pixmap = QPixmap.fromImage(qimage)
            
            # Update display
            self.photo_label.setPixmap(pixmap)
            self.photo_label.setText("")
            
        except Exception as e:
            print(f"Error loading photo: {e}")
            self.reset_photo_display()

    def save_photo(self, teacher_id):
        """Save photo to teachers directory"""
        if not self.photo_path:
            return None
            
        try:
            # Create filename
            filename = f"teacher_{teacher_id}.jpg"
            destination = os.path.join(self.photos_dir, filename)
            
            # Copy and resize image
            image = Image.open(self.photo_path)
            image = image.resize((300, 300), Image.Resampling.LANCZOS)
            image.save(destination, "JPEG", quality=85)
            
            return destination
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving photo: {str(e)}")
            return None

    def load_schools(self):
        """Load schools for dropdown"""
        try:
            self.cursor.execute("SELECT id, school_name FROM schools ORDER BY school_name")
            schools = self.cursor.fetchall()
            
            self.school_combo.clear()
            self.school_combo.addItem("", None)  # Empty option
            
            for school in schools:
                display_text = f"{school[1]} (ID: {school[0]})"
                self.school_combo.addItem(display_text, school[0])
                
        except Exception as e:
            print(f"Error loading schools: {e}")
            QMessageBox.critical(self, "Error", f"Error loading schools: {str(e)}")

    def get_school_id_from_selection(self):
        """Get school ID from combo box selection"""
        return self.school_combo.currentData()

    def validate_fields(self):
        """Enhanced validation with duplicate checking"""
        # Basic field validation
        if not self.teacher_id_entry.text().strip():
            QMessageBox.warning(self, "Validation Error", "Teacher ID Code is required!")
            return False

        if not self.first_name_entry.text().strip():
            QMessageBox.warning(self, "Validation Error", "First Name is required!")
            return False

        if not self.surname_entry.text().strip():
            QMessageBox.warning(self, "Validation Error", "Surname is required!")
            return False

        if not self.get_school_id_from_selection():
            QMessageBox.warning(self, "Validation Error", "School selection is required!")
            return False

        if not self.position_combo.currentText().strip():
            QMessageBox.warning(self, "Validation Error", "Position is required!")
            return False
        
        # Email format validation
        email = self.email_entry.text().strip()
        if email and "@" not in email:
            QMessageBox.warning(self, "Validation Error", "Please enter a valid email address!")
            return False
        
        # Check for duplicates
        teacher_id_code = self.teacher_id_entry.text().strip()
        national_id = self.national_id_entry.text().strip()
        
        duplicates = self.check_for_duplicates(
            teacher_id_code=teacher_id_code,
            email=email,
            national_id=national_id,
            exclude_id=self.current_teacher_id
        )
        
        if duplicates:
            duplicate_msg = "Duplicate records found:\n\n" + "\n".join(duplicates)
            QMessageBox.warning(self, "Duplicate Error", duplicate_msg)
            return False

        return True

    def check_for_duplicates(self, teacher_id_code=None, email=None, national_id=None, exclude_id=None):
        """Check for duplicate teachers based on key identifiers"""
        duplicates = []
        
        try:
            # Check Teacher ID Code
            if teacher_id_code:
                query = "SELECT id, full_name FROM teachers WHERE teacher_id_code = %s"
                params = [teacher_id_code.strip()]
                if exclude_id:
                    query += " AND id != %s"
                    params.append(exclude_id)
                    
                self.cursor.execute(query, params)
                result = self.cursor.fetchone()
                if result:
                    duplicates.append(f"Teacher ID '{teacher_id_code}' is already used by {result[1]}")
            
            # Check Email
            if email:
                query = "SELECT id, full_name FROM teachers WHERE email = %s"
                params = [email.strip().lower()]
                if exclude_id:
                    query += " AND id != %s"
                    params.append(exclude_id)
                    
                self.cursor.execute(query, params)
                result = self.cursor.fetchone()
                if result:
                    duplicates.append(f"Email '{email}' is already used by {result[1]}")
            
            # Check National ID
            if national_id:
                query = "SELECT id, full_name FROM teachers WHERE national_id_number = %s"
                params = [national_id.strip()]
                if exclude_id:
                    query += " AND id != %s"
                    params.append(exclude_id)
                    
                self.cursor.execute(query, params)
                result = self.cursor.fetchone()
                if result:
                    duplicates.append(f"National ID '{national_id}' is already used by {result[1]}")
                    
        except Exception as e:
            print(f"Error checking duplicates: {e}")
            
        return duplicates

    def add_teacher(self):
        """Add new teacher with full field-level audit logging"""
        if not has_permission(self.user_session, "create_teacher"):
            QMessageBox.warning(self, "Permission Denied", "You don't have permission to add teachers.")
            return
    
        if not self.validate_fields():
            return
    
        try:
            school_id = self.get_school_id_from_selection()
    
            # Get form data
            first_name = self.first_name_entry.text().strip()
            surname = self.surname_entry.text().strip()
            full_name = f"{first_name} {surname}".strip()
            teacher_id_code = self.teacher_id_entry.text().strip()
    
            # Update full name display
            self.full_name_entry.setText(full_name)
    
            # Build list of fields for audit log
            fields_created = []
            if teacher_id_code:
                fields_created.append(f"ID='{teacher_id_code}'")
            if first_name:
                fields_created.append(f"First Name='{first_name}'")
            if surname:
                fields_created.append(f"Surname='{surname}'")
            if full_name:
                fields_created.append(f"Full Name='{full_name}'")
            if self.email_entry.text().strip():
                fields_created.append(f"Email='{self.email_entry.text().strip()}'")
            if self.phone_contact_1_entry.text().strip():
                fields_created.append(f"Phone='{self.phone_contact_1_entry.text().strip()}'")
            if self.position_combo.currentText().strip():
                fields_created.append(f"Position='{self.position_combo.currentText().strip()}'")
            if self.staff_type_combo.currentText().strip():
                fields_created.append(f"Staff Type='{self.staff_type_combo.currentText().strip()}'")
            if self.subject_specialty_entry.text().strip():
                fields_created.append(f"Subject='{self.subject_specialty_entry.text().strip()}'")
            if self.department_combo.currentText().strip():
                fields_created.append(f"Department='{self.department_combo.currentText().strip()}'")
            if self.school_combo.currentText().strip():
                fields_created.append(f"School='{self.school_combo.currentText().strip()}'")
            if self.is_active_checkbox.isChecked():
                fields_created.append("Status='Active'")
    
            # Insert teacher record
            query = '''
                INSERT INTO teachers (
                    school_id, teacher_id_code, salutation, first_name, surname, full_name,
                    email, gender, phone_contact_1, day_phone, current_address, home_district,
                    subject_specialty, qualification, date_joined, emergency_contact_1,
                    emergency_contact_2, national_id_number, birth_date, bank_account_number,
                    next_of_kin, employment_status, is_active, staff_type, position, monthly_salary, department_id
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            '''
    
            try:
                monthly_salary = float(self.monthly_salary_entry.text().strip() or 0)
            except ValueError:
                monthly_salary = 0.0
    
            values = (
                school_id,
                teacher_id_code,
                self.salutation_combo.currentText(),
                first_name,
                surname,
                full_name,
                self.email_entry.text().strip(),
                self.gender_combo.currentText(),
                self.phone_contact_1_entry.text().strip(),
                self.day_phone_entry.text().strip(),
                self.current_address_entry.text().strip(),
                self.home_district_entry.text().strip(),
                self.subject_specialty_entry.text().strip(),
                self.qualification_entry.text().strip(),
                self.date_joined_edit.date().toString("yyyy-MM-dd"),
                self.emergency_contact_1_entry.text().strip(),
                self.emergency_contact_2_entry.text().strip(),
                self.national_id_entry.text().strip(),
                self.birth_date_edit.date().toString("yyyy-MM-dd"),
                self.bank_account_entry.text().strip(),
                self.next_of_kin_entry.text().strip(),
                self.employment_status_combo.currentText(),
                self.is_active_checkbox.isChecked(),
                self.staff_type_combo.currentText(),
                self.position_combo.currentText().strip().title(),
                monthly_salary,
                self.get_department_id_from_selection()
            )
    
            self.cursor.execute(query, values)
            teacher_id = self.cursor.lastrowid
    
            # Save photo if selected
            if self.photo_path:
                photo_path = self.save_photo(teacher_id)
                if photo_path:
                    self.cursor.execute("UPDATE teachers SET photo_path = %s WHERE id = %s",
                                      (photo_path, teacher_id))
    
            self.db_connection.commit()
    
            # ✅ Log detailed audit action
            description = f"Created teacher {full_name} (ID: {teacher_id_code}): " + ", ".join(fields_created)
            self.log_audit_action(
                action="CREATE",
                table_name="teachers",
                record_id=teacher_id,
                description=description
            )
    
            QMessageBox.information(self, "Success", "Teacher added successfully!")
            self.clear_fields()
            self.load_teachers()
    
        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Error adding teacher: {str(e)}")
            print(f"Full error: {traceback.format_exc()}")

    def get_department_id_from_selection(self):
        return self.department_combo.currentData()

    def update_teacher(self):
        """Update existing teacher with field-level audit logging"""
        if not self.current_teacher_id:
            QMessageBox.warning(self, "Error", "Please select a teacher to update!")
            return
    
        if not has_permission(self.user_session, "edit_teacher"):
            QMessageBox.warning(self, "Permission Denied", "You don't have permission to edit teachers.")
            return
    
        if not self.validate_fields():
            return
    
        try:
            school_id = self.get_school_id_from_selection()
    
            # --- Step 1: Get current (old) values ---
            self.cursor.execute("""
                SELECT school_id, teacher_id_code, salutation, first_name, surname, full_name,
                       email, gender, phone_contact_1, day_phone, current_address, home_district,
                       subject_specialty, qualification, date_joined, emergency_contact_1,
                       emergency_contact_2, national_id_number, birth_date, bank_account_number,
                       next_of_kin, employment_status, is_active, staff_type, position, monthly_salary, department_id
                FROM teachers WHERE id = %s
            """, (self.current_teacher_id,))
            old_data = self.cursor.fetchone()
            if not old_data:
                QMessageBox.warning(self, "Error", "Teacher not found.")
                return
    
            # Map old values
            old_fields = {
                'school_id': old_data[0],
                'teacher_id_code': old_data[1],
                'salutation': old_data[2],
                'first_name': old_data[3],
                'surname': old_data[4],
                'full_name': f"{old_data[3]} {old_data[4]}".strip(),
                'email': old_data[6],
                'gender': old_data[7],
                'phone_contact_1': old_data[8],
                'day_phone': old_data[9],
                'current_address': old_data[10],
                'home_district': old_data[11],
                'subject_specialty': old_data[12],
                'qualification': old_data[13],
                'date_joined': old_data[14].strftime('%Y-%m-%d') if old_data[14] else None,
                'emergency_contact_1': old_data[15],
                'emergency_contact_2': old_data[16],
                'national_id_number': old_data[17],
                'birth_date': old_data[18].strftime('%Y-%m-%d') if old_data[18] else None,
                'bank_account_number': old_data[19],
                'next_of_kin': old_data[20],
                'employment_status': old_data[21],
                'is_active': bool(old_data[22]),
                'staff_type': old_data[23],
                'position': old_data[24],
                'monthly_salary': float(old_data[25]) if old_data[25] else 0.0,
                'department_id': old_data[26]
            }
    
            # --- Step 2: Get new values ---
            first_name = self.first_name_entry.text().strip()
            surname = self.surname_entry.text().strip()
            full_name = f"{first_name} {surname}".strip()
    
            try:
                monthly_salary = float(self.monthly_salary_entry.text().strip() or 0)
            except ValueError:
                monthly_salary = 0.0
    
            new_fields = {
                'school_id': school_id,
                'teacher_id_code': self.teacher_id_entry.text().strip(),
                'salutation': self.salutation_combo.currentText(),
                'first_name': first_name,
                'surname': surname,
                'full_name': full_name,
                'email': self.email_entry.text().strip(),
                'gender': self.gender_combo.currentText(),
                'phone_contact_1': self.phone_contact_1_entry.text().strip(),
                'day_phone': self.day_phone_entry.text().strip(),
                'current_address': self.current_address_entry.text().strip(),
                'home_district': self.home_district_entry.text().strip(),
                'subject_specialty': self.subject_specialty_entry.text().strip(),
                'qualification': self.qualification_entry.text().strip(),
                'date_joined': self.date_joined_edit.date().toString("yyyy-MM-dd"),
                'emergency_contact_1': self.emergency_contact_1_entry.text().strip(),
                'emergency_contact_2': self.emergency_contact_2_entry.text().strip(),
                'national_id_number': self.national_id_entry.text().strip(),
                'birth_date': self.birth_date_edit.date().toString("yyyy-MM-dd"),
                'bank_account_number': self.bank_account_entry.text().strip(),
                'next_of_kin': self.next_of_kin_entry.text().strip(),
                'employment_status': self.employment_status_combo.currentText(),
                'is_active': self.is_active_checkbox.isChecked(),
                'staff_type': self.staff_type_combo.currentText(),
                'position': self.position_combo.currentText().strip().title(),
                'monthly_salary': monthly_salary,
                'department_id': self.get_department_id_from_selection()
            }
    
            # --- Step 3: Compare and build changes ---
            changes = []
            field_labels = {
                'teacher_id_code': 'Teacher ID',
                'salutation': 'Salutation',
                'first_name': 'First Name',
                'surname': 'Surname',
                'email': 'Email',
                'gender': 'Gender',
                'phone_contact_1': 'Phone',
                'current_address': 'Address',
                'subject_specialty': 'Subject',
                'qualification': 'Qualification',
                'date_joined': 'Join Date',
                'national_id_number': 'National ID',
                'employment_status': 'Employment Status',
                'is_active': 'Active',
                'staff_type': 'Staff Type',
                'position': 'Position',
                'monthly_salary': 'Monthly Salary',
                'department_id': 'Department'
            }
    
            for field in old_fields:
                old_val = old_fields[field]
                new_val = new_fields[field]
                if old_val != new_val:
                    label = field_labels.get(field, field.replace('_', ' ').title())
                    changes.append(f"{label} changed from '{old_val}' to '{new_val}'")
    
            if not changes:
                QMessageBox.information(self, "No Changes", "No data was changed.")
                return
    
            # --- Step 4: Perform update ---
            query = '''
                UPDATE teachers SET
                    school_id = %s, teacher_id_code = %s, salutation = %s, first_name = %s,
                    surname = %s, full_name = %s, email = %s, gender = %s, phone_contact_1 = %s,
                    day_phone = %s, current_address = %s, home_district = %s, subject_specialty = %s,
                    qualification = %s, date_joined = %s, emergency_contact_1 = %s,
                    emergency_contact_2 = %s, national_id_number = %s, birth_date = %s,
                    bank_account_number = %s, next_of_kin = %s, employment_status = %s,
                    is_active = %s, staff_type = %s, position = %s, monthly_salary = %s, department_id = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            '''
    
            values = tuple(new_fields.values()) + (self.current_teacher_id,)
            self.cursor.execute(query, values)
    
            # Handle photo update
            if self.photo_path:
                photo_path = self.save_photo(self.current_teacher_id)
                if photo_path:
                    self.cursor.execute("UPDATE teachers SET photo_path = %s WHERE id = %s",
                                      (photo_path, self.current_teacher_id))
    
            self.db_connection.commit()
    
            # --- Step 5: Log detailed audit ---
            change_desc = "; ".join(changes)
            self.log_audit_action(
                action="UPDATE",
                table_name="teachers",
                record_id=self.current_teacher_id,
                description=f"Updated teacher {full_name} (ID: {new_fields['teacher_id_code']}): {change_desc}"
            )
    
            QMessageBox.information(self, "Success", "Teacher updated successfully!")
            self.clear_fields()
            self.load_teachers()
    
        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Error updating teacher: {str(e)}")
            print(f"Full error: {traceback.format_exc()}")
        
    def delete_teacher(self):
        """Delete selected teacher with detailed audit logging"""
        if not self.current_teacher_id:
            QMessageBox.warning(self, "Error", "Please select a teacher to delete!")
            return
    
        if not has_permission(self.user_session, "delete_teacher"):
            QMessageBox.warning(self, "Permission Denied", "You don't have permission to delete teachers.")
            return
    
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            "Are you sure you want to delete this teacher? This action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
    
        if reply == QMessageBox.Yes:
            try:
                # Get teacher info before deletion
                self.cursor.execute("""
                    SELECT full_name, teacher_id_code, is_active, staff_type, position, photo_path 
                    FROM teachers WHERE id = %s
                """, (self.current_teacher_id,))
                result = self.cursor.fetchone()
                if not result:
                    QMessageBox.warning(self, "Error", "Teacher not found.")
                    return
    
                full_name = result[0] or "Unknown"
                teacher_id_code = result[1] or "N/A"
                is_active = bool(result[2])
                staff_type = result[3] or "Unknown"
                position = result[4] or "Unknown"
                photo_path = result[5]
    
                # Delete teacher record
                self.cursor.execute("DELETE FROM teachers WHERE id = %s", (self.current_teacher_id,))
                self.db_connection.commit()
    
                # ✅ Log detailed audit BEFORE deleting file
                photo_status = "with photo removed" if photo_path and os.path.exists(photo_path) else "no photo"
                description = (
                    f"Deleted teacher {full_name} (ID: {teacher_id_code}), "
                    f"was {'active' if is_active else 'inactive'}, "
                    f"employed as {position} ({staff_type}), {photo_status}"
                )
    
                self.log_audit_action(
                    action="DELETE",
                    table_name="teachers",
                    record_id=self.current_teacher_id,
                    description=description
                )
    
                # Delete photo file if exists
                if photo_path and os.path.exists(photo_path):
                    try:
                        os.remove(photo_path)
                    except Exception as e:
                        print(f"Warning: Could not delete photo file {photo_path}: {e}")
    
                QMessageBox.information(self, "Success", "Teacher deleted successfully!")
                self.clear_fields()
                self.load_teachers()
    
            except Exception as e:
                self.db_connection.rollback()
                QMessageBox.critical(self, "Error", f"Error deleting teacher: {str(e)}")
                print(f"Full error: {traceback.format_exc()}")

    def load_departments_combo(self):
        """Load departments into department combo box"""
        try:
            self.cursor.execute("""
                SELECT id, department_name 
                FROM departments 
                WHERE is_active = 1 
                ORDER BY department_name
            """)
            departments = self.cursor.fetchall()
            self.department_combo.clear()
            self.department_combo.addItem("", None)  # No department
            for dept in departments:
                self.department_combo.addItem(dept[1], dept[0])
        except Exception as e:
            print(f"Error loading departments: {e}")
        
    def clear_fields(self):
        """Clear all form fields and selection"""
        self.current_teacher_id = None
        self.photo_path = None
        
        # Clear all entry fields
        self.teacher_id_entry.clear()
        self.first_name_entry.clear()
        self.surname_entry.clear()
        self.full_name_entry.clear()
        self.email_entry.clear()
        self.phone_contact_1_entry.clear()
        self.day_phone_entry.clear()
        self.current_address_entry.clear()
        self.home_district_entry.clear()
        self.subject_specialty_entry.clear()
        self.qualification_entry.clear()
        self.emergency_contact_1_entry.clear()
        self.emergency_contact_2_entry.clear()
        self.national_id_entry.clear()
        self.bank_account_entry.clear()
        self.next_of_kin_entry.clear()
        self.monthly_salary_entry.clear()
        
        # Reset combo boxes
        self.salutation_combo.setCurrentText("Mr.")
        self.gender_combo.setCurrentText("Male")
        self.school_combo.setCurrentIndex(0)
        self.employment_status_combo.setCurrentText("Full-time")
        self.staff_type_combo.setCurrentText("Teaching")
        self.position_combo.setCurrentText("")
        self.department_combo.setCurrentIndex(0)
        
        # Reset dates
        self.birth_date_edit.setDate(QDate.currentDate())
        self.date_joined_edit.setDate(QDate.currentDate())
        
        # Reset checkbox
        self.is_active_checkbox.setChecked(True)
        
        # Clear photo
        self.reset_photo_display()
        
        print("All fields cleared and selection reset")

    def apply_button_permissions(self):
        """Enable/disable buttons based on user permissions"""
        can_create = has_permission(self.user_session, "create_teacher")
        can_edit = has_permission(self.user_session, "edit_teacher")
        can_delete = has_permission(self.user_session, "delete_teacher")
    
        self.add_btn.setEnabled(can_create)
        self.update_btn.setEnabled(can_edit)
        self.delete_btn.setEnabled(can_delete)
        self.edit_selected_btn.setEnabled(can_edit)
    
        # Optional: Always enabled
        self.clear_btn.setEnabled(True)
        self.refresh_btn.setEnabled(True)
    
        # Set tooltips for better UX
        self.add_btn.setToolTip("Add a new teacher" if can_create else "Permission required: create_teacher")
        self.update_btn.setToolTip("Update selected teacher" if can_edit else "Permission required: edit_teacher")
        self.delete_btn.setToolTip("Delete selected teacher" if can_delete else "Permission required: delete_teacher")
        self.edit_selected_btn.setToolTip("Edit selected teacher" if can_edit else "Permission required: edit_teacher")

    def load_teachers(self):
        """Load teachers data into the table"""
        try:
            # Update query to include the new fields
            query = '''
                SELECT
                    t.id, t.teacher_id_code, t.full_name, t.email, t.phone_contact_1,
                    t.subject_specialty, t.staff_type, t.employment_status, t.is_active,
                    t.position, t.date_joined, t.qualification, t.gender, t.current_address
                FROM teachers t
                ORDER BY t.full_name
            '''
            self.cursor.execute(query)
            teachers = self.cursor.fetchall()
    
            # Set table dimensions
            self.teachers_table.setRowCount(len(teachers))
            self.teachers_table.setColumnCount(len(self.teachers_table_headers))
            
            # Populate table
            for row_idx, teacher in enumerate(teachers):
                items = [
                    str(teacher[0]),                              # id
                    teacher[1] or "",                             # teacher_id_code
                    teacher[2] or "",                             # full_name
                    teacher[3] or "",                             # email
                    teacher[4] or "",                             # phone_contact_1
                    teacher[5] or "",                             # subject_specialty
                    teacher[6] or "",                             # staff_type
                    teacher[7] or "",                             # employment_status
                    "Yes" if teacher[8] else "No",                # is_active
                    teacher[9] or "",                             # position
                    str(teacher[10] or ""),                       # date_joined
                    teacher[11] or "",                            # qualification
                    teacher[12] or "",                            # gender
                    teacher[13] or ""                             # current_address
                ]
                
                for col_idx, item in enumerate(items):
                    table_item = QTableWidgetItem(str(item))
                    table_item.setData(Qt.UserRole, teacher[0])  # Store teacher ID
                    
                    #Right-align current address column
                    if col_idx == 13:  # current address column
                        table_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    
                    self.teachers_table.setItem(row_idx, col_idx, table_item)
    
            # Configure column sizing
            self.configure_table_columns()
            
        except Exception as e:
            print(f"Error loading teachers: {e}")
            QMessageBox.critical(self, "Error", f"Error loading teachers: {str(e)}")

    def configure_table_columns(self):
        """Configure table column sizing policies"""
        header = self.teachers_table.horizontalHeader()
        
        # Set resize modes for different columns
        # Fixed size columns
        fixed_columns = [0, 1, 7, 8, 10, 12]  # ID, Teacher ID, Emp. Status, Active, Date Joined, Gender
        for col in fixed_columns:
            header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        
        # Stretchable columns
        stretch_columns = [2, 3, 5, 9, 11, 12]  # Full Name, Email, Subject, Position, Qualification
        for col in stretch_columns:
            header.setSectionResizeMode(col, QHeaderView.Stretch)
        
        # Interactive columns (user can resize)
        interactive_columns = [4, 6, 13]  # Phone, Staff Type, current address
        for col in interactive_columns:
            header.setSectionResizeMode(col, QHeaderView.Interactive)
            # Set initial width
            if col == 13:  # current address column
                header.resizeSection(col, 100)
            else:
                header.resizeSection(col, 120)

    def search_teachers(self):
        """Search teachers based on search entry"""
        search_term = self.search_entry.text().strip()
        
        try:
            if not search_term:
                self.load_teachers()
                return
    
            # Update query to include new fields in search
            query = '''
                SELECT  
                    t.id, t.teacher_id_code, t.full_name, t.email, t.phone_contact_1,
                    t.subject_specialty, t.staff_type, t.employment_status, t.is_active,
                    t.position, t.date_joined, t.qualification, t.gender, t.current_address
                FROM teachers t
                WHERE  
                    LOWER(t.full_name) LIKE LOWER(%s) OR  
                    LOWER(t.first_name) LIKE LOWER(%s) OR  
                    LOWER(t.surname) LIKE LOWER(%s) OR
                    LOWER(t.teacher_id_code) LIKE LOWER(%s) OR
                    LOWER(t.email) LIKE LOWER(%s) OR
                    LOWER(t.subject_specialty) LIKE LOWER(%s) OR
                    LOWER(t.staff_type) LIKE LOWER(%s) OR
                    LOWER(t.employment_status) LIKE LOWER(%s) OR
                    LOWER(t.position) LIKE LOWER(%s) OR
                    LOWER(t.qualification) LIKE LOWER(%s) OR
                    LOWER(t.gender) LIKE LOWER(%s) OR
                    LOWER(t.current_address) LIKE LOWER(%s)
                ORDER BY t.full_name
            '''
    
            search_pattern = f"%{search_term}%"
            params = tuple([search_pattern] * 12)  # Increased to 12 for the new search fields
            self.cursor.execute(query, params)
            teachers = self.cursor.fetchall()
    
            # Update table with search results
            self.teachers_table.setRowCount(len(teachers))
            
            for row_idx, teacher in enumerate(teachers):
                items = [
                    str(teacher[0]),                              # id
                    teacher[1] or "",                             # teacher_id_code
                    teacher[2] or "",                             # full_name
                    teacher[3] or "",                             # email
                    teacher[4] or "",                             # phone_contact_1
                    teacher[5] or "",                             # subject_specialty
                    teacher[6] or "",                             # staff_type
                    teacher[7] or "",                             # employment_status
                    "Yes" if teacher[8] else "No",                # is_active
                    teacher[9] or "",                             # position
                    str(teacher[10] or ""),                       # date_joined
                    teacher[11] or "",                            # qualification
                    teacher[12] or "",                            # gender
                    teacher[13] or ""                             # current_address
                    #f"{float(teacher[14] or 0):,.2f}" if teacher[14] else "0.00"  # monthly_salary
                ]
                
                for col_idx, item in enumerate(items):
                    table_item = QTableWidgetItem(str(item))
                    table_item.setData(Qt.UserRole, teacher[0])
                    
                    # Right-align salary column
                    if col_idx == 13:  # current address column
                        table_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    
                    self.teachers_table.setItem(row_idx, col_idx, table_item)
    
            self.configure_table_columns()
    
        except Exception as e:
            print(f"Error searching teachers: {e}")
            QMessageBox.critical(self, "Error", f"Error searching teachers: {str(e)}")
        
    def clear_search(self):
        """Clear search and reload all teachers"""
        self.search_entry.clear()
        self.load_teachers()

    def on_teacher_select(self):
        """Handle teacher selection from table"""
        try:
            selected_items = self.teachers_table.selectedItems()
            if not selected_items:
                return
                
            # Get teacher ID from the first selected item (first column)
            row = selected_items[0].row()
            teacher_id_item = self.teachers_table.item(row, 0)
            
            if teacher_id_item:
                teacher_id = int(teacher_id_item.text())
                self.current_teacher_id = teacher_id
                self.load_teacher_data(teacher_id)
                
                # Show selection message
                teacher_name = self.teachers_table.item(row, 2).text()
                QMessageBox.information(
                    self, 
                    "Teacher Selected",
                    f"Selected: {teacher_name}\n\nYou can now Update or Delete this teacher."
                )
                
        except Exception as e:
            print(f"Error in teacher selection: {e}")
            QMessageBox.critical(self, "Error", f"Error selecting teacher: {str(e)}")

    def load_teacher_data(self, teacher_id):
        """Load teacher data into form fields"""
        try:
            # Get teacher data with school info
            query = '''
                SELECT 
                    t.id, t.school_id, t.teacher_id_code, t.salutation, t.first_name, 
                    t.surname, t.full_name, t.email, t.gender, t.phone_contact_1, 
                    t.day_phone, t.current_address, t.home_district, t.subject_specialty, 
                    t.qualification, t.date_joined, t.emergency_contact_1, t.emergency_contact_2, 
                    t.national_id_number, t.birth_date, t.bank_account_number, t.next_of_kin, 
                    t.photo_path, t.employment_status, t.is_active, t.staff_type, t.position, 
                    t.monthly_salary, t.department_id, s.school_name
                FROM teachers t
                LEFT JOIN schools s ON t.school_id = s.id
                WHERE t.id = %s
            '''
            self.cursor.execute(query, (teacher_id,))
            teacher = self.cursor.fetchone()
            
            if not teacher:
                QMessageBox.warning(self, "Error", "Teacher data not found")
                return
    
            # Clear all fields first
            self.clear_fields()
            
            # Set current teacher ID
            self.current_teacher_id = teacher_id
            
            # Populate form fields using tuple indices
            self.teacher_id_entry.setText(teacher[2] or "")        # teacher_id_code
            self.salutation_combo.setCurrentText(teacher[3] or "Mr.")  # salutation
            self.first_name_entry.setText(teacher[4] or "")       # first_name
            self.surname_entry.setText(teacher[5] or "")          # surname
            
            # Auto-generate full name
            self.update_full_name()
            
            self.email_entry.setText(teacher[7] or "")            # email
            self.gender_combo.setCurrentText(teacher[8] or "Male")  # gender
            self.phone_contact_1_entry.setText(teacher[9] or "")   # phone_contact_1
            self.day_phone_entry.setText(teacher[10] or "")       # day_phone
            self.current_address_entry.setText(teacher[11] or "") # current_address
            self.home_district_entry.setText(teacher[12] or "")   # home_district
            self.subject_specialty_entry.setText(teacher[13] or "") # subject_specialty
            self.qualification_entry.setText(teacher[14] or "")   # qualification
            
            # Handle dates
            if teacher[15]:  # date_joined
                try:
                    date_joined = QDate.fromString(str(teacher[15]), "yyyy-MM-dd")
                    self.date_joined_edit.setDate(date_joined)
                except:
                    self.date_joined_edit.setDate(QDate.currentDate())
            
            self.emergency_contact_1_entry.setText(teacher[16] or "") # emergency_contact_1
            self.emergency_contact_2_entry.setText(teacher[17] or "") # emergency_contact_2
            self.national_id_entry.setText(teacher[18] or "")      # national_id_number
            
            # Birth date
            if teacher[19]:  # birth_date
                try:
                    birth_date = QDate.fromString(str(teacher[19]), "yyyy-MM-dd")
                    self.birth_date_edit.setDate(birth_date)
                except:
                    self.birth_date_edit.setDate(QDate.currentDate())
            
            self.bank_account_entry.setText(teacher[20] or "")     # bank_account_number
            self.next_of_kin_entry.setText(teacher[21] or "")      # next_of_kin
            
            # Handle photo
            photo_path = teacher[22]  # photo_path
            if photo_path and os.path.exists(photo_path):
                self.load_photo(photo_path)
            else:
                self.reset_photo_display()
            
            self.employment_status_combo.setCurrentText(teacher[23] or "Full-time")  # employment_status
            self.is_active_checkbox.setChecked(bool(teacher[24]))   # is_active
            self.staff_type_combo.setCurrentText(teacher[25] or "Teaching")  # staff_type
            self.position_combo.setCurrentText(teacher[26] or "")   # position
            self.monthly_salary_entry.setText(str(teacher[27] or 0))  # monthly_salary
            
            # Handle department selection
            dept_id = teacher[28]  # department_id
            if dept_id:
                # Find the department in combo box by data
                for i in range(self.department_combo.count()):
                    if self.department_combo.itemData(i) == dept_id:
                        self.department_combo.setCurrentIndex(i)
                        break
            else:
                self.department_combo.setCurrentIndex(0)  # Select empty option
            
            # Handle school selection
            if teacher[1]:  # school_id exists
                for i in range(self.school_combo.count()):
                    if self.school_combo.itemData(i) == teacher[1]:
                        self.school_combo.setCurrentIndex(i)
                        break
            
            print(f"Teacher data loaded successfully for ID: {teacher_id}")
            
        except Exception as e:
            print(f"Error loading teacher data: {e}")
            QMessageBox.critical(self, "Error", f"Error loading teacher data: {str(e)}")
            import traceback
            print(f"Full traceback: {traceback.format_exc()}")

    def edit_selected_teacher(self):
        """Edit selected teacher from table"""
        if not self.current_teacher_id:
            QMessageBox.warning(self, "Error", "Please select a teacher from the table first!")
            return
        
        # Switch to teacher form tab
        self.tab_widget.setCurrentIndex(0)  # Index 0 is the form tab
        QMessageBox.information(self, "Info", "Teacher data loaded in the form. Make changes and click Update.")

    def refresh_data(self):
        """Refresh all data in the form with progress indication"""
        try:
            # Ensure connection first
            self._ensure_connection()
            
            self.db_connection.commit()
            self.status_label.setText("Refreshing...")
            self.status_label.setStyleSheet(f"color: {self.colors['info']}; font-weight: bold;")
            QApplication.processEvents()
    
            self.current_teacher_id = None
            self.load_teachers()
            self.load_schools()
            self.load_departments_combo()
            self.clear_fields()
            
            # Refresh analytics if the tab is active
            if self.tab_widget.currentIndex() == 2:  # Analytics tab index
                self.refresh_staff_analytics()
    
            # Reapply button permissions after refresh
            self.apply_button_permissions()
    
            self.status_label.setText("Data Refreshed")
            self.status_label.setStyleSheet(f"color: {self.colors['success']}; font-weight: bold;")
        except Error as e:
            QMessageBox.critical(self, "Error", f"Failed to refresh data: {e}")
            self.status_label.setText("Refresh Failed")
            self.status_label.setStyleSheet(f"color: {self.colors['danger']}; font-weight: bold;")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Unexpected error during refresh: {e}")
            self.status_label.setText("Refresh Failed")
            self.status_label.setStyleSheet(f"color: {self.colors['danger']}; font-weight: bold;")
        
    def export_teachers_data(self):
        """Export teachers data using shared export_with_green_header method"""
        try:
            # Fetch data from database
            self.cursor.execute('''
                SELECT
                    t.teacher_id_code, t.salutation, t.first_name, t.surname, t.full_name,
                    t.email, t.gender, t.phone_contact_1, t.day_phone, t.current_address,
                    t.home_district, t.subject_specialty, t.qualification, t.date_joined,
                    t.emergency_contact_1, t.emergency_contact_2, t.national_id_number,
                    t.birth_date, t.bank_account_number, t.next_of_kin, t.employment_status,
                    CASE WHEN t.is_active = 1 THEN 'Yes' ELSE 'No' END,
                    t.staff_type, t.position, t.monthly_salary, s.school_name
                FROM teachers t
                LEFT JOIN schools s ON t.school_id = s.id
                ORDER BY t.full_name
            ''')
            teachers = self.cursor.fetchall()
    
            if not teachers:
                QMessageBox.information(self, "No Data", "No teacher data found to export.")
                return
    
            # Define headers (must match SELECT order)
            headers = [
                "Teacher ID", "Salutation", "First Name", "Surname", "Full Name",
                "Email", "Gender", "Phone Contact 1", "Day Phone", "Current Address",
                "Home District", "Subject Specialty", "Qualification", "Date Joined",
                "Emergency Contact 1", "Emergency Contact 2", "National ID",
                "Birth Date", "Bank Account", "Next of Kin", "Employment Status",
                "Active Status", "Staff Type", "Position", "Monthly Salary", "School Name"
            ]
    
            # Get school name for title
            school_info = self.get_school_info()
            title = f"{school_info['name']} - TEACHERS DATA"
    
            # Use shared export method
            self.export_with_green_header(
                data=teachers,
                headers=headers,
                filename_prefix="teachers_export",
                title=title
            )
    
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export teacher data:\n{str(e)}")
    
    def generate_teacher_report(self):
        """Generate a teacher summary report"""
        try:
            from datetime import datetime
    
            # Get statistics including staff_type breakdown
            self.cursor.execute("SELECT COUNT(*) FROM teachers WHERE is_active = 1")
            active_count = self.cursor.fetchone()[0]
            
            self.cursor.execute("SELECT COUNT(*) FROM teachers WHERE is_active = 0")
            inactive_count = self.cursor.fetchone()[0]
            
            # Employment status breakdown
            self.cursor.execute('''
                SELECT employment_status, COUNT(*)
                FROM teachers
                GROUP BY employment_status
            ''')
            employment_stats = self.cursor.fetchall()
            
            # Staff type breakdown
            self.cursor.execute('''
                SELECT staff_type, COUNT(*)
                FROM teachers
                GROUP BY staff_type
            ''')
            staff_type_stats = self.cursor.fetchall()
            
            # School distribution
            self.cursor.execute('''
                SELECT s.school_name, COUNT(t.id) as teacher_count
                FROM schools s
                LEFT JOIN teachers t ON s.id = t.school_id
                GROUP BY s.id, s.school_name
                ORDER BY teacher_count DESC
            ''')
            school_stats = self.cursor.fetchall()
    
            # Generate report with staff type information
            report = f"""
    TEACHERS SUMMARY REPORT
    Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    
    OVERVIEW:
    - Active Teachers: {active_count}
    - Inactive Teachers: {inactive_count}
    - Total Teachers: {active_count + inactive_count}
    
    STAFF TYPE BREAKDOWN:
    """
            for staff_type, count in staff_type_stats:
                report += f"- {staff_type or 'Unknown'}: {count}\n"
    
            report += "\nEMPLOYMENT STATUS BREAKDOWN:\n"
            for status, count in employment_stats:
                report += f"- {status or 'Unknown'}: {count}\n"
    
            report += "\nSCHOOL DISTRIBUTION:\n"
            for school, count in school_stats:
                report += f"- {school or 'Unknown'}: {count}\n"
    
            # Show report in a new window
            report_window = QDialog(self)
            report_window.setWindowTitle("Teachers Report")
            report_window.setMinimumSize(600, 500)
            
            layout = QVBoxLayout(report_window)
            
            text_edit = QTextEdit()
            text_edit.setReadOnly(True)
            text_edit.setFont(QFont("Courier", 10))
            text_edit.setText(report)
            
            close_button = QPushButton("Close")
            close_button.clicked.connect(report_window.close)
            
            layout.addWidget(text_edit)
            layout.addWidget(close_button)
            
            report_window.exec()
    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error generating report: {str(e)}")


    # Updated method for TeachersForm to generate PDF bytes correctly from teachers form
    def generate_teacher_profile_pdf(self, teacher_id=None):
        """Generate teacher profile PDF and return proper PDF bytes"""
        try:
            # Ensure DB connection first
            self._ensure_connection()
            
            import os
            import io  # Add this import that was missing
            from fpdf import FPDF
            from datetime import datetime
            import tempfile
    
            # Validate cursor
            if not hasattr(self, 'cursor') or not self.cursor:
                raise Exception("Database cursor not available")
    
            # Use provided teacher_id or current selection
            target_teacher_id = teacher_id or getattr(self, 'current_teacher_id', None)
            if not target_teacher_id:
                raise ValueError("No teacher selected")
    
            # Fetch teacher data with photo
            query = '''
                SELECT 
                    t.teacher_id_code, t.salutation, t.first_name, t.surname, t.email, 
                    t.gender, t.phone_contact_1, t.day_phone, t.current_address, 
                    t.home_district, t.subject_specialty, t.qualification, t.date_joined,
                    t.emergency_contact_1, t.emergency_contact_2, t.national_id_number,
                    t.birth_date, t.bank_account_number, t.next_of_kin, t.employment_status,
                    t.is_active, t.staff_type, t.position, t.school_id, t.photo_path
                FROM teachers t
                WHERE t.id = %s
            '''
            self.cursor.execute(query, (target_teacher_id,))
            teacher = self.cursor.fetchone()
            if not teacher:
                raise ValueError("Teacher data not found")
    
            # Fetch school info - CORRECT QUERY for your table structure
            school_id = teacher[23]  # school_id position
            school = None
            
            if school_id:
                try:
                    school_query = '''
                        SELECT 
                            school_name, address, phone, email, logo_path
                        FROM schools 
                        WHERE id = %s
                    '''
                    self.cursor.execute(school_query, (school_id,))
                    school = self.cursor.fetchone()
                except Exception as school_error:
                    print(f"School query failed: {school_error}")
                    school = None
    
            # Extract school data with defaults
            school_name = school[0] if school else "CBCentra School"
            school_address = school[1] if school else "P.O. Box 12345"
            school_phone = school[2] if school else "Tel: +254 700 000000"
            school_email = school[3] if school else "info@cbcentra.edu"
            default_logo = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                "static", "images", "logo.png"
            )
            school_logo = school[4] if school and school[4] else default_logo
    
            class TeacherPDF(FPDF):
                def __init__(self, teacher_photo=None):
                    super().__init__(orientation='P', unit='mm', format='A4')
                    self.set_margins(15, 15, 15)
                    self.set_auto_page_break(auto=False)  # Manual page breaks for better control
                    self.teacher_photo = teacher_photo
    
                def header(self):
                    # School logo (left side)
                    if os.path.exists(school_logo):
                        try:
                            self.image(school_logo, 15, 10, 25)
                        except:
                            pass  # Skip if logo can't be loaded
                    
                    # Teacher photo (right side of header)
                    if self.teacher_photo and os.path.exists(self.teacher_photo):
                        try:
                            self.image(self.teacher_photo, 165, 5, 30, 30)  # Positioned in header
                        except:
                            pass  # Skip if photo can't be loaded
                    
                    # School information (centered) - only show if available
                    self.set_y(10)
                    if school_name:
                        self.set_font("Arial", "B", 16)
                        self.cell(0, 8, school_name, 0, 1, "C")
                    
                    if school_address:
                        self.set_font("Arial", "", 10)
                        self.cell(0, 5, school_address, 0, 1, "C")
                    
                    if school_phone or school_email:
                        contact_info = ""
                        if school_phone:
                            contact_info += school_phone
                        if school_phone and school_email:
                            contact_info += " | "
                        if school_email:
                            contact_info += school_email
                        
                        self.set_font("Arial", "", 10)
                        self.cell(0, 5, contact_info, 0, 1, "C")
                    
                    # Report title
                    self.ln(3)
                    self.set_font("Arial", "B", 14)
                    self.set_text_color(70, 70, 70)  # Dark gray
                    self.cell(0, 8, "TEACHER PROFILE REPORT", 0, 1, "C")
                    
                    # Generation info
                    self.set_font("Arial", "I", 8)
                    self.set_text_color(100, 100, 100)
                    gen_info = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                    if hasattr(self, 'user_session') and self.user_session:
                        user_name = getattr(self.user_session, 'full_name', 'System')
                        gen_info += f" | By: {user_name}"
                    self.cell(0, 4, gen_info, 0, 1, "C")
                    
                    # Line separator
                    self.line(15, self.get_y() + 2, 195, self.get_y() + 2)
                    self.ln(6)
    
                def footer(self):
                    self.set_y(-15)
                    self.set_font("Arial", "I", 8)
                    self.set_text_color(128, 128, 128)
                    self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "C")
    
                def section_header(self, title, highlight=False):
                    self.ln(3)  # Space before section
                    self.set_font("Arial", "B", 11)
                    if highlight:
                        self.set_fill_color(173, 216, 230)  # Light blue highlight
                        self.set_text_color(0, 0, 0)      # Black text
                    else:
                        self.set_fill_color(230, 230, 230)  # Light gray background
                        self.set_text_color(70, 70, 70)     # Dark gray text
                    self.cell(0, 7, title, 0, 1, "L", True)
                    self.ln(3)  # Space after section header
    
                def add_field(self, label, value, width1=50, width2=0):
                    self.set_font("Arial", "B", 10)
                    self.cell(width1, 6, label, 0, 0)
                    self.set_font("Arial", "", 10)
                    if width2 == 0:
                        width2 = 195 - 15 - width1  # Calculate remaining width
                    self.cell(width2, 6, str(value) if value else "N/A", 0, 1)
    
                def add_multiline_field(self, label, value):
                    self.set_font("Arial", "B", 10)
                    self.cell(0, 6, label, 0, 1)
                    self.set_font("Arial", "", 10)
                    text = str(value) if value else "N/A"
                    self.multi_cell(0, 6, text)
                    self.ln(2)
    
            # Create PDF instance with teacher photo
            photo_path = teacher[24]  # photo_path from query (last field)
            pdf = TeacherPDF(photo_path)
            pdf.add_page()
    
            # Personal Information Section
            pdf.section_header("PERSONAL INFORMATION")
            
            pdf.add_field("Teacher ID Code:", teacher[0])
            full_name = f"{teacher[1] or ''} {teacher[2] or ''} {teacher[3] or ''}".strip()
            pdf.add_field("Full Name:", full_name)
            
            # Two-column layout for compact display
            current_y = pdf.get_y()
            pdf.add_field("Gender:", teacher[5], 50, 45)
            pdf.set_y(current_y)
            pdf.set_x(110)
            pdf.add_field("Birth Date:", teacher[16].strftime("%Y-%m-%d") if teacher[16] else "N/A", 40, 45)
            
            pdf.add_field("National ID Number:", teacher[15])
            pdf.add_field("Email:", teacher[4])
            
            current_y = pdf.get_y()
            pdf.add_field("Phone Contact 1:", teacher[6], 50, 45)
            pdf.set_y(current_y)
            pdf.set_x(110)
            pdf.add_field("Day Phone:", teacher[7], 40, 45)
    
            # Professional Information Section
            pdf.section_header("PROFESSIONAL INFORMATION")
            
            pdf.add_field("Subject Specialty:", teacher[10])
            pdf.add_field("Qualification:", teacher[11])
            pdf.add_field("Date Joined:", teacher[12].strftime("%Y-%m-%d") if teacher[12] else "N/A")
            
            current_y = pdf.get_y()
            pdf.add_field("Staff Type:", teacher[21], 50, 45)
            pdf.set_y(current_y)
            pdf.set_x(110)
            pdf.add_field("Position:", teacher[22], 40, 45)
            
            current_y = pdf.get_y()
            pdf.add_field("Employment Status:", teacher[19], 50, 45)
            pdf.set_y(current_y)
            pdf.set_x(110)
            pdf.add_field("Bank Account:", teacher[17], 40, 45)
    
            # School Information Section
            pdf.section_header("SCHOOL INFORMATION")
            
            pdf.add_field("School Name:", school_name)
            pdf.add_field("School Address:", school_address)
            
            current_y = pdf.get_y()
            pdf.add_field("School Phone:", school_phone, 50, 45)
            pdf.set_y(current_y)
            pdf.set_x(110)
            pdf.add_field("School Email:", school_email, 40, 45)
    
            # Address & Emergency Contact Section
            pdf.section_header("ADDRESS & EMERGENCY CONTACT")
            
            pdf.add_field("Current Address:", teacher[8])
            pdf.add_field("Home District:", teacher[9])
            pdf.add_field("Next of Kin:", teacher[18])
            
            current_y = pdf.get_y()
            pdf.add_field("Emergency Contact 1:", teacher[13], 50, 45)
            pdf.set_y(current_y)
            pdf.set_x(110)
            pdf.add_field("Emergency Contact 2:", teacher[14], 40, 45)
    
            # Declaration Section
            pdf.section_header("DECLARATION")
            pdf.set_font("Arial", "", 10)
            declaration_text = (
                "I, _____________________________________________________________, "
                "hereby declare that the information provided in this form is true "
                "and accurate to the best of my knowledge.\n\n\n"
                "Signature: ________________________________________________    "
                "Date: _________________________\n\n\n"
            )
            pdf.multi_cell(0, 3, declaration_text)
    
            # FOR OFFICIAL USE ONLY Section - Highlighted
            pdf.section_header("FOR OFFICIAL USE ONLY", highlight=True)
            pdf.set_font("Arial", "", 10)
            official_text = (
                "Recommended for appointment on (Date): ___________________________________\n\n\n"
                "Administrator Signature: ___________________________________     "
                "Date: _________________________\n\n"
                "                                                                                                    (Administrator)"
            )
            pdf.multi_cell(0, 3, official_text)
    
            # Use temporary file approach to get PDF bytes
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
                temp_path = temp_file.name
            
            try:
                # Output to temporary file
                pdf.output(temp_path)
                
                # Read the file back as bytes
                with open(temp_path, 'rb') as f:
                    pdf_bytes = f.read()
                
            finally:
                # Clean up temporary file
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
    
            # Cache for reuse
            self.current_pdf_bytes = pdf_bytes
            self.current_file_type = "pdf"
    
            return pdf_bytes
    
        except Exception as e:
            raise Exception(f"Error generating teacher profile PDF: {str(e)}")
    
    def export_teacher_to_pdf_form(self):
        """Export selected teacher to PDF and show viewer"""
        try:
            if not self.current_teacher_id:
                QMessageBox.warning(self, "No Selection", "Please select a teacher first")
                return
    
            # Ensure connection is alive
            self._ensure_connection()
    
            # Generate PDF
            pdf_bytes = self.generate_teacher_profile_pdf(self.current_teacher_id)
    
            # Use standard viewer
            try:
                from utils.pdf_utils import view_pdf
                view_pdf(pdf_bytes, parent=self)
            except ImportError:
                # Fallback if utils module not available
                QMessageBox.information(self, "PDF Generated", "PDF generated successfully but viewer not available")
    
        except Exception as e:
            import traceback
            print(f"Full error: {traceback.format_exc()}")
            QMessageBox.critical(self, "Export Error", f"Failed to generate or view PDF:\n{str(e)}")

    def setup_analytics_tab(self):
        """Setup the analytics tab for staff statistics and charts"""
        # Main scroll area
        analytics_scroll = QScrollArea()
        analytics_scroll.setWidgetResizable(True)
        analytics_scroll.setFrameShape(QFrame.NoFrame)
        analytics_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        analytics_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        # Container widget
        analytics_container = QWidget()
        main_layout = QVBoxLayout(analytics_container)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        analytics_scroll.setWidget(analytics_container)
        
        # Set scroll area as main layout
        self.analytics_tab.setLayout(QVBoxLayout())
        self.analytics_tab.layout().addWidget(analytics_scroll)
        self.analytics_tab.layout().setContentsMargins(0, 0, 0, 0)
        
        # Title and refresh button
        header_layout = QHBoxLayout()
        title = QLabel("Staff Analytics Dashboard")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #1f538d; padding: 10px;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        refresh_analytics_btn = QPushButton("Refresh Data")
        refresh_analytics_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                font-weight: bold;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #218838; }
        """)
        refresh_analytics_btn.clicked.connect(self.refresh_staff_analytics)
        header_layout.addWidget(refresh_analytics_btn)
        
        main_layout.addLayout(header_layout)
        
        # === OVERVIEW STATISTICS CARDS ===
        stats_container = QWidget()
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)
        
        # Total Staff Card
        self.total_staff_card = self.create_stats_card("Total Staff", "0", "#007bff")
        stats_layout.addWidget(self.total_staff_card)
        
        # Active Staff Card
        self.active_staff_card = self.create_stats_card("Active Staff", "0", "#28a745")
        stats_layout.addWidget(self.active_staff_card)
        
        # Inactive Staff Card
        self.inactive_staff_card = self.create_stats_card("Inactive Staff", "0", "#dc3545")
        stats_layout.addWidget(self.inactive_staff_card)
        
        # Male/Female Ratio Card
        self.gender_ratio_card = self.create_stats_card("Gender Ratio", "M: 0% | F: 0%", "#6f42c1")
        stats_layout.addWidget(self.gender_ratio_card)
        
        main_layout.addWidget(stats_container)
        
        # === CHARTS SECTION ===
        charts_container = QWidget()
        charts_layout = QHBoxLayout(charts_container)
        charts_layout.setSpacing(20)
        
        # Left side - Staff Type Distribution
        left_chart_group = QGroupBox("Distribution by Staff Type")
        left_chart_group.setMinimumHeight(400)
        left_chart_layout = QVBoxLayout()
        left_chart_group.setLayout(left_chart_layout)
        
        # Create matplotlib figure for staff type distribution
        self.staff_type_figure = Figure(figsize=(8, 6))
        self.staff_type_canvas = FigureCanvas(self.staff_type_figure)
        self.staff_type_canvas.setMinimumHeight(350)
        left_chart_layout.addWidget(self.staff_type_canvas)
        charts_layout.addWidget(left_chart_group, 1)
        
        # Right side - Position Distribution  
        right_chart_group = QGroupBox("Distribution by Position")
        right_chart_group.setMinimumHeight(400)
        right_chart_layout = QVBoxLayout()
        right_chart_group.setLayout(right_chart_layout)
        
        # Create matplotlib figure for position distribution
        self.position_figure = Figure(figsize=(8, 6))
        self.position_canvas = FigureCanvas(self.position_figure)
        self.position_canvas.setMinimumHeight(350)
        right_chart_layout.addWidget(self.position_canvas)
        charts_layout.addWidget(right_chart_group, 1)
        
        main_layout.addWidget(charts_container)
        
        # === DETAILED TABLES ===
        tables_container = QWidget()
        tables_layout = QHBoxLayout(tables_container)
        tables_layout.setSpacing(20)
        
        # Staff type breakdown table
        staff_type_table_group = QGroupBox("Staff Type Breakdown")
        staff_type_table_group.setMinimumHeight(250)
        staff_type_table_layout = QVBoxLayout()
        staff_type_table_group.setLayout(staff_type_table_layout)
        
        self.staff_type_stats_table = QTableWidget()
        self.staff_type_stats_table.setColumnCount(4)
        self.staff_type_stats_table.setHorizontalHeaderLabels(["Staff Type", "Male", "Female", "Total"])
        self.staff_type_stats_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.staff_type_stats_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.staff_type_stats_table.setMinimumHeight(250)
        self.staff_type_stats_table.setAlternatingRowColors(True)
        self.staff_type_stats_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
            }
        """)
        staff_type_table_layout.addWidget(self.staff_type_stats_table)
        tables_layout.addWidget(staff_type_table_group, 1)
        
        # Position breakdown table
        position_table_group = QGroupBox("Position Breakdown")
        position_table_group.setMinimumHeight(250)
        position_table_layout = QVBoxLayout()
        position_table_group.setLayout(position_table_layout)
        
        self.position_stats_table = QTableWidget()
        self.position_stats_table.setColumnCount(4)
        self.position_stats_table.setHorizontalHeaderLabels(["Position", "Male", "Female", "Total"])
        self.position_stats_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.position_stats_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.position_stats_table.setMinimumHeight(250)
        self.position_stats_table.setAlternatingRowColors(True)
        self.position_stats_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
            }
        """)
        position_table_layout.addWidget(self.position_stats_table)
        tables_layout.addWidget(position_table_group, 1)
        
        main_layout.addWidget(tables_container)
        
        # Add stretch to push content to top and allow scrolling
        main_layout.addStretch()
        
        # Load initial analytics data
        QTimer.singleShot(500, self.refresh_staff_analytics)
    
    def create_stats_card(self, title, value, color):
        """Create a statistics card widget for staff analytics"""
        card = QFrame()
        card.setFrameStyle(QFrame.Box)
        card.setMinimumHeight(120)
        card.setMinimumWidth(200)
        card.setStyleSheet(f"""
            QFrame {{
                border: 2px solid {color};
                border-radius: 10px;
                background-color: white;
                padding: 15px;
            }}
            QLabel {{
                border: none;
                background: transparent;
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setAlignment(Qt.AlignCenter)
        layout.setSpacing(8)
        
        # Title
        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(f"font-size: 14px; font-weight: bold; color: {color}; margin-bottom: 5px;")
        title_label.setWordWrap(True)
        layout.addWidget(title_label)
        
        # Value
        value_label = QLabel(value)
        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #333;")
        value_label.setWordWrap(True)
        layout.addWidget(value_label)
        
        # Store value label for updates
        card.value_label = value_label
        
        return card
    
    def refresh_staff_analytics(self):
        """Refresh all staff analytics data"""
        try:
            # Show loading cursor
            QApplication.setOverrideCursor(Qt.WaitCursor)
            
            # Load all analytics components
            self.load_staff_overview_stats()
            self.load_staff_type_distribution()
            self.load_position_distribution()
            self.update_staff_charts()
            
            # Show success popup only if manually triggered (not initial load)
            if hasattr(self, '_analytics_initial_load_complete'):
                QApplication.restoreOverrideCursor()
                QMessageBox.information(self, "Refresh Complete", 
                                      "Staff analytics data has been refreshed successfully!")
            else:
                self._analytics_initial_load_complete = True
                QApplication.restoreOverrideCursor()
            
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "Error", f"Failed to load staff analytics: {str(e)}")
    
    def load_staff_overview_stats(self):
        """Load overview statistics for staff"""
        try:
            # Total staff
            self.cursor.execute("SELECT COUNT(*) FROM teachers")
            total_staff = self.cursor.fetchone()[0]
            
            # Active staff
            self.cursor.execute("SELECT COUNT(*) FROM teachers WHERE is_active = TRUE")
            total_active = self.cursor.fetchone()[0]
            
            # Inactive staff
            total_inactive = total_staff - total_active
            
            # Gender distribution (active staff only)
            self.cursor.execute("""
                SELECT gender, COUNT(*) FROM teachers 
                WHERE is_active = TRUE 
                GROUP BY gender
            """)
            gender_data = self.cursor.fetchall()
            
            male_count = 0
            female_count = 0
            
            for gender, count in gender_data:
                if gender and gender.lower() in ['male', 'm']:
                    male_count = count
                elif gender and gender.lower() in ['female', 'f']:
                    female_count = count
            
            # Calculate percentages
            if total_active > 0:
                male_percent = round((male_count / total_active) * 100, 1)
                female_percent = round((female_count / total_active) * 100, 1)
            else:
                male_percent = female_percent = 0
            
            # Update cards
            self.total_staff_card.value_label.setText(str(total_staff))
            self.active_staff_card.value_label.setText(str(total_active))
            self.inactive_staff_card.value_label.setText(str(total_inactive))
            self.gender_ratio_card.value_label.setText(f"M: {male_percent}% | F: {female_percent}%")
            
            # Force UI update
            self.total_staff_card.value_label.update()
            self.active_staff_card.value_label.update()
            self.inactive_staff_card.value_label.update()
            self.gender_ratio_card.value_label.update()
            
        except Exception as e:
            print(f"Error loading staff overview stats: {e}")
    
    def load_staff_type_distribution(self):
        """Load distribution by staff type"""
        try:
            query = """
                SELECT 
                    staff_type,
                    gender,
                    COUNT(*) as count
                FROM teachers 
                WHERE is_active = TRUE 
                GROUP BY staff_type, gender
                ORDER BY staff_type, gender
            """
            self.cursor.execute(query)
            staff_type_data = self.cursor.fetchall()
            
            # Process data into a dictionary
            staff_type_stats = {}
            for staff_type, gender, count in staff_type_data:
                if staff_type not in staff_type_stats:
                    staff_type_stats[staff_type] = {'Male': 0, 'Female': 0, 'Total': 0}
                
                if gender and gender.lower() in ['male', 'm']:
                    staff_type_stats[staff_type]['Male'] = count
                elif gender and gender.lower() in ['female', 'f']:
                    staff_type_stats[staff_type]['Female'] = count
                
                staff_type_stats[staff_type]['Total'] += count
            
            # Update table
            staff_types = list(staff_type_stats.keys())
            self.staff_type_stats_table.setRowCount(len(staff_types))
            
            for row_idx, staff_type in enumerate(staff_types):
                stats = staff_type_stats.get(staff_type, {'Male': 0, 'Female': 0, 'Total': 0})
                
                self.staff_type_stats_table.setItem(row_idx, 0, QTableWidgetItem(staff_type or "Unknown"))
                self.staff_type_stats_table.setItem(row_idx, 1, QTableWidgetItem(str(stats['Male'])))
                self.staff_type_stats_table.setItem(row_idx, 2, QTableWidgetItem(str(stats['Female'])))
                self.staff_type_stats_table.setItem(row_idx, 3, QTableWidgetItem(str(stats['Total'])))
            
            self.staff_type_stats_data = staff_type_stats
            
        except Exception as e:
            print(f"Error loading staff type distribution: {e}")
    
    def load_position_distribution(self):
        """Load distribution by position"""
        try:
            query = """
                SELECT 
                    position,
                    gender,
                    COUNT(*) as count
                FROM teachers 
                WHERE is_active = TRUE 
                AND position IS NOT NULL 
                AND position != ''
                GROUP BY position, gender
                ORDER BY position, gender
            """
            self.cursor.execute(query)
            position_data = self.cursor.fetchall()
            
            # Process data
            position_stats = {}
            for position, gender, count in position_data:
                if position not in position_stats:
                    position_stats[position] = {'Male': 0, 'Female': 0, 'Total': 0}
                
                if gender and gender.lower() in ['male', 'm']:
                    position_stats[position]['Male'] = count
                elif gender and gender.lower() in ['female', 'f']:
                    position_stats[position]['Female'] = count
                
                position_stats[position]['Total'] += count
            
            # Update table - show top 10 positions by total count
            sorted_positions = sorted(position_stats.items(), key=lambda x: x[1]['Total'], reverse=True)
            top_positions = [item[0] for item in sorted_positions[:10]]  # Top 10 positions
            
            self.position_stats_table.setRowCount(len(top_positions))
            
            for row_idx, position in enumerate(top_positions):
                stats = position_stats.get(position, {'Male': 0, 'Female': 0, 'Total': 0})
                
                # Calculate percentages
                total = stats['Total']
                if total > 0:
                    male_percent = f"{stats['Male']} ({round(stats['Male']/total*100, 1)}%)"
                    female_percent = f"{stats['Female']} ({round(stats['Female']/total*100, 1)}%)"
                else:
                    male_percent = "0 (0%)"
                    female_percent = "0 (0%)"
                
                self.position_stats_table.setItem(row_idx, 0, QTableWidgetItem(position))
                self.position_stats_table.setItem(row_idx, 1, QTableWidgetItem(male_percent))
                self.position_stats_table.setItem(row_idx, 2, QTableWidgetItem(female_percent))
                self.position_stats_table.setItem(row_idx, 3, QTableWidgetItem(str(total)))
            
            self.position_stats_data = position_stats
            
        except Exception as e:
            print(f"Error loading position distribution: {e}")
    
    def update_staff_charts(self):
        """Update both staff charts with current data"""
        self.update_staff_type_chart()
        self.update_position_chart()
    
    def update_staff_type_chart(self):
        """Update the staff type distribution chart"""
        try:
            self.staff_type_figure.clear()
            ax = self.staff_type_figure.add_subplot(111)
            
            if hasattr(self, 'staff_type_stats_data') and self.staff_type_stats_data:
                staff_types = list(self.staff_type_stats_data.keys())
                male_counts = [self.staff_type_stats_data.get(staff_type, {}).get('Male', 0) for staff_type in staff_types]
                female_counts = [self.staff_type_stats_data.get(staff_type, {}).get('Female', 0) for staff_type in staff_types]
                
                x = range(len(staff_types))
                width = 0.35
                
                bars1 = ax.bar([i - width/2 for i in x], male_counts, width, 
                              label='Male', color='#4472C4', alpha=0.8)
                bars2 = ax.bar([i + width/2 for i in x], female_counts, width,
                              label='Female', color='#E15759', alpha=0.8)
                
                ax.set_xlabel('Staff Type', fontsize=12, fontweight='bold')
                ax.set_ylabel('Number of Staff', fontsize=12, fontweight='bold')
                ax.set_title('Staff Distribution by Type', fontsize=14, fontweight='bold', pad=20)
                ax.set_xticks(x)
                ax.set_xticklabels(staff_types, fontsize=11, rotation=45, ha='right')
                ax.legend(fontsize=11)
                ax.grid(True, alpha=0.3)
                
                # Add value labels on bars
                for bar in bars1:
                    height = bar.get_height()
                    if height > 0:
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{int(height)}', ha='center', va='bottom', 
                               fontsize=9, fontweight='bold')
                
                for bar in bars2:
                    height = bar.get_height()
                    if height > 0:
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{int(height)}', ha='center', va='bottom', 
                               fontsize=9, fontweight='bold')
            else:
                # Show placeholder when no data
                ax.text(0.5, 0.5, 'No Data Available\nAdd staff to see charts', 
                       horizontalalignment='center', verticalalignment='center',
                       transform=ax.transAxes, fontsize=14, color='gray')
                ax.set_title('Staff Distribution by Type', fontsize=14, fontweight='bold')
            
            self.staff_type_figure.tight_layout()
            self.staff_type_canvas.draw()
            
        except Exception as e:
            print(f"Error updating staff type chart: {e}")
    
    def update_position_chart(self):
        """Update the position distribution chart"""
        try:
            self.position_figure.clear()
            ax = self.position_figure.add_subplot(111)
            
            if hasattr(self, 'position_stats_data') and self.position_stats_data:
                # Get top 10 positions by total count
                sorted_positions = sorted(self.position_stats_data.items(), key=lambda x: x[1]['Total'], reverse=True)
                top_positions = [item[0] for item in sorted_positions[:10]]
                
                male_counts = [self.position_stats_data.get(position, {}).get('Male', 0) for position in top_positions]
                female_counts = [self.position_stats_data.get(position, {}).get('Female', 0) for position in top_positions]
                
                x = range(len(top_positions))
                width = 0.35
                
                bars1 = ax.bar([i - width/2 for i in x], male_counts, width,
                              label='Male', color='#70AD47', alpha=0.8)
                bars2 = ax.bar([i + width/2 for i in x], female_counts, width,
                              label='Female', color='#FFC000', alpha=0.8)
                
                ax.set_xlabel('Position', fontsize=12, fontweight='bold')
                ax.set_ylabel('Number of Staff', fontsize=12, fontweight='bold')
                ax.set_title('Staff Distribution by Position (Top 10)', fontsize=14, fontweight='bold', pad=20)
                ax.set_xticks(x)
                ax.set_xticklabels(top_positions, fontsize=11, rotation=45, ha='right')
                ax.legend(fontsize=11)
                ax.grid(True, alpha=0.3)
                
                # Add value labels on bars
                for bar in bars1:
                    height = bar.get_height()
                    if height > 0:
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{int(height)}', ha='center', va='bottom', 
                               fontsize=9, fontweight='bold')
                
                for bar in bars2:
                    height = bar.get_height()
                    if height > 0:
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{int(height)}', ha='center', va='bottom', 
                               fontsize=9, fontweight='bold')
            else:
                # Show placeholder when no data
                ax.text(0.5, 0.5, 'No Data Available\nAdd staff to see charts', 
                       horizontalalignment='center', verticalalignment='center',
                       transform=ax.transAxes, fontsize=14, color='gray')
                ax.set_title('Staff Distribution by Position', fontsize=14, fontweight='bold')
            
            self.position_figure.tight_layout()
            self.position_canvas.draw()
            
        except Exception as e:
            print(f"Error updating position chart: {e}")
        
    def import_teachers_data(self):
        """Import teachers data from CSV or Excel file"""
        try:
            # Get file path from user
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Select Import File",
                "",
                "Data Files (*.csv *.xlsx *.xls);;CSV Files (*.csv);;Excel Files (*.xlsx *.xls)"
            )
            
            if not file_path:
                return  # User cancelled
            
            # Show import options dialog
            dialog = ImportOptionsDialog(self)
            if dialog.exec() != QDialog.Accepted:
                return
            
            options = dialog.get_options()
            
            # Process the file based on extension
            if file_path.lower().endswith('.csv'):
                teachers_data = self.process_csv_file(file_path, options)
            elif file_path.lower().endswith(('.xlsx', '.xls')):
                teachers_data = self.process_excel_file(file_path, options)
            else:
                QMessageBox.warning(self, "Error", "Unsupported file format")
                return
            
            if not teachers_data:
                QMessageBox.warning(self, "Error", "No valid data found in the file")
                return
            
            # Show preview and get confirmation
            preview_dialog = ImportPreviewDialog(teachers_data, self)
            if preview_dialog.exec() != QDialog.Accepted:
                return
            
            # Import the data
            success_count, error_count = self.import_to_database(teachers_data, options.get('update_existing', False))
            
            # Show results
            QMessageBox.information(
                self,
                "Import Complete",
                f"Import completed successfully!\n\n"
                f"Successfully imported: {success_count} teachers\n"
                f"Failed to import: {error_count} teachers"
            )
            
            # Refresh the table
            self.load_teachers()
            
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Error during import: {str(e)}")
            print(f"Import error: {traceback.format_exc()}")
    
    def process_csv_file(self, file_path, options):
        """Process CSV file and extract teacher data"""
        teachers_data = []
        
        try:
            with open(file_path, 'r', encoding=options.get('encoding', 'utf-8')) as file:
                # Detect delimiter
                sample = file.read(1024)
                file.seek(0)
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(sample)
                
                reader = csv.DictReader(file, dialect=dialect)
                
                for row_num, row in enumerate(reader, 1):
                    try:
                        teacher = self.map_csv_row_to_teacher(row, options.get('column_mapping', {}))
                        if teacher:
                            teachers_data.append(teacher)
                    except Exception as e:
                        print(f"Error processing row {row_num}: {e}")
                        
        except Exception as e:
            raise Exception(f"Error processing CSV file: {str(e)}")
        
        return teachers_data
    
    def process_excel_file(self, file_path, options):
        """Process Excel file and extract teacher data"""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value or f"Column_{cell.column}")
            
            teachers_data = []
            
            # Process each row
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    row_dict = dict(zip(headers, row))
                    teacher = self.map_csv_row_to_teacher(row_dict, options.get('column_mapping', {}))
                    if teacher:
                        teachers_data.append(teacher)
                except Exception as e:
                    print(f"Error processing row {row_num}: {e}")
                    
            return teachers_data
            
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")
    
    def map_csv_row_to_teacher(self, row, column_mapping):
        """Map CSV row data to teacher dictionary"""
        if not row:
            return None
        
        # Apply column mapping if provided
        if column_mapping:
            mapped_row = {}
            for db_field, csv_column in column_mapping.items():
                if csv_column in row:
                    mapped_row[db_field] = row[csv_column]
            row = mapped_row
        
        # Create teacher dictionary with default values
        teacher = {
            'teacher_id_code': row.get('teacher_id_code', '').strip(),
            'salutation': row.get('salutation', 'Mr.').strip(),
            'first_name': row.get('first_name', '').strip(),
            'surname': row.get('surname', '').strip(),
            'full_name': f"{row.get('first_name', '').strip()} {row.get('surname', '').strip()}".strip(),
            'email': row.get('email', '').strip().lower(),
            'gender': row.get('gender', 'Male').strip(),
            'phone_contact_1': row.get('phone_contact_1', '').strip(),
            'day_phone': row.get('day_phone', '').strip(),
            'current_address': row.get('current_address', '').strip(),
            'home_district': row.get('home_district', '').strip(),
            'subject_specialty': row.get('subject_specialty', '').strip(),
            'qualification': row.get('qualification', '').strip(),
            'date_joined': row.get('date_joined', ''),
            'emergency_contact_1': row.get('emergency_contact_1', '').strip(),
            'emergency_contact_2': row.get('emergency_contact_2', '').strip(),
            'national_id_number': row.get('national_id_number', '').strip(),
            'birth_date': row.get('birth_date', ''),
            'bank_account_number': row.get('bank_account_number', '').strip(),
            'next_of_kin': row.get('next_of_kin', '').strip(),
            'employment_status': row.get('employment_status', 'Full-time').strip(),
            'is_active': str(row.get('is_active', 'True')).lower() in ('true', 'yes', '1', 'y'),
            'staff_type': row.get('staff_type', 'Teaching').strip(),
            'position': row.get('position', '').strip(),
            'monthly_salary': float(row.get('monthly_salary', 0) or 0),
            'school_id': None  # Will be set based on school name mapping
        }
        
        # Handle school mapping
        school_name = row.get('school_name', '').strip()
        if school_name:
            teacher['school_name'] = school_name
        
        return teacher
    
    def import_to_database(self, teachers_data, update_existing=False):
        """Import teachers data into database"""
        success_count = 0
        error_count = 0
        
        try:
            for teacher_data in teachers_data:
                try:
                    # Get school ID
                    school_id = None
                    if 'school_name' in teacher_data:
                        school_id = self.get_school_id_by_name(teacher_data['school_name'])
                    
                    if not school_id:
                        print(f"Warning: School not found for {teacher_data.get('full_name')}")
                        error_count += 1
                        continue
                    
                    teacher_data['school_id'] = school_id
                    
                    # Check if teacher already exists
                    existing_teacher_id = None
                    if teacher_data['teacher_id_code']:
                        self.cursor.execute(
                            "SELECT id FROM teachers WHERE teacher_id_code = %s",
                            (teacher_data['teacher_id_code'],)
                        )
                        result = self.cursor.fetchone()
                        if result:
                            existing_teacher_id = result[0]
                    
                    if existing_teacher_id and update_existing:
                        # Update existing teacher
                        query = '''
                            UPDATE teachers SET
                                school_id = %s, salutation = %s, first_name = %s, surname = %s,
                                full_name = %s, email = %s, gender = %s, phone_contact_1 = %s,
                                day_phone = %s, current_address = %s, home_district = %s,
                                subject_specialty = %s, qualification = %s, date_joined = %s,
                                emergency_contact_1 = %s, emergency_contact_2 = %s,
                                national_id_number = %s, birth_date = %s, bank_account_number = %s,
                                next_of_kin = %s, employment_status = %s, is_active = %s,
                                staff_type = %s, position = %s, monthly_salary = %s,
                                updated_at = CURRENT_TIMESTAMP
                            WHERE id = %s
                        '''
                        values = (
                            teacher_data['school_id'], teacher_data['salutation'],
                            teacher_data['first_name'], teacher_data['surname'],
                            teacher_data['full_name'], teacher_data['email'],
                            teacher_data['gender'], teacher_data['phone_contact_1'],
                            teacher_data['day_phone'], teacher_data['current_address'],
                            teacher_data['home_district'], teacher_data['subject_specialty'],
                            teacher_data['qualification'], teacher_data['date_joined'],
                            teacher_data['emergency_contact_1'], teacher_data['emergency_contact_2'],
                            teacher_data['national_id_number'], teacher_data['birth_date'],
                            teacher_data['bank_account_number'], teacher_data['next_of_kin'],
                            teacher_data['employment_status'], teacher_data['is_active'],
                            teacher_data['staff_type'], teacher_data['position'],
                            teacher_data['monthly_salary'], existing_teacher_id
                        )
                    else:
                        # Insert new teacher
                        query = '''
                            INSERT INTO teachers (
                                school_id, teacher_id_code, salutation, first_name, surname, full_name,
                                email, gender, phone_contact_1, day_phone, current_address, home_district,
                                subject_specialty, qualification, date_joined, emergency_contact_1,
                                emergency_contact_2, national_id_number, birth_date, bank_account_number,
                                next_of_kin, employment_status, is_active, staff_type, position, monthly_salary
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        '''
                        values = (
                            teacher_data['school_id'], teacher_data['teacher_id_code'],
                            teacher_data['salutation'], teacher_data['first_name'],
                            teacher_data['surname'], teacher_data['full_name'],
                            teacher_data['email'], teacher_data['gender'],
                            teacher_data['phone_contact_1'], teacher_data['day_phone'],
                            teacher_data['current_address'], teacher_data['home_district'],
                            teacher_data['subject_specialty'], teacher_data['qualification'],
                            teacher_data['date_joined'], teacher_data['emergency_contact_1'],
                            teacher_data['emergency_contact_2'], teacher_data['national_id_number'],
                            teacher_data['birth_date'], teacher_data['bank_account_number'],
                            teacher_data['next_of_kin'], teacher_data['employment_status'],
                            teacher_data['is_active'], teacher_data['staff_type'],
                            teacher_data['position'], teacher_data['monthly_salary']
                        )
                    
                    self.cursor.execute(query, values)
                    success_count += 1
                    
                except Exception as e:
                    print(f"Error importing teacher {teacher_data.get('full_name')}: {e}")
                    error_count += 1
            
            self.db_connection.commit()
            
        except Exception as e:
            self.db_connection.rollback()
            raise e
        
        return success_count, error_count
    
    def get_school_id_by_name(self, school_name):
        """Get school ID by name"""
        try:
            self.cursor.execute(
                "SELECT id FROM schools WHERE school_name = %s",
                (school_name,)
            )
            result = self.cursor.fetchone()
            return result[0] if result else None
        except Exception as e:
            print(f"Error getting school ID: {e}")
            return None

class ImportOptionsDialog(QDialog):
    """Dialog for import options"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Import Options")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        self.options = {
            'encoding': 'utf-8',
            'update_existing': False,
            'column_mapping': {}
        }
        
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Encoding selection
        encoding_group = QGroupBox("File Encoding")
        encoding_layout = QVBoxLayout(encoding_group)
        
        self.encoding_combo = QComboBox()
        self.encoding_combo.addItems(['utf-8', 'latin-1', 'cp1252', 'ascii'])
        encoding_layout.addWidget(QLabel("Select file encoding:"))
        encoding_layout.addWidget(self.encoding_combo)
        
        layout.addWidget(encoding_group)
        
        # Update option
        self.update_checkbox = QCheckBox("Update existing teachers if found")
        layout.addWidget(self.update_checkbox)
        
        # Buttons
        button_layout = QHBoxLayout()
        self.ok_btn = QPushButton("OK")
        self.ok_btn.clicked.connect(self.accept)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.ok_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addLayout(button_layout)
    
    def get_options(self):
        self.options['encoding'] = self.encoding_combo.currentText()
        self.options['update_existing'] = self.update_checkbox.isChecked()
        return self.options


class ImportPreviewDialog(QDialog):
    """Dialog to preview import data"""
    
    def __init__(self, teachers_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Import Preview")
        self.setModal(True)
        self.setMinimumSize(800, 500)
        
        self.teachers_data = teachers_data
        
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Preview table
        self.preview_table = QTableWidget()
        self.preview_table.setRowCount(min(10, len(self.teachers_data)))  # Show max 10 rows
        self.preview_table.setColumnCount(5)  # Show key columns
        
        headers = ["Teacher ID", "Full Name", "Email", "Position", "School"]
        self.preview_table.setHorizontalHeaderLabels(headers)
        
        # Populate table
        for row, teacher in enumerate(self.teachers_data[:10]):
            self.preview_table.setItem(row, 0, QTableWidgetItem(teacher.get('teacher_id_code', '')))
            self.preview_table.setItem(row, 1, QTableWidgetItem(teacher.get('full_name', '')))
            self.preview_table.setItem(row, 2, QTableWidgetItem(teacher.get('email', '')))
            self.preview_table.setItem(row, 3, QTableWidgetItem(teacher.get('position', '')))
            self.preview_table.setItem(row, 4, QTableWidgetItem(teacher.get('school_name', '')))
        
        self.preview_table.resizeColumnsToContents()
        layout.addWidget(QLabel(f"Preview (showing first 10 of {len(self.teachers_data)} records):"))
        layout.addWidget(self.preview_table)
        
        # Summary
        layout.addWidget(QLabel(f"Total records to import: {len(self.teachers_data)}"))
        
        # Buttons
        button_layout = QHBoxLayout()
        self.import_btn = QPushButton("Import")
        self.import_btn.clicked.connect(self.accept)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.import_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addLayout(button_layout)

        
    def closeEvent(self, event):
        """Handle close event to clean up database connection"""
        try:
            if hasattr(self, 'cursor'):
                self.cursor.close()
            if hasattr(self, 'db_connection'):
                self.db_connection.close()
        except:
            pass
        event.accept()

    def __del__(self):
        """Close database connection when object is destroyed"""
        try:
            if hasattr(self, 'cursor'):
                self.cursor.close()
            if hasattr(self, 'db_connection'):
                self.db_connection.close()
        except:
            pass
