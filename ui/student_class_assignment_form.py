import sys
import os
import traceback
from datetime import datetime
from typing import Optional, Dict, Any, List
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QMessageBox, QFileDialog, QScrollArea, QFrame, QSizePolicy,
    QGroupBox, QGridLayout, QSpacerItem, QComboBox, QFormLayout, 
    QTabWidget, QMenu, QCheckBox, QDateEdit, QTextEdit, QApplication,
    QSplitter, QListWidget, QListWidgetItem, QProgressDialog
)
from PySide6.QtGui import QFont, QPalette, QIcon, QPixmap, QPainter, QAction, QColor, QTextCursor
from PySide6.QtCore import Qt, Signal, QSize, QDate, QTimer, QDateTime
import mysql.connector
from mysql.connector import Error
from fpdf import FPDF
import platform
import subprocess
from ui.audit_base_form import AuditBaseForm
from models.models import get_db_connection
import pandas as pd
import openpyxl



class SearchableComboBox(QComboBox):
    """Enhanced ComboBox with search functionality"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.original_values = []
        self.lineEdit().textEdited.connect(self.filter_values)
        self.lineEdit().returnPressed.connect(self.on_return_pressed)
        
    def setValues(self, values):
        """Set the values for the combobox"""
        self.original_values = values
        self.clear()
        self.addItems(values)
        
    def filter_values(self, text):
        """Filter values based on typed text"""
        if not text:
            self.clear()
            self.addItems(self.original_values)
            return
            
        filtered_values = [
            value for value in self.original_values 
            if text.lower() in value.lower()
        ]
        
        self.clear()
        self.addItems(filtered_values)
        self.showPopup()
        
    def on_return_pressed(self):
        """Handle return pressed to select the first item"""
        if self.count() > 0:
            self.setCurrentIndex(0)


class StudentClassAssignmentForm(AuditBaseForm):
    def __init__(self, parent=None, user_session=None):
        super().__init__(parent, user_session)
        self.user_session = user_session
        self.selected_assignment_id = None
        
        # Database connection
        try:
            self.db_connection = get_db_connection()
            self.cursor = self.db_connection.cursor(buffered=True)
        except Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to connect to database: {e}")
            return
        
        # Data storage for dropdowns
        self.all_students = []
        self.all_classes = []
        self.all_terms = []
        self.all_academic_years = []
        self.all_grade_levels = []
        self.current_table_data = []
        self.current_assignments_data = []
        
        # Filtered data for the new UI flow
        self.filtered_classes_by_level = []
        self.filtered_students_by_class = []
        self.available_streams = []
        
        self.setup_ui()
        self.load_dropdown_data()
        self.load_data()
        
    def setup_ui(self):
        """Setup the main UI components with tabbed interface"""
        self.setWindowTitle("Student Class Assignments")
        self.setMinimumSize(1200, 800)
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Create tab widget
        self.tab_widget = QTabWidget()
        self.tab_widget.setProperty("class", "main-tabs")
        
        # Create tabs
        self.create_assignment_tab()
        self.create_table_tab()
        
        # Add tab widget to main layout
        main_layout.addWidget(self.tab_widget)
            
    def create_assignment_tab(self):
        """Create the assignment form tab with proper scrolling"""
        # Create assignment tab
        assignment_widget = QWidget()
        assignment_layout = QVBoxLayout(assignment_widget)
        assignment_layout.setContentsMargins(0, 0, 0, 0)
        assignment_layout.setSpacing(0)
        
        # Form title
        title_frame = QFrame()
        title_frame.setProperty("class", "title-frame")
        title_layout = QVBoxLayout(title_frame)
        title_layout.setContentsMargins(20, 15, 20, 15)
        
        form_title = QLabel("Student Class Assignment Form")
        form_title.setProperty("class", "page-title")
        form_title.setAlignment(Qt.AlignCenter)
        title_layout.addWidget(form_title)
        
        assignment_layout.addWidget(title_frame)
        
        # Create scroll area with proper configuration
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setFrameStyle(QFrame.NoFrame)
        scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Create scroll content widget
        scroll_content = QWidget()
        scroll_content.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Main form layout for scroll content
        main_form_layout = QVBoxLayout(scroll_content)
        main_form_layout.setContentsMargins(20, 10, 20, 20)
        main_form_layout.setSpacing(15)
        
        # Add form sections
        self.create_assignment_form_sections(main_form_layout)
        
        # Add stretch to push content to top and ensure proper spacing
        main_form_layout.addStretch(1)
        
        # Set scroll content
        scroll_area.setWidget(scroll_content)
        assignment_layout.addWidget(scroll_area)
        
        # Add tab
        self.tab_widget.addTab(assignment_widget, "New Assignment")
        
    def create_assignment_form_sections(self, form_layout):
        """Create the form sections within the assignment tab using two-pane layout"""
        
        # Create container widget for the splitter to ensure proper sizing
        container_widget = QWidget()
        container_layout = QVBoxLayout(container_widget)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)
        
        # Create horizontal splitter for two panes
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        splitter.setHandleWidth(8)
        splitter.setStyleSheet("""
            QSplitter::handle {
                background: #cccccc;
                border: 1px solid #aaaaaa;
                border-radius: 2px;
            }
            QSplitter::handle:hover {
                background: #aaaaaa;
            }
        """)
        
        # === LEFT PANE - Selection and Assignment Details ===
        left_pane = QWidget()
        left_pane.setMinimumWidth(400)  # Minimum width for left pane
        left_layout = QVBoxLayout(left_pane)
        left_layout.setContentsMargins(10, 0, 5, 0)
        left_layout.setSpacing(15)
        
        # Student & Class Selection Section
        selection_group = QGroupBox("Student & Class Selection")
        selection_group.setProperty("class", "form-section")
        selection_group.setMinimumHeight(200)  # Ensure minimum height
        selection_layout = QGridLayout(selection_group)
        selection_layout.setContentsMargins(16, 20, 16, 12)
        selection_layout.setSpacing(12)
        selection_layout.setColumnStretch(1, 1)
        
        # STEP 1: Level (O-Level/A-Level)
        level_label = QLabel("Education Level *")
        level_label.setProperty("class", "field-label")
        selection_layout.addWidget(level_label, 0, 0)
        
        self.level_dropdown = QComboBox()
        self.level_dropdown.setProperty("class", "form-control")
        self.level_dropdown.setMinimumHeight(35)
        self.level_dropdown.currentTextChanged.connect(self.on_level_selection_changed)
        selection_layout.addWidget(self.level_dropdown, 0, 1)
        
        # STEP 2: Class Name (filtered by level)
        class_label = QLabel("Class Name *")
        class_label.setProperty("class", "field-label")
        selection_layout.addWidget(class_label, 1, 0)
        
        self.class_name_dropdown = QComboBox()
        self.class_name_dropdown.setProperty("class", "form-control")
        self.class_name_dropdown.setMinimumHeight(35)
        self.class_name_dropdown.currentTextChanged.connect(self.on_class_name_selection_changed)
        selection_layout.addWidget(self.class_name_dropdown, 1, 1)
        
        # STEP 3: Stream (filtered by class)
        stream_label = QLabel("Stream *")
        stream_label.setProperty("class", "field-label")
        selection_layout.addWidget(stream_label, 2, 0)
        
        self.stream_dropdown = QComboBox()
        self.stream_dropdown.setProperty("class", "form-control")
        self.stream_dropdown.setMinimumHeight(35)
        self.stream_dropdown.currentTextChanged.connect(self.on_stream_selection_changed)
        selection_layout.addWidget(self.stream_dropdown, 2, 1)
        
        # STEP 4: Students (filtered by class and stream)
        student_label = QLabel("Student *")
        student_label.setProperty("class", "field-label")
        selection_layout.addWidget(student_label, 3, 0)
        
        self.student_dropdown = SearchableComboBox()
        self.student_dropdown.setProperty("class", "form-control")
        self.student_dropdown.setMinimumHeight(35)
        selection_layout.addWidget(self.student_dropdown, 3, 1)
        
        left_layout.addWidget(selection_group)
        
        # Assignment Details Section
        details_group = QGroupBox("Assignment Details")
        details_group.setProperty("class", "form-section")
        details_group.setMinimumHeight(200)  # Ensure minimum height
        details_layout = QGridLayout(details_group)
        details_layout.setContentsMargins(16, 20, 16, 12)
        details_layout.setSpacing(12)
        details_layout.setColumnStretch(1, 1)
        
        # Academic Year dropdown
        year_label = QLabel("Academic Year *")
        year_label.setProperty("class", "field-label")
        details_layout.addWidget(year_label, 0, 0)
        
        self.academic_year_dropdown = QComboBox()
        self.academic_year_dropdown.setProperty("class", "form-control")
        self.academic_year_dropdown.setMinimumHeight(35)
        details_layout.addWidget(self.academic_year_dropdown, 0, 1)
        
        # Term dropdown
        term_label = QLabel("Term *")
        term_label.setProperty("class", "field-label")
        details_layout.addWidget(term_label, 1, 0)
        
        self.term_dropdown = QComboBox()
        self.term_dropdown.setProperty("class", "form-control")
        self.term_dropdown.setMinimumHeight(35)
        details_layout.addWidget(self.term_dropdown, 1, 1)
        
        # Status dropdown
        status_label = QLabel("Status *")
        status_label.setProperty("class", "field-label")
        details_layout.addWidget(status_label, 2, 0)
        
        self.status_dropdown = QComboBox()
        self.status_dropdown.setProperty("class", "form-control")
        self.status_dropdown.setMinimumHeight(35)
        self.status_dropdown.addItems(["Promoted", "Completed", "Repeated"])
        details_layout.addWidget(self.status_dropdown, 2, 1)
        
        # Assignment Date (readonly display)
        date_label = QLabel("Assignment Date")
        date_label.setProperty("class", "field-label")
        details_layout.addWidget(date_label, 3, 0)
        
        self.assignment_date_entry = QLineEdit()
        self.assignment_date_entry.setProperty("class", "form-control")
        self.assignment_date_entry.setMinimumHeight(35)
        self.assignment_date_entry.setReadOnly(True)
        self.assignment_date_entry.setPlaceholderText("Will be set automatically")
        details_layout.addWidget(self.assignment_date_entry, 3, 1)
        
        left_layout.addWidget(details_group)
        left_layout.addStretch()  # Push content to top
        
        # === RIGHT PANE - Additional Options and Actions ===
        right_pane = QWidget()
        right_pane.setMinimumWidth(350)  # Minimum width for right pane
        right_layout = QVBoxLayout(right_pane)
        right_layout.setContentsMargins(5, 0, 10, 0)
        right_layout.setSpacing(15)
        
        # Additional Options Section
        options_group = QGroupBox("Additional Options")
        options_group.setProperty("class", "form-section")
        options_layout = QVBoxLayout(options_group)
        options_layout.setContentsMargins(16, 20, 16, 12)
        options_layout.setSpacing(12)
        
        # Is Current checkbox
        self.is_current_checkbox = QCheckBox("✅ Current Assignment")
        self.is_current_checkbox.setChecked(True)
        self.is_current_checkbox.setProperty("class", "form-checkbox")
        self.is_current_checkbox.setMinimumHeight(30)
        options_layout.addWidget(self.is_current_checkbox)
        
        # Notes
        notes_label = QLabel("Notes")
        notes_label.setProperty("class", "field-label")
        options_layout.addWidget(notes_label)
        
        self.notes_textbox = QTextEdit()
        self.notes_textbox.setProperty("class", "form-control")
        self.notes_textbox.setMinimumHeight(120)
        self.notes_textbox.setMaximumHeight(150)
        self.notes_textbox.setPlaceholderText("Enter any additional notes about this assignment...")
        options_layout.addWidget(self.notes_textbox)
        
        right_layout.addWidget(options_group)
        
    
        # Action Buttons Section - REORGANIZED WITH TWO COLUMNS
        buttons_group = QGroupBox("Actions")
        buttons_group.setProperty("class", "form-section")
        buttons_layout = QVBoxLayout(buttons_group)
        buttons_layout.setContentsMargins(16, 20, 16, 12)
        buttons_layout.setSpacing(15)
        
        # Create a horizontal layout for the two columns of buttons
        buttons_columns_layout = QHBoxLayout()
        buttons_columns_layout.setSpacing(20)  # Space between columns
        
        # === LEFT COLUMN - Primary Actions ===
        primary_column = QVBoxLayout()
        primary_column.setSpacing(10)
        
        primary_label = QLabel("🟢 Primary Actions")
        primary_label.setProperty("class", "section-subtitle")
        primary_label.setAlignment(Qt.AlignCenter)
        primary_column.addWidget(primary_label)
        
        # Save button
        save_btn = QPushButton("Save Assignment")
        save_btn.setProperty("class", "success")
        save_btn.setIcon(QIcon("static/icons/save.png"))
        save_btn.setIconSize(QSize(16, 16))
        save_btn.setMinimumHeight(45)
        save_btn.clicked.connect(self.save_assignment)
        primary_column.addWidget(save_btn)
        
        # Update button
        update_btn = QPushButton("Update Assignment")
        update_btn.setProperty("class", "primary")
        update_btn.setIcon(QIcon("static/icons/edit.png"))
        update_btn.setIconSize(QSize(16, 16))
        update_btn.setMinimumHeight(45)
        update_btn.clicked.connect(self.update_assignment)
        primary_column.addWidget(update_btn)
        
        # Delete button
        delete_btn = QPushButton("Delete Assignment")
        delete_btn.setProperty("class", "danger")
        delete_btn.setIcon(QIcon("static/icons/delete.png"))
        delete_btn.setIconSize(QSize(16, 16))
        delete_btn.setMinimumHeight(45)
        delete_btn.clicked.connect(self.delete_assignment)
        primary_column.addWidget(delete_btn)
        
        # Add stretch to push buttons to top
        primary_column.addStretch()
        
        # === RIGHT COLUMN - Secondary Actions ===
        secondary_column = QVBoxLayout()
        secondary_column.setSpacing(10)
        
        secondary_label = QLabel("Secondary Actions")
        secondary_label.setProperty("class", "section-subtitle")
        secondary_label.setAlignment(Qt.AlignCenter)
        secondary_column.addWidget(secondary_label)
        
        # Clear button
        clear_btn = QPushButton("Clear Form")
        clear_btn.setProperty("class", "secondary")
        clear_btn.setIcon(QIcon("static/icons/clear.png"))
        clear_btn.setIconSize(QSize(16, 16))
        clear_btn.setMinimumHeight(45)
        clear_btn.clicked.connect(self.clear_form)
        secondary_column.addWidget(clear_btn)
        
        # Refresh button
        refresh_btn = QPushButton("Refresh Data")
        refresh_btn.setProperty("class", "info")
        refresh_btn.setIcon(QIcon("static/icons/refresh.png"))
        refresh_btn.setIconSize(QSize(16, 16))
        refresh_btn.setMinimumHeight(45)
        refresh_btn.clicked.connect(self.refresh_all_data)
        secondary_column.addWidget(refresh_btn)
        
        # View Assignments button (switches to table tab)
        view_assignments_btn = QPushButton("View Assignments")
        view_assignments_btn.setProperty("class", "warning")
        view_assignments_btn.setIcon(QIcon("static/icons/view.png"))
        view_assignments_btn.setIconSize(QSize(16, 16))
        view_assignments_btn.setMinimumHeight(40)
        view_assignments_btn.clicked.connect(lambda: self.tab_widget.setCurrentIndex(1))
        secondary_column.addWidget(view_assignments_btn)
        
        
        # Add stretch to push buttons to top
        secondary_column.addStretch()
        
        # Add both columns to the horizontal layout
        buttons_columns_layout.addLayout(primary_column)
        buttons_columns_layout.addLayout(secondary_column)
        
        # Add the columns layout to the main buttons layout
        buttons_layout.addLayout(buttons_columns_layout)
        
        right_layout.addWidget(buttons_group)
        
        # Status Section
        status_group = QGroupBox("Status")
        status_group.setProperty("class", "form-section")
        status_layout = QVBoxLayout(status_group)
        status_layout.setContentsMargins(16, 20, 16, 12)
        
        self.status_label = QLabel("Ready to assign students")
        self.status_label.setProperty("class", "status-label")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setWordWrap(True)
        self.status_label.setMinimumHeight(60)
        status_layout.addWidget(self.status_label)
        
        right_layout.addWidget(status_group)
        right_layout.addStretch()  # Push content to top
        
        # Add panes to splitter
        splitter.addWidget(left_pane)
        splitter.addWidget(right_pane)
        
        # Set initial splitter sizes (60% left, 40% right)
        splitter.setSizes([600, 400])
        
        # Add splitter to container layout
        container_layout.addWidget(splitter)
        
        # Add container to form layout
        form_layout.addWidget(container_widget)
        
    def create_table_tab(self):
        """Create the table view tab"""
        # Create table tab
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        table_layout.setContentsMargins(20, 20, 20, 20)
        table_layout.setSpacing(15)
        
        # Table title and stats
        title_frame = QHBoxLayout()
        
        # Title
        title_label = QLabel("Student Class Assignments Overview")
        title_label.setProperty("class", "page-title")
        title_frame.addWidget(title_label)
        
        title_frame.addStretch()
        
        # Quick stats label
        self.stats_label = QLabel("Loading...")
        self.stats_label.setProperty("class", "stats-label")
        title_frame.addWidget(self.stats_label)
        
        table_layout.addLayout(title_frame)
        
        # Action buttons frame
        action_frame = QHBoxLayout()
        action_frame.setSpacing(10)
        
        # Promotion button
        promote_btn = QPushButton("Promote Students")
        promote_btn.setProperty("class", "success")
        promote_btn.setIcon(QIcon("static/icons/promote.png"))
        promote_btn.setIconSize(QSize(16, 16))
        promote_btn.clicked.connect(self.open_promotion_popup)
        action_frame.addWidget(promote_btn)
        
        # Demotion button
        demote_btn = QPushButton("Demote Students")
        demote_btn.setProperty("class", "danger")
        demote_btn.setIcon(QIcon("static/icons/demote.png"))
        demote_btn.setIconSize(QSize(16, 16))
        demote_btn.clicked.connect(self.open_demotion_popup)
        action_frame.addWidget(demote_btn)
        
        # Export button
        export_btn = QPushButton("Export to PDF")
        export_btn.setProperty("class", "primary")
        export_btn.setIcon(QIcon("static/icons/export.png"))
        export_btn.setIconSize(QSize(16, 16))
        export_btn.clicked.connect(self.export_to_pdf)
        action_frame.addWidget(export_btn)

        # In the action_frame section, add this button:
        export_excel_btn = QPushButton("Export to Excel")
        export_excel_btn.setProperty("class", "primary")
        export_excel_btn.setIcon(QIcon("static/icons/excel.png"))
        export_excel_btn.setIconSize(QSize(16, 16))
        export_excel_btn.clicked.connect(self.export_to_excel)
        action_frame.addWidget(export_excel_btn)
        
        action_frame.addStretch()
        
        table_layout.addLayout(action_frame)
        
        # Search components in a group
        search_group = QGroupBox("Search & Filter")
        search_group.setProperty("class", "search-section")
        search_layout = QHBoxLayout(search_group)
        search_layout.setContentsMargins(12, 16, 12, 8)
        search_layout.setSpacing(8)
        
        search_label = QLabel("Search:")
        search_label.setProperty("class", "field-label")
        search_label.setFixedWidth(50)
        search_layout.addWidget(search_label)
        
        self.search_entry = QLineEdit()
        self.search_entry.setProperty("class", "form-control")
        self.search_entry.setPlaceholderText("Search by student name, class, term, year, or status...")
        self.search_entry.textChanged.connect(self.search_assignments)
        search_layout.addWidget(self.search_entry)
        
        search_btn = QPushButton("Search")
        search_btn.setProperty("class", "primary")
        search_btn.setIcon(QIcon("static/icons/search.png"))
        search_btn.setIconSize(QSize(16, 16))
        search_btn.clicked.connect(self.search_assignments)
        search_layout.addWidget(search_btn)
        
        clear_search_btn = QPushButton("Clear")
        clear_search_btn.setProperty("class", "secondary")
        clear_search_btn.setIcon(QIcon("static/icons/clear.png"))
        clear_search_btn.setIconSize(QSize(16, 16))
        clear_search_btn.clicked.connect(self.clear_search)
        search_layout.addWidget(clear_search_btn)
        
        table_layout.addWidget(search_group)
        
        # Table - SET CORRECT NUMBER OF COLUMNS
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(7)  # 7 columns for all headers
        self.table_widget.setHorizontalHeaderLabels(["Student Name", "Grade", "Class/Stream", "Term", "Year", "Status", "Current"])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_widget.cellClicked.connect(self.on_table_row_click)
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setProperty("class", "data-table")
        
        table_layout.addWidget(self.table_widget)
        
        # Table info footer
        table_info_frame = QHBoxLayout()
        
        self.table_info_label = QLabel("Select a row to edit assignment")
        self.table_info_label.setProperty("class", "info-label")
        table_info_frame.addWidget(self.table_info_label)
        
        table_info_frame.addStretch()
        
        refresh_table_btn = QPushButton("Refresh Table")
        refresh_table_btn.setProperty("class", "secondary")
        refresh_table_btn.setIcon(QIcon("static/icons/refresh.png"))
        refresh_table_btn.setIconSize(QSize(14, 14))
        refresh_table_btn.clicked.connect(self.load_data)
        table_info_frame.addWidget(refresh_table_btn)
        
        table_layout.addLayout(table_info_frame)
        
        # Add tab
        self.tab_widget.addTab(table_widget, "View Assignments")
        
        
    def clear_search(self):
        """Clear search and reload all data"""
        self.search_entry.clear()
        self.load_data()
        
    # [Keep all the existing methods for data handling, form operations, etc.]
    # The methods below remain unchanged from your original code:
    
    def load_dropdown_data(self):
        """Load all dropdown data with the new structure"""
        try:
            # Load distinct education levels (O-Level/A-Level)
            self.cursor.execute("""
                SELECT DISTINCT level 
                FROM classes 
                WHERE level IS NOT NULL AND level != ''
                ORDER BY level
            """)
            levels = [row[0] for row in self.cursor.fetchall()]
            self.level_dropdown.clear()
            if levels:
                self.level_dropdown.addItems(levels)
            else:
                self.level_dropdown.addItem("No levels found")
            self.all_grade_levels = levels
            
            # Load only ACTIVE students with their grade information
            self.cursor.execute("""
                SELECT id, full_name, grade_applied_for 
                FROM students 
                WHERE full_name IS NOT NULL AND full_name != '' AND is_active = 1
                ORDER BY full_name
            """)
            self.all_students = self.cursor.fetchall()
            
            print(f"DEBUG: Loaded {len(self.all_students)} ACTIVE students from database:")
            for i, student in enumerate(self.all_students[:10]):  # Show first 10
                print(f"DEBUG: Student {i+1}: ID={student[0]}, Name='{student[1]}', Grade='{student[2]}'")

            # Load only ACTIVE students with their grade information
            self.cursor.execute("""
                SELECT id, full_name, grade_applied_for 
                FROM students 
                WHERE full_name IS NOT NULL AND full_name != '' AND is_active = 1
                ORDER BY full_name
            """)
            self.all_students = self.cursor.fetchall()
            
            # Clean up duplicate students (optional)
            self.cleanup_inactive_students()
            
            # Load all classes with their details
            self.cursor.execute("""
                SELECT id, class_name, stream, level 
                FROM classes 
                WHERE class_name IS NOT NULL AND class_name != ''
                ORDER BY 
                    level,
                    CASE 
                        WHEN class_name LIKE 'S%' THEN CAST(SUBSTR(class_name, 2) AS SIGNED)
                        ELSE 999 
                    END,
                    stream
            """)
            self.all_classes = self.cursor.fetchall()
            
            print(f"DEBUG: Loaded {len(self.all_classes)} classes from database:")
            for i, cls in enumerate(self.all_classes[:10]):  # Show first 10
                print(f"DEBUG: Class {i+1}: ID={cls[0]}, Name='{cls[1]}', Stream='{cls[2]}', Level='{cls[3]}'")
            
            # Load terms
            self.cursor.execute("SELECT id, term_name FROM terms ORDER BY term_name")
            self.all_terms = self.cursor.fetchall()
            self.term_dropdown.clear()
            if self.all_terms:
                terms = [row[1] for row in self.all_terms]
                self.term_dropdown.addItems(terms)
            else:
                self.term_dropdown.addItem("No terms found")
            
            # Load academic years
            self.cursor.execute("SELECT id, year_name FROM academic_years ORDER BY year_name DESC")
            self.all_academic_years = self.cursor.fetchall()
            self.academic_year_dropdown.clear()
            if self.all_academic_years:
                academic_years = [row[1] for row in self.all_academic_years]
                self.academic_year_dropdown.addItems(academic_years)
            else:
                self.academic_year_dropdown.addItem("No academic years found")
            
            self.status_label.setText(f"Data loaded: {len(self.all_students)} active students, {len(self.all_classes)} classes")
            
        except Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to load data: {e}")
            print(f"Database error: {e}")
            traceback.print_exc()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load data: {e}")
            print(f"Unexpected error: {e}")
            traceback.print_exc()

    def cleanup_inactive_students(self):
        """Optional: Clean up inactive students from the dropdown data"""
        try:
            # Remove duplicates by keeping only the latest active student for each name
            unique_students = {}
            for student_id, full_name, grade_applied in self.all_students:
                if full_name in unique_students:
                    # Keep the student with the higher ID (assuming newer records have higher IDs)
                    existing_id, existing_grade = unique_students[full_name]
                    if student_id > existing_id:
                        unique_students[full_name] = (student_id, grade_applied)
                else:
                    unique_students[full_name] = (student_id, grade_applied)
            
            # Convert back to list format
            self.all_students = [(student_id, full_name, grade_applied) 
                               for full_name, (student_id, grade_applied) in unique_students.items()]
            
            print(f"DEBUG: After cleanup: {len(self.all_students)} unique active students")
            
        except Exception as e:
            print(f"Error cleaning up students: {e}")
            
    def on_level_selection_changed(self, selected_level):
        """Filter classes by selected education level"""
        print(f"DEBUG: Level selected: {selected_level}")
        
        if not selected_level or selected_level == "No levels found":
            self.class_name_dropdown.clear()
            self.class_name_dropdown.addItem("Select education level first")
            self.student_dropdown.setValues(["Select level first..."])
            self.stream_dropdown.clear()
            self.stream_dropdown.addItem("Select level first...")
            return
        
        # Get unique class names for this level
        unique_classes = set()
        self.filtered_classes_by_level = []
        
        for class_id, class_name, stream, level in self.all_classes:
            if level == selected_level:
                unique_classes.add(class_name)
                self.filtered_classes_by_level.append((class_id, class_name, stream, level))
        
        print(f"DEBUG: Found {len(self.filtered_classes_by_level)} classes for level '{selected_level}'")
        print(f"DEBUG: Unique class names: {sorted(list(unique_classes))}")
        
        # Update class dropdown
        class_names = sorted(list(unique_classes))
        self.class_name_dropdown.clear()
        if class_names:
            self.class_name_dropdown.addItems(class_names)
            # AUTO-SELECT THE FIRST CLASS AND TRIGGER THE CHANGE
            self.class_name_dropdown.setCurrentIndex(0)
            # Manually trigger the class selection change
            current_class = self.class_name_dropdown.currentText()
            self.on_class_name_selection_changed(current_class)
        else:
            self.class_name_dropdown.addItem("No classes found")
            self.student_dropdown.setValues(["Select class first..."])
            self.stream_dropdown.clear()
            self.stream_dropdown.addItem("Select class first...")
        
    def on_class_name_selection_changed(self, selected_class_name):
        """Load streams for selected class and show ALL students in that class"""
        selected_level = self.level_dropdown.currentText()
        
        print(f"DEBUG: Class selected: {selected_class_name}, Level: {selected_level}")
        
        if not selected_class_name or not selected_level or selected_class_name == "Select education level first" or selected_class_name == "No classes found":
            self.student_dropdown.setValues(["Select class first..."])
            self.stream_dropdown.clear()
            self.stream_dropdown.addItem("Select class first...")
            return
        
        try:
            # Find all matching classes for this class name and level
            matching_classes = []
            
            for class_id, class_name, stream, level in self.filtered_classes_by_level:
                if class_name == selected_class_name and level == selected_level:
                    matching_classes.append((class_id, class_name, stream, level))
            
            print(f"DEBUG: Found {len(matching_classes)} matching classes")
            
            # Load available streams for this class
            streams = set()
            has_streams = False
            
            for class_id, cls_name, cls_stream, cls_level in matching_classes:
                if cls_stream and cls_stream.strip():  # Only add if stream is not None, empty, or just whitespace
                    streams.add(cls_stream)
                    has_streams = True
                    print(f"DEBUG: Adding stream: {cls_stream}")
            
            # If no streams found, use the class name itself as the stream
            if not has_streams:
                streams.add(selected_class_name)  # Use class name as stream
                print(f"DEBUG: No streams found, using class name as stream: {selected_class_name}")
            
            self.available_streams = sorted(list(streams))
            self.stream_dropdown.clear()
            
            print(f"DEBUG: Available streams: {self.available_streams}")
            
            if self.available_streams:
                self.stream_dropdown.addItems(self.available_streams)
                # Auto-select the first stream if available
                if len(self.available_streams) == 1:
                    self.stream_dropdown.setCurrentIndex(0)
            else:
                self.stream_dropdown.addItem("No streams available")
            
            # Load ALL students for this class (regardless of stream assignment)
            self.load_all_students_for_class(selected_class_name, selected_level)
                
        except Exception as e:
            print(f"Error in class name selection: {e}")
            QMessageBox.critical(self, "Error", f"Failed to process class selection: {e}")
            traceback.print_exc()
    
    def load_all_students_for_class(self, class_name, level):
        """Load ALL ACTIVE students for a specific class/level (for assignment to streams)"""
        try:
            student_list = []
            unassigned_count = 0
            assigned_count = 0
            
            print(f"DEBUG: Looking for ACTIVE students with grade_applied_for = '{class_name}'")
            
            # Get ALL ACTIVE students in this class (based on grade_applied_for)
            for student_id, full_name, grade_applied in self.all_students:
                # Debug: Print what we're checking
                print(f"DEBUG: Student {student_id}: '{full_name}' has grade_applied_for = '{grade_applied}'")
                
                # Check if student's grade_applied_for matches the class name
                if grade_applied and grade_applied.upper().strip() == class_name.upper().strip():
                    print(f"DEBUG: MATCH FOUND! Student {student_id} matches class {class_name}")
                    
                    # Check if student already has ANY current assignment
                    self.cursor.execute("""
                        SELECT sca.id, c.class_name, c.stream 
                        FROM student_class_assignments sca
                        JOIN classes c ON sca.class_id = c.id
                        WHERE sca.student_id = %s AND sca.is_current = 1
                    """, (student_id,))
                    current_assignment = self.cursor.fetchone()
                    
                    # Get additional student info for better display
                    self.cursor.execute("""
                        SELECT regNo, enrollment_date 
                        FROM students 
                        WHERE id = %s
                    """, (student_id,))
                    student_info = self.cursor.fetchone()
                    reg_no = student_info[0] if student_info and student_info[0] else "No Reg"
                    enrollment_date = student_info[1] if student_info and student_info[1] else "Unknown"
                    
                    display_text = f"{full_name} (Reg: {reg_no}, ID: {student_id})"
                    
                    if current_assignment:
                        current_class = current_assignment[1]
                        current_stream = current_assignment[2] if current_assignment[2] else "No Stream"
                        display_text += f" [Already in {current_class} {current_stream}]"
                        assigned_count += 1
                    else:
                        display_text += " [Unassigned - Ready for stream assignment]"
                        unassigned_count += 1
                    
                    student_list.append(display_text)
                else:
                    print(f"DEBUG: NO MATCH: '{grade_applied}' != '{class_name}'")
            
            print(f"DEBUG: Found {len(student_list)} ACTIVE students for class {class_name}")
            
            if student_list:
                self.student_dropdown.setValues(student_list)
                status_msg = f"{class_name}: {len(student_list)} active students ({unassigned_count} unassigned, {assigned_count} already assigned)"
            else:
                self.student_dropdown.setValues([f"No active students found for {class_name}"])
                status_msg = f"No active students found for {class_name}"
            
            self.status_label.setText(status_msg)
            
        except Exception as e:
            print(f"Error loading students: {e}")
            self.student_dropdown.setValues([f"Error loading students: {e}"])
        
        
    def on_stream_selection_changed(self, selected_stream):
        """When stream is selected, update status but don't filter students"""
        try:
            selected_level = self.level_dropdown.currentText()
            selected_class_name = self.class_name_dropdown.currentText()
            
            print(f"DEBUG: Stream selected: {selected_stream}, Class: {selected_class_name}, Level: {selected_level}")
            
            if not selected_stream or not selected_level or not selected_class_name or selected_stream == "Select class first...":
                return
            
            # Update status to show what stream we're preparing to assign students to
            current_text = self.status_label.text()
            if "students total" in current_text:
                # Extract the student count info and add stream info
                parts = current_text.split(" students total")
                if len(parts) > 1:
                    new_status = f"{parts[0]} students total{parts[1]} - Ready to assign to {selected_stream}"
                    self.status_label.setText(new_status)
                else:
                    self.status_label.setText(f"{current_text} - Ready to assign to {selected_stream}")
            else:
                self.status_label.setText(f"Ready to assign students to {selected_class_name} {selected_stream}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Stream selection error: {e}")
            print(f"Stream selection error: {e}")
            traceback.print_exc()
            
    def get_selected_id(self, dropdown, data_list=None):
        """Extract ID from dropdown selection"""
        try:
            selection = dropdown.currentText()
            if not selection:
                print("DEBUG: No selection in dropdown")
                return None
                
            if dropdown == self.student_dropdown:
                print(f"DEBUG: Extracting ID from: '{selection}'")
                
                # Extract ID from format: "Name (Reg: X, ID: Y) [Status]"
                import re
                
                # Look for "ID: 123" pattern
                id_match = re.search(r'ID: (\d+)', selection)
                if id_match:
                    student_id = int(id_match.group(1))
                    print(f"DEBUG: Successfully extracted ID: {student_id}")
                    return student_id
                
                print("DEBUG: Could not find ID pattern in selection")
                return None
                    
            elif dropdown == self.term_dropdown:
                for term in self.all_terms:
                    if term[1] == selection:
                        return term[0]
            elif dropdown == self.academic_year_dropdown:
                for year in self.all_academic_years:
                    if year[1] == selection:
                        return year[0]
            
            return None
                
        except (IndexError, ValueError) as e:
            print(f"Error getting selected ID: {e}")
            return None
            
    def get_selected_class_id(self):
        """Get the class ID for the selected level, class name, and stream combination"""
        try:
            selected_level = self.level_dropdown.currentText()
            selected_class_name = self.class_name_dropdown.currentText()
            selected_stream = self.stream_dropdown.currentText()
            
            print(f"DEBUG: Getting class ID for: Level={selected_level}, Class={selected_class_name}, Stream={selected_stream}")
            
            if not all([selected_level, selected_class_name, selected_stream]):
                print("DEBUG: Missing selection for class ID lookup")
                return None
            
            # Find matching class
            for class_id, class_name, stream, level in self.filtered_classes_by_level:
                # Handle the case where stream is the same as class name (meaning no actual stream)
                if selected_stream == selected_class_name:
                    # This means we want a class with no specific stream (stream is None, empty, or matches class name)
                    stream_match = (stream is None or stream == "" or stream == selected_class_name)
                    print(f"DEBUG: Looking for class with no specific stream. stream={stream}, match={stream_match}")
                else:
                    # Looking for a specific stream
                    stream_match = (stream == selected_stream)
                    print(f"DEBUG: Looking for specific stream. stream={stream}, looking_for={selected_stream}, match={stream_match}")
                
                level_match = (level == selected_level)
                class_match = (class_name == selected_class_name)
                
                if class_match and level_match and stream_match:
                    print(f"DEBUG: Found matching class ID: {class_id}")
                    return class_id
            
            print("DEBUG: No matching class found in filtered_classes_by_level")
            
            # If no match found in filtered classes, search in all classes
            print("DEBUG: Searching in all classes...")
            for class_id, class_name, stream, level in self.all_classes:
                if selected_stream == selected_class_name:
                    stream_match = (stream is None or stream == "" or stream == selected_class_name)
                else:
                    stream_match = (stream == selected_stream)
                
                if class_name == selected_class_name and level == selected_level and stream_match:
                    print(f"DEBUG: Found matching class ID in all classes: {class_id}")
                    return class_id
            
            print("DEBUG: No matching class found in all classes either")
            return None
            
        except Exception as e:
            print(f"Error getting class ID: {e}")
            return None
            
    def validate_form(self):
        """Validate form inputs"""
        print("DEBUG: Starting form validation...")
        
        if not self.level_dropdown.currentText():
            print("DEBUG: Validation failed - No level selected")
            QMessageBox.warning(self, "Validation Error", "Please select an education level")
            return False
            
        if not self.class_name_dropdown.currentText() or self.class_name_dropdown.currentText() == "No classes found":
            print("DEBUG: Validation failed - No class name selected")
            QMessageBox.warning(self, "Validation Error", "Please select a class name")
            return False
        
        if not self.stream_dropdown.currentText() or self.stream_dropdown.currentText() == "No streams available":
            print("DEBUG: Validation failed - No stream selected")
            QMessageBox.warning(self, "Validation Error", "Please select a stream")
            return False
            
        student_id = self.get_selected_id(self.student_dropdown)
        print(f"DEBUG: Extracted student ID: {student_id}")
        
        if not student_id:
            print("DEBUG: Validation failed - No valid student ID extracted")
            QMessageBox.warning(self, "Validation Error", "Please select a valid student")
            return False
        
        class_id = self.get_selected_class_id()
        print(f"DEBUG: Extracted class ID: {class_id}")
        
        if not class_id:
            print("DEBUG: Validation failed - No valid class ID extracted")
            
            # Provide more specific error message
            selected_level = self.level_dropdown.currentText()
            selected_class = self.class_name_dropdown.currentText()
            selected_stream = self.stream_dropdown.currentText()
            
            error_msg = f"No class configuration found for:\nLevel: {selected_level}\nClass: {selected_class}"
            
            if selected_stream != selected_class:
                error_msg += f"\nStream: {selected_stream}"
            
            error_msg += "\n\nPlease check if this class configuration exists or create it first."
            
            QMessageBox.warning(self, "Configuration Error", error_msg)
            return False
        
        academic_year_id = self.get_selected_id(self.academic_year_dropdown)
        print(f"DEBUG: Extracted academic year ID: {academic_year_id}")
        
        if not academic_year_id:
            print("DEBUG: Validation failed - No academic year selected")
            QMessageBox.warning(self, "Validation Error", "Please select an academic year")
            return False
        
        term_id = self.get_selected_id(self.term_dropdown)
        print(f"DEBUG: Extracted term ID: {term_id}")
        
        if not term_id:
            print("DEBUG: Validation failed - No term selected")
            QMessageBox.warning(self, "Validation Error", "Please select a term")
            return False
        
        if not self.status_dropdown.currentText():
            print("DEBUG: Validation failed - No status selected")
            QMessageBox.warning(self, "Validation Error", "Please select a status")
            return False
        
        print("DEBUG: Form validation passed!")
        return True
        
    def check_duplicate_assignment(self, student_id, class_id, academic_year_id, term_id, exclude_id=None):
        """Check if the assignment already exists"""
        query = """
            SELECT id FROM student_class_assignments 
            WHERE student_id = %s AND class_id = %s AND academic_year_id = %s AND term_id = %s
        """
        params = [student_id, class_id, academic_year_id, term_id]
        
        if exclude_id:
            query += " AND id != %s"
            params.append(exclude_id)
        
        self.cursor.execute(query, params)
        return self.cursor.fetchone() is not None

    def get_school_id(self):
        """Get school_id from user_session or database"""
        # First try to get from user session
        if self.user_session and 'school_id' in self.user_session:
            school_id = self.user_session['school_id']
            print(f"DEBUG: Using school_id from session: {school_id}")
            return school_id
        
        # If not in session, try to get from database
        try:
            self.cursor.execute("SELECT id FROM schools WHERE is_active = 1 LIMIT 1")
            result = self.cursor.fetchone()
            if result:
                school_id = result[0]
                print(f"DEBUG: Using school_id from database: {school_id}")
                return school_id
            else:
                # Create a default school if none exists
                self.cursor.execute("""
                    INSERT INTO schools (school_name, is_active) 
                    VALUES ('Default School', 1)
                """)
                school_id = self.cursor.lastrowid
                self.db_connection.commit()
                print(f"DEBUG: Created default school with ID: {school_id}")
                return school_id
                
        except Exception as e:
            print(f"Error getting school_id: {e}")
            # Ultimate fallback
            return 1
            
    def save_assignment(self):
        """Save new student class assignment (assign student to specific stream)"""
        # Validate required fields first
        if not self.validate_form():
            return
        
        try:
            # Get all selected values
            student_id = self.get_selected_id(self.student_dropdown)
            selected_level = self.level_dropdown.currentText()
            selected_class_name = self.class_name_dropdown.currentText()
            selected_stream = self.stream_dropdown.currentText()
            academic_year_id = self.get_selected_id(self.academic_year_dropdown)
            term_id = self.get_selected_id(self.term_dropdown)
            status = self.status_dropdown.currentText()
            notes = self.notes_textbox.toPlainText().strip()
            is_current = self.is_current_checkbox.isChecked()
            
            # Additional validation
            if not all([student_id, selected_level, selected_class_name, selected_stream]):
                QMessageBox.critical(self, "Error", "Missing required selection")
                return
            
            # Find the EXACT class with matching class name, level, and stream
            target_class_id = self.get_selected_class_id()
            
            # If class doesn't exist yet, we need to create it with proper stream handling
            if not target_class_id:
                # Format the message based on whether we're using a stream or not
                if selected_stream == selected_class_name:
                    message = f"Class '{selected_class_name}' doesn't exist for {selected_level} level.\nDo you want to create it?"
                else:
                    message = f"Class '{selected_class_name}' with stream '{selected_stream}' doesn't exist for {selected_level} level.\nDo you want to create it?"
                
                reply = QMessageBox.question(
                    self,
                    "Create New Class",
                    message,
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                
                if reply == QMessageBox.Yes:
                    # Create the new class - FIRST get school_id and ensure it's valid
                    school_id = self.get_school_id()
                    
                    # SAFETY CHECK: Ensure school_id is never None
                    if school_id is None:
                        print("DEBUG: school_id was None, using hardcoded fallback")
                        school_id = 1
                    
                    print(f"DEBUG: Creating class with school_id: {school_id}")
                    
                    # If stream is the same as class name, it means no actual stream
                    actual_stream = None if selected_stream == selected_class_name else selected_stream
                    
                    self.cursor.execute("""
                        INSERT INTO classes (school_id, class_name, stream, level, is_active)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (
                        school_id,
                        selected_class_name,
                        actual_stream,  # Will be NULL if no actual stream
                        selected_level,
                        True
                    ))
                    
                    target_class_id = self.cursor.lastrowid
                    print(f"DEBUG: Created new class with ID: {target_class_id}")
                    
                    # Refresh the classes data
                    self.cursor.execute("""
                        SELECT id, class_name, stream, level 
                        FROM classes 
                        WHERE class_name IS NOT NULL AND class_name != ''
                        ORDER BY level, class_name, stream
                    """)
                    self.all_classes = self.cursor.fetchall()
                    self.filtered_classes_by_level = [cls for cls in self.all_classes if cls[3] == selected_level]
                    print(f"DEBUG: Reloaded {len(self.all_classes)} classes")
                    
                else:
                    # User chose not to create the class
                    return
            
            if not target_class_id:
                QMessageBox.critical(
                    self, 
                    "Error", 
                    f"No matching class found for:\nLevel: {selected_level}\nClass: {selected_class_name}\nStream: {selected_stream}"
                )
                return
            
            # Check for duplicate assignment (same student, class, year, term)
            if self.check_duplicate_assignment(student_id, target_class_id, academic_year_id, term_id):
                # Get student and class details for better error message
                student_name = next((s[1] for s in self.all_students if s[0] == student_id), "Unknown Student")
                
                # Format class display - if stream is same as class name, just show class name
                if selected_stream == selected_class_name:
                    class_display = selected_class_name
                else:
                    class_display = f"{selected_class_name} {selected_stream}"
                
                QMessageBox.warning(
                    self,
                    "Duplicate Assignment", 
                    f"Student '{student_name}' already has an assignment to:\n"
                    f"{class_display} for the selected term and academic year.\n\n"
                    f"Please check existing assignments or select different parameters."
                )
                return
            
            # Get school_id for the assignment - with safety check
            school_id = self.get_school_id()
            print(f"DEBUG: Using school_id for assignment: {school_id}")
            
            # SAFETY CHECK: Ensure school_id is never None
            if school_id is None:
                print("DEBUG: school_id was None for assignment, using hardcoded fallback")
                school_id = 1
            
            # If this is a current assignment, deactivate other current assignments for this student
            if is_current:
                self.cursor.execute("""
                    UPDATE student_class_assignments 
                    SET is_current = 0 
                    WHERE student_id = %s AND is_current = 1
                """, (student_id,))
            
            # Insert new assignment (assigning student to the specific stream)
            print(f"DEBUG: Inserting assignment with school_id: {school_id}")
            self.cursor.execute("""
                INSERT INTO student_class_assignments (
                    school_id, student_id, class_id, academic_year_id, term_id,
                    assignment_date, is_current, status, notes
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                school_id, 
                student_id, 
                target_class_id, 
                academic_year_id, 
                term_id,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                1 if is_current else 0, 
                status, 
                notes
            ))
            
            # For current assignments, update student's grade_applied_for if needed
            if is_current:
                self.cursor.execute("""
                    UPDATE students 
                    SET grade_applied_for = %s
                    WHERE id = %s
                """, (selected_class_name, student_id))
            
            self.db_connection.commit()
            
            # Get student name for success message
            student_name = next((s[1] for s in self.all_students if s[0] == student_id), "Student")
            
            # Format class display for success message
            if selected_stream == selected_class_name:
                class_display = selected_class_name
            else:
                class_display = f"{selected_class_name} {selected_stream}"
            
            QMessageBox.information(
                self,
                "Success", 
                f"Successfully assigned student to class:\n"
                f"Student: {student_name}\n"
                f"Class: {class_display}\n"
                f"Status: {status}\n"
                f"Current: {'Yes' if is_current else 'No'}"
            )
            
            # Refresh data and clear form
            self.clear_form()
            self.load_data()
            
        except Error as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Database Error", f"Failed to save assignment:\n{e}")
            print(f"Database error: {e}")
            traceback.print_exc()
        
        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"An unexpected error occurred:\n{e}")
            print(f"Unexpected error: {e}")
            traceback.print_exc()
            
    def update_assignment(self):
        """Update existing student class assignment"""
        if not self.selected_assignment_id:
            QMessageBox.warning(self, "Warning", "Please select an assignment to update")
            return
        
        if not self.validate_form():
            return
        
        try:
            student_id = self.get_selected_id(self.student_dropdown)
            class_id = self.get_selected_class_id()
            academic_year_id = self.get_selected_id(self.academic_year_dropdown)
            term_id = self.get_selected_id(self.term_dropdown)
            
            # Check for duplicate assignment (excluding current record)
            if self.check_duplicate_assignment(student_id, class_id, academic_year_id, term_id, self.selected_assignment_id):
                QMessageBox.warning(
                    self,
                    "Duplicate Assignment", 
                    "This student is already assigned to this class for the selected term and academic year."
                )
                return
            
            status = self.status_dropdown.currentText()
            notes = self.notes_textbox.toPlainText().strip()
            
            self.cursor.execute("""
                UPDATE student_class_assignments SET 
                    student_id = %s, class_id = %s, academic_year_id = %s, term_id = %s,
                    is_current = %s, status = %s, notes = %s
                WHERE id = %s
            """, (
                student_id, class_id, academic_year_id, term_id,
                self.is_current_checkbox.isChecked(), status, notes,
                self.selected_assignment_id
            ))
            
            self.db_connection.commit()
            QMessageBox.information(self, "Success", "Student class assignment updated successfully!")
            self.clear_form()
            self.load_data()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update assignment: {e}")
            print(f"Update error: {e}")
            traceback.print_exc()
            
    def delete_assignment(self):
        """Delete selected student class assignment"""
        if not self.selected_assignment_id:
            QMessageBox.warning(self, "Warning", "Please select an assignment to delete")
            return
        
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            "Are you sure you want to delete this student class assignment?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                self.cursor.execute("DELETE FROM student_class_assignments WHERE id = %s", (self.selected_assignment_id,))
                self.db_connection.commit()
                QMessageBox.information(self, "Success", "Student class assignment deleted successfully!")
                self.clear_form()
                self.load_data()
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete assignment: {e}")
                print(f"Delete error: {e}")
                traceback.print_exc()
                
    def clear_form(self):
        """Clear all form fields"""
        self.level_dropdown.setCurrentIndex(-1)
        self.class_name_dropdown.clear()
        self.class_name_dropdown.addItem("Select education level first")
        self.stream_dropdown.clear()
        self.stream_dropdown.addItem("Select class name first...")
        
        # Show all students when form is cleared
        student_values = [f"{s[1]} (ID: {s[0]})" for s in self.all_students]
        self.student_dropdown.setValues(student_values)
        self.student_dropdown.setCurrentIndex(-1)
        
        self.academic_year_dropdown.setCurrentIndex(-1)
        self.term_dropdown.setCurrentIndex(-1)
        self.status_dropdown.setCurrentIndex(-1)
        self.assignment_date_entry.clear()
        self.notes_textbox.clear()
        self.is_current_checkbox.setChecked(True)
        self.selected_assignment_id = None
        
        self.filtered_classes_by_level = []
        self.filtered_students_by_class = []
        
        self.status_label.setText(f"Form cleared - Showing all {len(self.all_students)} students")
        
    def refresh_all_data(self):
        """Refresh all data with simple popup feedback"""
        try:
            # Ensure connection first
            self._ensure_connection()
            
            self.db_connection.commit()
            
            # Refresh data
            self.load_dropdown_data()
            self.load_data()
            
            # Get counts for the popup message
            student_count = len(self.all_students)
            class_count = len(self.all_classes)
            assignment_count = len(self.current_assignments_data) if hasattr(self, 'current_assignments_data') else 0
            
            # Show success popup
            QMessageBox.information(
                self, 
                "Refresh Complete", 
                f"✅ All data refreshed successfully!\n\n"
                f"• Students: {student_count}\n"
                f"• Classes: {class_count}\n"
                f"• Assignments: {assignment_count}",
                QMessageBox.Ok
            )
            
            # Update status label
            self.status_label.setText(f"Data refreshed: {student_count} students, {class_count} classes, {assignment_count} assignments")
            
        except Exception as e:
            # Show error popup
            QMessageBox.critical(
                self, 
                "Refresh Failed", 
                f"❌ Failed to refresh data:\n\n{str(e)}",
                QMessageBox.Ok
            )
            
            # Update status label with error
            self.status_label.setText("Refresh failed - check connection")
            
            print(f"Refresh error: {e}")
            traceback.print_exc()
        
    def load_data(self):
        """Load student class assignments data into table"""
        try:
            self.cursor.execute("""
                SELECT sca.id, s.full_name, s.grade_applied_for,
                       c.class_name, 
                       c.stream,
                       t.term_name, ay.year_name, 
                       sca.status, sca.is_current,
                       sca.assignment_date, sca.notes
                FROM student_class_assignments sca
                JOIN students s ON sca.student_id = s.id
                JOIN classes c ON sca.class_id = c.id
                JOIN terms t ON sca.term_id = t.id
                JOIN academic_years ay ON sca.academic_year_id = ay.id
                ORDER BY s.full_name, c.class_name, c.stream
            """)
            assignments = self.cursor.fetchall()
            
            self.update_assignments_table(assignments)
            
            # Update stats
            if hasattr(self, 'stats_label'):
                total_count = len(assignments)
                current_count = sum(1 for a in assignments if a[8])  # is_current
                self.stats_label.setText(f"Total: {len(assignments)} assignments | Current: {current_count}")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load data: {e}")
            print(f"Load data error: {e}")
            traceback.print_exc()
            
    def update_assignments_table(self, assignments):
        """Update the assignments table with new data"""
        self.table_widget.setRowCount(0)
        self.current_assignments_data = assignments
        
        # Ensure we have the correct number of columns
        if self.table_widget.columnCount() != 7:
            self.table_widget.setColumnCount(7)
            self.table_widget.setHorizontalHeaderLabels(["Student Name", "Grade", "Class/Stream", "Term", "Year", "Status", "Current"])
        
        for row, assignment in enumerate(assignments):
            self.table_widget.insertRow(row)
            
            # Format data for display
            # assignment structure: [id, full_name, grade_applied_for, class_name, stream, term_name, year_name, status, is_current, ...]
            student_name = assignment[1] or "N/A"
            grade = assignment[2] or "N/A"
            class_name = assignment[3] or "N/A"
            stream = assignment[4] or ""
            term = assignment[5] or "N/A"
            year = assignment[6] or "N/A"
            status = assignment[7] or "Active"
            is_current = assignment[8]
            
            # DEBUG: Print what we're working with
            print(f"DEBUG: Class='{class_name}', Stream='{stream}'")
            
            # Create proper Class/Stream display - FIXED
            if stream and stream.strip() and stream != class_name:
                # If stream exists and is different from class name, show "Class Stream"
                class_stream = f"{class_name} {stream}"
                # Check if stream already contains class name to avoid duplicates
                if stream.startswith(class_name):
                    # If stream is like "S1 EAST", just use the stream as-is
                    class_stream = stream
                else:
                    class_stream = f"{class_name} {stream}"
            else:
                # If no stream or stream is same as class name, just show class name
                class_stream = class_name
            
            # Clean up any duplicate class names (e.g., "S1 S1 EAST" -> "S1 EAST")
            if class_stream and class_name and class_stream.startswith(class_name + " " + class_name):
                # Fix patterns like "S1 S1 EAST" -> "S1 EAST"
                class_stream = class_stream.replace(class_name + " " + class_name, class_name, 1)
            
            # Add items to table - ALL 7 COLUMNS
            self.table_widget.setItem(row, 0, QTableWidgetItem(self.truncate_text(student_name, 20)))
            self.table_widget.setItem(row, 1, QTableWidgetItem(self.truncate_text(grade, 10)))
            self.table_widget.setItem(row, 2, QTableWidgetItem(self.truncate_text(class_stream, 15)))
            self.table_widget.setItem(row, 3, QTableWidgetItem(self.truncate_text(term, 10)))
            self.table_widget.setItem(row, 4, QTableWidgetItem(self.truncate_text(year, 10)))
            self.table_widget.setItem(row, 5, QTableWidgetItem(self.truncate_text(status, 15)))
            self.table_widget.setItem(row, 6, QTableWidgetItem("Yes" if is_current else "No"))
        
        # Resize columns to fit content nicely
        self.table_widget.resizeColumnsToContents()
        
        # Make some columns a bit wider for better readability
        self.table_widget.setColumnWidth(0, 200)  # Student Name
        self.table_widget.setColumnWidth(2, 150)  # Class/Stream
        self.table_widget.setColumnWidth(6, 80)   # Current
        
        # Update table info
        if hasattr(self, 'table_info_label'):
            self.table_info_label.setText(f"Showing {len(assignments)} assignment(s) - Select a row to edit")
        
    def truncate_text(self, text, max_length):
        """Truncate text to fit in table cells"""
        if not text:
            return "N/A"
        text_str = str(text)
        return text_str[:max_length] + "..." if len(text_str) > max_length else text_str
        
    def on_table_row_click(self, row, column):
        """Handle row selection from table"""
        try:
            if row < 0 or row >= len(self.current_assignments_data):
                return
                
            # Get assignment data
            assignment_data = self.current_assignments_data[row]
            assignment_id = assignment_data[0]
            
            # Switch to the assignment tab
            self.tab_widget.setCurrentIndex(0)
            
            self.load_assignment_data(assignment_id)
            
        except Exception as e:
            print(f"Error selecting assignment: {e}")
            traceback.print_exc()
            
    def load_assignment_data(self, assignment_id):
        """Load assignment data into form"""
        try:
            self.cursor.execute("""
                SELECT sca.id, sca.student_id, sca.class_id, sca.academic_year_id,
                       sca.term_id, sca.status, sca.is_current, sca.assignment_date,
                       sca.notes, s.full_name, s.grade_applied_for, c.class_name, 
                       c.stream, c.level, t.term_name, ay.year_name
                FROM student_class_assignments sca
                JOIN students s ON sca.student_id = s.id
                JOIN classes c ON sca.class_id = c.id
                JOIN terms t ON sca.term_id = t.id
                JOIN academic_years ay ON sca.academic_year_id = ay.id
                WHERE sca.id = %s
            """, (assignment_id,))
            
            assignment_data = self.cursor.fetchone()
            
            if not assignment_data:
                QMessageBox.critical(self, "Error", "Assignment not found")
                return
            
            self.selected_assignment_id = assignment_id
            
            # Extract data
            level = assignment_data[13]           # c.level
            class_name = assignment_data[11]      # c.class_name
            stream = assignment_data[12]          # c.stream
            student_name = assignment_data[9]     # s.full_name
            student_id = assignment_data[1]       # sca.student_id
            term_name = assignment_data[14]       # t.term_name
            year_name = assignment_data[15]       # ay.year_name
            
            # Set level first (this will trigger the cascade)
            if level:
                index = self.level_dropdown.findText(level)
                if index >= 0:
                    self.level_dropdown.setCurrentIndex(index)
                    self.on_level_selection_changed(level)
            
            # Set class name
            if class_name:
                index = self.class_name_dropdown.findText(class_name)
                if index >= 0:
                    self.class_name_dropdown.setCurrentIndex(index)
                    self.on_class_name_selection_changed(class_name)
            
            # Set stream
            stream_display = stream if stream else "No Stream"
            index = self.stream_dropdown.findText(stream_display)
            if index >= 0:
                self.stream_dropdown.setCurrentIndex(index)
                self.on_stream_selection_changed(stream_display)
            
            # Set student after a short delay to allow dropdowns to populate
            QTimer.singleShot(100, lambda: self._set_student_selection(student_id, student_name))
            
            # Set other dropdown values
            if year_name:
                index = self.academic_year_dropdown.findText(year_name)
                if index >= 0:
                    self.academic_year_dropdown.setCurrentIndex(index)
            if term_name:
                index = self.term_dropdown.findText(term_name)
                if index >= 0:
                    self.term_dropdown.setCurrentIndex(index)
            if assignment_data[5]:  # Status
                index = self.status_dropdown.findText(assignment_data[5])
                if index >= 0:
                    self.status_dropdown.setCurrentIndex(index)
            
            # Set assignment date
            if assignment_data[7]:
                self.assignment_date_entry.setText(str(assignment_data[7]))
            
            # Set notes
            if assignment_data[8]:  # notes
                self.notes_textbox.setPlainText(assignment_data[8])
            
            # Set is_current checkbox
            self.is_current_checkbox.setChecked(assignment_data[6])  # is_current
            
            self.status_label.setText(f"Assignment ID {assignment_id} loaded successfully")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load assignment data: {e}")
            print(f"Load assignment error: {e}")
            traceback.print_exc()
            
    def _set_student_selection(self, student_id, student_name):
        """Helper method to set student selection after dropdown is populated"""
        try:
            if hasattr(self, 'filtered_students_by_class') and self.filtered_students_by_class:
                for student in self.filtered_students_by_class:
                    if student[0] == student_id:
                        # Find the exact text in the dropdown
                        for i in range(self.student_dropdown.count()):
                            if f"(ID: {student_id})" in self.student_dropdown.itemText(i):
                                self.student_dropdown.setCurrentIndex(i)
                                return
                
                # If not found in current dropdown, try to set by name
                self.student_dropdown.setCurrentText(f"{student_name} (ID: {student_id}) [Not in current list]")
            else:
                self.student_dropdown.setCurrentText(f"{student_name} (ID: {student_id}) [Load class first]")
                
        except Exception as e:
            print(f"Error setting student selection: {e}")
            
    def search_assignments(self):
        """Search assignments based on search entry"""
        try:
            search_term = self.search_entry.text().strip().lower()
            
            if not search_term:
                self.load_data()
                return
            
            self.cursor.execute("""
                SELECT sca.id, s.full_name, s.grade_applied_for,
                       c.class_name, 
                       c.stream,
                       t.term_name, ay.year_name, 
                       sca.status, sca.is_current,
                       sca.assignment_date, sca.notes
                FROM student_class_assignments sca
                JOIN students s ON sca.student_id = s.id
                JOIN classes c ON sca.class_id = c.id
                JOIN terms t ON sca.term_id = t.id
                JOIN academic_years ay ON sca.academic_year_id = ay.id
                WHERE LOWER(s.full_name) LIKE %s 
                   OR LOWER(c.class_name) LIKE %s
                   OR LOWER(c.stream) LIKE %s
                   OR LOWER(sca.status) LIKE %s
                   OR LOWER(s.grade_applied_for) LIKE %s
                   OR LOWER(t.term_name) LIKE %s
                   OR LOWER(ay.year_name) LIKE %s
                ORDER BY s.full_name, c.class_name, c.stream
            """, tuple([f'%{search_term}%'] * 7))
            
            assignments = self.cursor.fetchall()
            self.update_assignments_table(assignments)
            
            if hasattr(self, 'stats_label'):
                self.stats_label.setText(f"Search Results: {len(assignments)} assignments found")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Search failed: {e}")
            print(f"Search error: {e}")
            traceback.print_exc()
            
    def open_promotion_popup(self):
        """Open promotion dialog"""
        QMessageBox.information(self, "Info", "Promotion feature will be implemented in the next version")
        
    def open_demotion_popup(self):
        """Open demotion dialog"""
        QMessageBox.information(self, "Info", "Demotion feature will be implemented in the next version")

    def export_to_excel(self):
        """Export student class assignments to Excel with green header style"""
        try:
            # Validation: Check if there's data to export
            if not hasattr(self, 'current_assignments_data') or not self.current_assignments_data:
                QMessageBox.warning(self, "Warning", "No assignment data to export")
                return
    
            # Get school info for the title
            school_info = self.get_school_info()
            
            # Prepare data for export - convert to list of lists
            export_data = []
            for assignment in self.current_assignments_data:
                # assignment structure: [id, full_name, grade_applied_for, class_name, stream, term_name, year_name, status, is_current, ...]
                student_name = assignment[1] or "N/A"
                grade = assignment[2] or "N/A"
                class_name = assignment[3] or "N/A"
                stream = assignment[4] or ""
                term = assignment[5] or "N/A"
                year = assignment[6] or "N/A"
                status = assignment[7] or "Active"
                is_current = assignment[8]
                assignment_date = assignment[9] or "N/A"
                notes = assignment[10] or ""
                
                # Format Class/Stream display
                if stream and stream.strip() and stream != class_name:
                    class_stream = f"{class_name} {stream}"
                else:
                    class_stream = class_name
                
                row_data = [
                    student_name,
                    grade,
                    class_stream,
                    term,
                    year,
                    status,
                    "Yes" if is_current else "No",
                    str(assignment_date),
                    notes
                ]
                export_data.append(row_data)
    
            # Define headers
            headers = [
                'Student Name', 'Grade', 'Class/Stream', 'Term', 'Academic Year', 
                'Status', 'Current', 'Assignment Date', 'Notes'
            ]
            
            # Create title
            title = f"{school_info['name']} - STUDENT CLASS ASSIGNMENTS"
            
            # Use shared export method
            self.export_with_green_header(
                data=export_data,
                headers=headers,
                filename_prefix="student_assignments_export",
                title=title
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export to Excel: {e}")
            traceback.print_exc()
        
    def export_to_pdf(self):
        """Export student class assignments to PDF"""
        try:
            # Validation: Check if there's data to export
            if not hasattr(self, 'current_assignments_data') or not self.current_assignments_data:
                QMessageBox.warning(self, "Warning", "No assignment data to export")
                return

            # File save dialog
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Student Assignments Report As",
                f"student_class_assignments_{timestamp}.pdf",
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return

            # Custom PDF class with footer
            class PDF(FPDF):
                def __init__(self, *args, **kwargs):
                    super().__init__(*args, **kwargs)
                    self.school_name = "WINSPIRE LEARNING HUB"
                    
                def footer(self):
                    """Add footer with page number and school name"""
                    self.set_y(-15)
                    self.set_font("Arial", 'I', 8)
                    self.cell(0, 10, f'{self.school_name} - Page {self.page_no()}', 0, 0, 'C')

            # Initialize PDF
            pdf = PDF(orientation='P', unit='mm', format='A4')
            pdf.set_margins(15, 20, 15)
            pdf.add_page()
            pdf.set_auto_page_break(True, 25)

            # === SCHOOL HEADER SECTION ===
            self._add_school_header(pdf)
            
            # === REPORT SUMMARY ===
            self._add_report_summary(pdf)
            
            # === ASSIGNMENTS TABLE ===
            self._add_assignments_table(pdf)
            
            # === REPORT FOOTER ===
            self._add_report_footer(pdf)

            # Save PDF
            pdf.output(file_path)
            
            # Open PDF automatically
            self._open_pdf(file_path)

            # Success message
            QMessageBox.information(
                self,
                "Export Successful",
                f"Student assignments report exported successfully!\n\n"
                f"Total Records: {len(self.current_assignments_data)}\n"
                f"File Location: {os.path.dirname(file_path)}\n"
                f"Filename: {os.path.basename(file_path)}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export PDF:\n{e}")
            traceback.print_exc()
            
    def _add_school_header(self, pdf):
        """Add school header with logo beside the school information"""
        # School configuration
        school_config = {
            'name': "WINSPIRE LEARNING HUB",
            'address': "P.O.Box 12345",
            'location': "Nairobi, Kenya",
            'tel': "Tel: +254 700 000000",
            'email': "info@winspirelearning.ac.ke",
            'logo_path': "static/images/logo.png"
        }

        # Add school information
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, school_config['name'], 0, 1, 'C')
        pdf.set_font("Arial", '', 12)
        pdf.cell(0, 8, school_config['address'], 0, 1, 'C')
        pdf.cell(0, 8, school_config['location'], 0, 1, 'C')
        pdf.cell(0, 8, school_config['tel'], 0, 1, 'C')
        pdf.ln(10)
        
        # Report title
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(0, 10, "STUDENT CLASS ASSIGNMENTS REPORT", 0, 1, 'C')
        pdf.set_font("Arial", 'I', 10)
        pdf.cell(0, 8, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 1, 'C')
        pdf.ln(10)
        
    def _add_report_summary(self, pdf):
        """Add summary information about the assignments"""
        # Calculate summary statistics - UPDATED indices
        total_assignments = len(self.current_assignments_data)
        active_assignments = sum(1 for assignment in self.current_assignments_data if assignment[8])  # is_current is now at index 8
        
        # Summary section
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "REPORT SUMMARY", 0, 1)
        pdf.ln(2)
    
        # Summary table
        pdf.set_font("Arial", '', 10)
        
        summary_data = [
            ("Total Assignment Records:", str(total_assignments)),
            ("Current Assignments:", str(active_assignments)),
            ("Inactive Assignments:", str(total_assignments - active_assignments)),
        ]
    
        for label, value in summary_data:
            pdf.cell(80, 6, label)
            pdf.cell(40, 6, value, 0, 1)
        
        pdf.ln(8)
        
    def _add_assignments_table(self, pdf):
        """Add the main assignments table with combined Class/Stream (without Grade column)"""
        # Table headers
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(0, 8, "ASSIGNMENT DETAILS", 0, 1)
        pdf.ln(3)
    
        # Define headers and column widths - REMOVED Grade column
        headers = ["Student Name", "Class", "Stream", "Term", "Year", "Status", "Current"]
        col_widths = [45, 20, 20, 20, 20, 15, 12]  # Adjusted widths
    
        # Table headers with styling
        pdf.set_fill_color(70, 130, 180)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", 'B', 9)
        
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 8, header, 1, 0, 'C', True)
        pdf.ln()
    
        # Reset text color for data rows
        pdf.set_text_color(0, 0, 0)
    
        # Table data
        pdf.set_font("Arial", '', 8)
        row_height = 6
        
        for i, assignment in enumerate(self.current_assignments_data):
            # Check if we need a new page
            if pdf.get_y() + row_height > pdf.h - 25:
                pdf.add_page()
                # Repeat headers on new page
                pdf.set_fill_color(70, 130, 180)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font("Arial", 'B', 9)
                for j, header in enumerate(headers):
                    pdf.cell(col_widths[j], 8, header, 1, 0, 'C', True)
                pdf.ln()
                pdf.set_text_color(0, 0, 0)
                pdf.set_font("Arial", '', 8)
    
            # Format data for display
            def truncate_text(text, max_length):
                text_str = str(text) if text is not None else "N/A"
                return text_str[:max_length] + "..." if len(text_str) > max_length else text_str
    
            # assignment structure: [id, full_name, grade_applied_for, class_name, stream, term_name, year_name, status, is_current, ...]
            student_name = assignment[1] or "N/A"
            class_name = assignment[3] or "N/A"  # Class name
            stream = assignment[4] or ""         # Stream (might be None or empty)
            term = assignment[5] or "N/A"
            year = assignment[6] or "N/A"
            status = assignment[7] or "Active"
            is_current = assignment[8]
    
            # Clean up stream display - if stream is same as class name or empty, show "N/A"
            if not stream or stream == class_name or stream == "N/A":
                stream_display = "N/A"
            else:
                stream_display = stream
    
            row_data = [
                truncate_text(student_name, 40),  # Student name
                truncate_text(class_name, 15),    # Class (removed grade)
                truncate_text(stream_display, 15), # Stream
                truncate_text(term, 15),          # Term
                truncate_text(year, 15),          # Year
                truncate_text(status, 12),        # Status
                "Yes" if is_current else "No"     # Current
            ]
            
            # Alternate row colors
            if i % 2 == 0:
                pdf.set_fill_color(248, 249, 250)
                fill = True
            else:
                pdf.set_fill_color(255, 255, 255)
                fill = True
            
            # Add row data
            for j, cell_data in enumerate(row_data):
                align = 'C' if j in [6] else 'L'  # Center align only Current column
                pdf.cell(col_widths[j], row_height, str(cell_data), 1, 0, align, fill)
            pdf.ln()
            
    def _add_report_footer(self, pdf):
        """Add footer section with report information"""
        pdf.ln(10)
        
        # Footer text
        pdf.set_font("Arial", 'I', 9)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(0, 6, "This report contains confidential student information.", 0, 1, 'C')
        pdf.cell(0, 6, "Please handle with appropriate care and in accordance with data protection policies.", 0, 1, 'C')
        pdf.ln(3)
        
        # Report generation info
        pdf.set_font("Arial", '', 8)
        pdf.cell(0, 5, f"Report generated by Winspire Learning Hub Management System", 0, 1, 'C')
        
        # Generation timestamp
        pdf.set_font("Arial", 'I', 8)
        pdf.cell(0, 5, f"Generated on: {datetime.now().strftime('%A, %B %d, %Y at %I:%M %p')}", 0, 1, 'C')
        
    def _open_pdf(self, path):
        """Open PDF file with the system's default viewer"""
        try:
            system = platform.system()
            
            if system == 'Windows':
                os.startfile(path)
            elif system == 'Darwin':  # macOS
                subprocess.run(['open', path], check=True)
            else:  # Linux and others
                try:
                    subprocess.run(['xdg-open', path], check=True)
                except (subprocess.CalledProcessError, FileNotFoundError):
                    # Fallback options for Linux
                    try:
                        subprocess.run(['gnome-open', path], check=True)
                    except (subprocess.CalledProcessError, FileNotFoundError):
                        try:
                            subprocess.run(['kde-open', path], check=True)
                        except (subprocess.CalledProcessError, FileNotFoundError):
                            try:
                                subprocess.run(['evince', path], check=True)  # GNOME PDF viewer
                            except (subprocess.CalledProcessError, FileNotFoundError):
                                try:
                                    subprocess.run(['okular', path], check=True)  # KDE PDF viewer
                                except (subprocess.CalledProcessError, FileNotFoundError):
                                    raise Exception("No suitable PDF viewer found")
            
        except Exception as e:
            print(f"Failed to open PDF automatically: {e}")
            # Show a message with the file location instead of error
            QMessageBox.information(
                self,
                "PDF Saved Successfully", 
                f"PDF report has been saved successfully!\n\n"
                f"Location: {path}\n\n"
                f"The file couldn't open automatically, but you can open it manually from the saved location."
            )

    def closeEvent(self, event):
        """Cleanup when the form is closed"""
        try:
            if hasattr(self, 'cursor') and self.cursor:
                self.cursor.close()
            if hasattr(self, 'db_connection') and self.db_connection:
                self.db_connection.close()
        except Exception as e:
            print(f"Error closing database connection: {e}")
        
        event.accept()

    def __del__(self):
        """Close database connection when object is destroyed"""
        try:
            if hasattr(self, 'cursor'):
                self.cursor.close()
            if hasattr(self, 'db_connection'):
                self.db_connection.close()
        except:
            pass