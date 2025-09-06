# ui/departments_form.py
import sys
import os
import csv
from datetime import datetime
from typing import Optional
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QMessageBox, QFileDialog, QTabWidget, QGroupBox, QComboBox,
    QTextEdit, QCheckBox, QMenu, QGridLayout
)
from PySide6.QtGui import QFont, QPalette, QIcon, QAction
from PySide6.QtCore import Qt, Signal, QSize

from utils.permissions import has_permission
from ui.audit_base_form import AuditBaseForm
from models.models import get_db_connection
from utils.pdf_utils import view_pdf  # Assuming you have this from teachers_form
import csv
import openpyxl  # For Excel support
from openpyxl import load_workbook
from io import StringIO

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class DepartmentsForm(AuditBaseForm):
    """
    Comprehensive Departments Management with:
    - Audit logging
    - CSV/PDF export
    - staff count
    - Inactive filter
    - Two-tab UI (Form + Data)
    """
    department_selected = Signal(int)

    def __init__(self, parent=None, user_session=None):
        super().__init__(parent, user_session)
        self.user_session = user_session
        self.current_department_id = None
        self.schools_map = {}
        self.teachers_map = {}
        self.all_teachers = []

        try:
            self.db_connection = get_db_connection()
            self.cursor = self.db_connection.cursor(dictionary=True)
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Failed to connect to database: {e}")
            return

        # Initialize UI components first
        self.setup_ui()
        
        # Load data
        self.load_schools()
        self.load_teachers()
        self.load_departments()
        
        # Apply permissions after UI is fully setup
        self.apply_permissions()

    def setup_ui(self):
        """Set up the tabbed UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)

        self.tab_widget = QTabWidget()
        self.department_form_tab = QWidget()
        self.department_data_tab = QWidget()

        self.tab_widget.addTab(self.department_form_tab, "Department Form")
        self.tab_widget.addTab(self.department_data_tab, "Departments Data")

        layout.addWidget(self.tab_widget)

        self.setup_department_form_tab()
        self.setup_department_data_tab()

    
    def setup_department_form_tab(self):
        """Form tab: Add/Edit Department"""
        layout = QVBoxLayout(self.department_form_tab)
        layout.setSpacing(10)
    
        # Title
        title = QLabel("Department Management")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setFont(self.fonts['section'])
        layout.addWidget(title)
    
        # === FORM GROUP ===
        form_group = QGroupBox("Department Details")
        form_layout = QGridLayout(form_group)
    
        # Department Name
        form_layout.addWidget(QLabel("Department Name:"), 0, 0)
        self.name_entry = QLineEdit()
        form_layout.addWidget(self.name_entry, 0, 1)
    
        # School
        form_layout.addWidget(QLabel("School:"), 0, 2)
        self.school_combo = QComboBox()
        self.school_combo.currentTextChanged.connect(self.on_school_selected)
        form_layout.addWidget(self.school_combo, 0, 3)
    
        # Head
        form_layout.addWidget(QLabel("Department Head:"), 1, 0)
        self.head_combo = QComboBox()
        self.head_combo.setEditable(True)
        self.head_combo.setPlaceholderText("Search teacher...")
        self.head_combo.editTextChanged.connect(self.on_head_search)
        self.head_combo.currentTextChanged.connect(self.on_head_selection_changed)
        form_layout.addWidget(self.head_combo, 1, 1)
    
        # Description
        form_layout.addWidget(QLabel("Description:"), 1, 2)
        self.description_box = QTextEdit()
        self.description_box.setMaximumHeight(60)
        form_layout.addWidget(self.description_box, 1, 3)
    
        layout.addWidget(form_group)
    
        # === ACTION BUTTONS ===
        btn_layout = QHBoxLayout()
    
        self.add_btn = QPushButton("Add Department")
        self.add_btn.setProperty("class", "success")
        self.add_btn.setIcon(QIcon("static/icons/add.png"))
        self.add_btn.setIconSize(QSize(18, 18))
        self.add_btn.clicked.connect(self.add_department)
        btn_layout.addWidget(self.add_btn)
    
        self.update_btn = QPushButton("Update")
        self.update_btn.setProperty("class", "primary")
        self.update_btn.setIcon(QIcon("static/icons/update.png"))
        self.update_btn.setIconSize(QSize(18, 18))
        self.update_btn.clicked.connect(self.update_department)
        btn_layout.addWidget(self.update_btn)
    
        self.delete_btn = QPushButton("Delete")
        self.delete_btn.setProperty("class", "danger")
        self.delete_btn.setIcon(QIcon("static/icons/delete.png"))
        self.delete_btn.setIconSize(QSize(18, 18))
        self.delete_btn.clicked.connect(self.delete_department)
        btn_layout.addWidget(self.delete_btn)
    
        self.clear_btn = QPushButton("Clear")
        self.clear_btn.setProperty("class", "secondary")
        self.clear_btn.setIcon(QIcon("static/icons/clear.png"))
        self.clear_btn.setIconSize(QSize(18, 18))
        self.clear_btn.clicked.connect(self.clear_fields)
        btn_layout.addWidget(self.clear_btn)
    
        layout.addLayout(btn_layout)

    
    def setup_department_data_tab(self):
        """Data tab: View/Search Departments - with styled actions"""
        layout = QVBoxLayout(self.department_data_tab)
        layout.setSpacing(10)
    
        # ---------------- Filters Section ----------------
        filter_group = QGroupBox("Filters")
        filter_layout = QGridLayout(filter_group)
    
        # Search input
        search_label = QLabel("Search:")
        search_label.setProperty("class", "field-label")
        filter_layout.addWidget(search_label, 0, 0)
    
        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText("Search by name...")
        self.search_entry.textChanged.connect(self.load_departments)
        filter_layout.addWidget(self.search_entry, 0, 1)
    
        # Active/Inactive toggle
        self.active_filter = QCheckBox("Show Inactive Departments")
        self.active_filter.stateChanged.connect(self.load_departments)
        filter_layout.addWidget(self.active_filter, 0, 2)
    
        # Filter button with icon
        self.filter_btn = QPushButton("Filter")
        self.filter_btn.setProperty("class", "primary")
        self.filter_btn.setIcon(QIcon("static/icons/filter.png"))
        self.filter_btn.setIconSize(QSize(18, 18))
        self.filter_btn.clicked.connect(self.load_departments)
        filter_layout.addWidget(self.filter_btn, 0, 3)
    
        layout.addWidget(filter_group)

        # === ACTION BUTTONS ===
        action_layout = QHBoxLayout()
    
        self.export_csv_btn = QPushButton("Export to Excel")
        self.export_csv_btn.setProperty("class", "success")
        self.export_csv_btn.setIcon(QIcon("static/icons/export.png"))
        self.export_csv_btn.setIconSize(QSize(18, 18))
        self.export_csv_btn.clicked.connect(self.export_to_csv)
        action_layout.addWidget(self.export_csv_btn)
    
        self.export_pdf_btn = QPushButton("Export to PDF")
        self.export_pdf_btn.setProperty("class", "info")
        self.export_pdf_btn.setIcon(QIcon("static/icons/export.png"))
        self.export_pdf_btn.setIconSize(QSize(18, 18))
        self.export_pdf_btn.clicked.connect(self.export_to_pdf)
        action_layout.addWidget(self.export_pdf_btn)
    
        self.export_summary_btn = QPushButton("Export Summary")
        self.export_summary_btn.setProperty("class", "warning")
        self.export_summary_btn.setIcon(QIcon("static/icons/summary.png"))
        self.export_summary_btn.setIconSize(QSize(18, 18))
        self.export_summary_btn.clicked.connect(self.export_department_summary)
        action_layout.addWidget(self.export_summary_btn)
    
        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.setProperty("class", "secondary")
        self.refresh_btn.setIcon(QIcon("static/icons/refresh.png"))
        self.refresh_btn.setIconSize(QSize(18, 18))
        self.refresh_btn.clicked.connect(self.force_refresh_data)
        action_layout.addWidget(self.refresh_btn)
    
        action_layout.addStretch()
        layout.addLayout(action_layout)
    
        # === TABLE ===
        table_group = QGroupBox("Departments List")
        table_layout = QVBoxLayout(table_group)
    
        self.departments_table = QTableWidget()
        self.departments_table.setColumnCount(8)
        self.departments_table.setHorizontalHeaderLabels([
            "ID", "School", "Department Name", "Head", "Staff Count",
            "Description", "Created At", "Active"
        ])
        self.departments_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.departments_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.departments_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.departments_table.setSortingEnabled(True)
        self.departments_table.itemSelectionChanged.connect(self.on_department_select)
        self.departments_table.itemDoubleClicked.connect(self.edit_selected_department)
    
        table_layout.addWidget(self.departments_table)
        layout.addWidget(table_group)


    def on_school_selected(self):
        """Reload teachers when school changes"""
        self.load_teachers()
        # Clear the head combo when school changes
        self.head_combo.blockSignals(True)
        self.head_combo.clearEditText()
        self.head_combo.setCurrentIndex(-1)
        self.head_combo.blockSignals(False)

    def load_schools(self):
        """Load schools into combo box"""
        try:
            self.cursor.execute("SELECT id, school_name FROM schools ORDER BY school_name")
            schools = self.cursor.fetchall()
            self.schools_map = {row['school_name']: row['id'] for row in schools}
            self.school_combo.clear()
            self.school_combo.addItem("All Schools", None)
            for name in self.schools_map.keys():
                self.school_combo.addItem(name, self.schools_map[name])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load schools: {e}")

    def load_teachers(self):
        """Load teachers based on selected school"""
        try:
            school_id = self.school_combo.currentData()
            query = "SELECT id, full_name FROM teachers WHERE is_active = 1"
            params = []

            if school_id:
                query += " AND school_id = %s"
                params.append(school_id)

            query += " ORDER BY full_name"
            self.cursor.execute(query, params)
            teachers = self.cursor.fetchall()

            self.teachers_map = {t['full_name']: t['id'] for t in teachers}
            self.all_teachers = [t['full_name'] for t in teachers]

            # Store current selection before clearing
            current_text = self.head_combo.currentText()
            
            # Block signals to prevent unwanted triggers
            self.head_combo.blockSignals(True)
            self.head_combo.clear()
            self.head_combo.addItems(self.all_teachers)
            
            # Restore selection if it exists in new list
            if current_text in self.all_teachers:
                self.head_combo.setCurrentText(current_text)
            
            self.head_combo.blockSignals(False)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load teachers: {e}")

    def on_head_search(self, text):
        """Filter head combo based on search text"""
        if not text.strip():
            # If empty, show all teachers
            self.head_combo.blockSignals(True)
            self.head_combo.clear()
            self.head_combo.addItems(self.all_teachers)
            self.head_combo.blockSignals(False)
            return
        
        # Filter teachers based on search text
        filtered = [t for t in self.all_teachers if text.lower() in t.lower()]
        
        # Block signals to prevent recursion
        self.head_combo.blockSignals(True)
        self.head_combo.clear()
        
        if filtered:
            self.head_combo.addItems(filtered)
            # If exact match exists, select it
            if text in filtered:
                self.head_combo.setCurrentText(text)
        else:
            # Show all if no matches found
            self.head_combo.addItems(self.all_teachers)
        
        self.head_combo.blockSignals(False)
        
        # Show dropdown with filtered results
        self.head_combo.showPopup()

    def on_head_selection_changed(self, text):
        """Handle when head selection changes"""
        if text in self.teachers_map:
            print(f"Selected department head: {text}")
    
    def load_departments(self):
        """Load departments with filters"""
        try:
            # Force commit any pending transactions and refresh connection
            self.db_connection.commit()
            
            # Optional: Reconnect cursor for fresh data
            if hasattr(self, 'cursor'):
                self.cursor.close()
                self.cursor = self.db_connection.cursor(dictionary=True)
            
            search_term = self.search_entry.text().strip()
            show_inactive = self.active_filter.isChecked()

            query = """
                SELECT 
                    d.id, s.school_name, d.department_name, COALESCE(t.full_name, 'N/A') as head_name,
                    d.description, d.created_at, d.is_active,
                    (SELECT COUNT(*) FROM teachers t WHERE t.department_id = d.id AND t.is_active = 1) as teacher_count
                FROM departments d
                LEFT JOIN schools s ON d.school_id = s.id
                LEFT JOIN teachers t ON d.department_head_id = t.id
            """
            params = []

            if not show_inactive:
                query += " WHERE d.is_active = 1"
            else:
                query += " WHERE 1=1"

            if search_term:
                query += " AND LOWER(d.department_name) LIKE %s"
                params.append(f"%{search_term.lower()}%")

            query += " ORDER BY s.school_name, d.department_name"

            self.cursor.execute(query, params)
            rows = self.cursor.fetchall()

            self.departments_table.setRowCount(len(rows))
            for row_idx, row in enumerate(rows):
                items = [
                    str(row['id']),
                    row['school_name'] or 'N/A',
                    row['department_name'] or 'N/A',
                    row['head_name'] or 'N/A',
                    str(row['teacher_count']),
                    row['description'] or 'N/A',
                    row['created_at'].strftime('%Y-%m-%d %H:%M') if row['created_at'] else 'N/A',
                    'Yes' if row['is_active'] else 'No'
                ]
                for col_idx, item in enumerate(items):
                    table_item = QTableWidgetItem(item)
                    table_item.setData(Qt.UserRole, row['id'])
                    if col_idx == 4:  # staff count → right align
                        table_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.departments_table.setItem(row_idx, col_idx, table_item)

            self.departments_table.resizeColumnsToContents()
            self.clear_selection()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load departments: {e}")

    def on_department_select(self):
        """Handle row selection"""
        selected = self.departments_table.selectedItems()
        if not selected:
            return
        row = selected[0].row()
        item = self.departments_table.item(row, 0)
        if item:
            dept_id = int(item.data(Qt.UserRole))
            self.current_department_id = dept_id
            self.load_department_data(dept_id)

    def edit_selected_department(self):
        """Switch to form tab on double-click"""
        if self.current_department_id:
            self.tab_widget.setCurrentIndex(0)

    def load_department_data(self, dept_id):
        """Load department data into form"""
        try:
            query = """
                SELECT d.department_name, d.school_id, d.department_head_id, d.description, d.is_active
                FROM departments d
                WHERE d.id = %s
            """
            self.cursor.execute(query, (dept_id,))
            row = self.cursor.fetchone()
            if not row:
                return

            self.name_entry.setText(row['department_name'])
            self.description_box.setPlainText(row['description'] or "")

            # School
            idx = self.school_combo.findData(row['school_id'])
            if idx >= 0:
                self.school_combo.setCurrentIndex(idx)
                # Reload teachers for selected school
                self.load_teachers()

            # Head - Set after teachers are loaded
            if row['department_head_id']:
                head_name = next(
                    (name for name, tid in self.teachers_map.items() if tid == row['department_head_id']), 
                    ""
                )
                if head_name:
                    self.head_combo.setCurrentText(head_name)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load department: {e}")

    def add_department(self):
        """Add new department"""
        if not has_permission(self.user_session, "create_department"):
            QMessageBox.warning(self, "Permission Denied", "You don't have permission to add departments.")
            return

        name = self.name_entry.text().strip()
        school_id = self.school_combo.currentData()
        head_text = self.head_combo.currentText().strip()
        description = self.description_box.toPlainText().strip()

        if not name or not school_id:
            QMessageBox.warning(self, "Validation Error", "Department name and school are required.")
            return

        head_id = self.teachers_map.get(head_text)

        try:
            # Check duplicates
            self.cursor.execute(
                "SELECT id FROM departments WHERE school_id = %s AND department_name = %s AND is_active = 1",
                (school_id, name)
            )
            if self.cursor.fetchone():
                QMessageBox.warning(self, "Duplicate", "A department with this name already exists.")
                return

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cursor.execute("""
                INSERT INTO departments (
                    school_id, department_name, department_head_id, description, 
                    is_active, created_at, updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (school_id, name, head_id, description, 1, now, now))

            dept_id = self.cursor.lastrowid
            self.db_connection.commit()

            # Audit Log
            self.log_audit_action(
                action="CREATE",
                table_name="departments",
                record_id=dept_id,
                description=f"Created department '{name}' in {self.school_combo.currentText()}"
            )

            QMessageBox.information(self, "Success", "Department added successfully!")
            self.clear_fields()
            self.load_departments()

        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Failed to add department: {e}")

    def update_department(self):
        """Update existing department"""
        if not has_permission(self.user_session, "edit_department"):
            QMessageBox.warning(self, "Permission Denied", "You don't have permission to edit departments.")
            return

        if not self.current_department_id:
            QMessageBox.warning(self, "Error", "No department selected.")
            return

        name = self.name_entry.text().strip()
        school_id = self.school_combo.currentData()
        head_text = self.head_combo.currentText().strip()
        description = self.description_box.toPlainText().strip()

        if not name or not school_id:
            QMessageBox.warning(self, "Validation Error", "Department name and school are required.")
            return

        head_id = self.teachers_map.get(head_text)

        try:
            # Check duplicates
            self.cursor.execute(
                "SELECT id FROM departments WHERE school_id = %s AND department_name = %s AND id != %s AND is_active = 1",
                (school_id, name, self.current_department_id)
            )
            if self.cursor.fetchone():
                QMessageBox.warning(self, "Duplicate", "Another department with this name exists in the school.")
                return

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cursor.execute("""
                UPDATE departments SET
                    school_id = %s, department_name = %s, department_head_id = %s,
                    description = %s, updated_at = %s
                WHERE id = %s
            """, (school_id, name, head_id, description, now, self.current_department_id))

            self.db_connection.commit()

            # Audit Log
            self.log_audit_action(
                action="UPDATE",
                table_name="departments",
                record_id=self.current_department_id,
                description=f"Updated department '{name}' (ID: {self.current_department_id})"
            )

            QMessageBox.information(self, "Success", "Department updated successfully!")
            self.clear_fields()
            self.load_departments()

        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Failed to update department: {e}")

    def delete_department(self):
        """Soft delete department"""
        if not has_permission(self.user_session, "delete_department"):
            QMessageBox.warning(self, "Permission Denied", "You don't have permission to delete departments.")
            return

        if not self.current_department_id:
            QMessageBox.warning(self, "Error", "No department selected.")
            return

        reply = QMessageBox.question(
            self, "Confirm Delete",
            "Deactivate this department? Teachers will still be assigned.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        try:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cursor.execute(
                "UPDATE departments SET is_active = 0, updated_at = %s WHERE id = %s",
                (now, self.current_department_id)
            )
            self.db_connection.commit()

            # Audit Log
            self.log_audit_action(
                action="DELETE",
                table_name="departments",
                record_id=self.current_department_id,
                description=f"Deactivated department (ID: {self.current_department_id})"
            )

            QMessageBox.information(self, "Success", "Department deactivated.")
            self.clear_fields()
            self.load_departments()

        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Failed to delete department: {e}")

    def clear_fields(self):
        """Clear all form fields"""
        self.current_department_id = None
        self.name_entry.clear()
        self.school_combo.setCurrentIndex(0)
        
        # Clear head combo properly
        self.head_combo.blockSignals(True)
        self.head_combo.clearEditText()
        self.head_combo.setCurrentIndex(-1)
        self.head_combo.blockSignals(False)
        
        self.description_box.clear()

    def clear_selection(self):
        """Clear selection and form"""
        self.clear_fields()
        self.departments_table.clearSelection()

    def apply_permissions(self):
        """Apply UI permissions"""
        # Safety check to ensure buttons exist
        if not hasattr(self, 'add_btn') or not hasattr(self, 'update_btn') or not hasattr(self, 'delete_btn'):
            print("Warning: UI buttons not fully initialized, skipping permissions")
            return
            
        can_create = has_permission(self.user_session, "create_department")
        can_edit = has_permission(self.user_session, "edit_department")
        can_delete = has_permission(self.user_session, "delete_department")

        self.add_btn.setEnabled(can_create)
        self.update_btn.setEnabled(can_edit)
        self.delete_btn.setEnabled(can_delete)

        self.add_btn.setToolTip("Add Department" if can_create else "Permission required")
        self.update_btn.setToolTip("Update Department" if can_edit else "Permission required")
        self.delete_btn.setToolTip("Delete Department" if can_delete else "Permission required")

    def export_to_csv(self):
        """Export departments to CSV with enhanced formatting"""
        try:
            # Get the data
            self.cursor.execute("""
                SELECT 
                    s.school_name, d.department_name, COALESCE(t.full_name, 'N/A') as head,
                    d.description, COUNT(te.id) as teacher_count, d.created_at, d.is_active
                FROM departments d
                LEFT JOIN schools s ON d.school_id = s.id
                LEFT JOIN teachers t ON d.department_head_id = t.id
                LEFT JOIN teachers te ON te.department_id = d.id AND te.is_active = 1
                GROUP BY d.id
                ORDER BY s.school_name, d.department_name
            """)
            rows = self.cursor.fetchall()
    
            if not rows:
                QMessageBox.information(self, "No Data", "No departments found to export.")
                return
    
            # Prepare data for CSV
            headers = [
                "School", "Department Name", "Department Head", "Description",
                "Staff Count", "Created Date", "Status"
            ]
            
            csv_data = []
            for row in rows:
                csv_data.append([
                    row['school_name'] or 'N/A',
                    row['department_name'] or 'N/A',
                    row['head'] or 'N/A',
                    row['description'] or 'N/A',
                    str(row['teacher_count']),
                    row['created_at'].strftime('%Y-%m-%d %H:%M') if row['created_at'] else 'N/A',
                    'Active' if row['is_active'] else 'Inactive'
                ])
    
            # Use the enhanced export method for Excel
            self.export_with_green_header(
                data=csv_data,
                headers=headers,
                filename_prefix="departments_export",
                title="DEPARTMENTS REPORT"
            )
    
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export departments: {e}")
    
    def export_to_pdf(self):
        """Generate clean Departments Report PDF with left-aligned school header and centered text"""
        try:
            from fpdf import FPDF
            from datetime import datetime
            import os
            import io
    
            # Ensure DB connection
            self._ensure_connection()
    
            def safe_encode(text):
                """Safely encode text for FPDF (latin-1 compatible)"""
                if not text:
                    return "N/A"
                replacements = {
                    '\u2013': '-', '\u2014': '--', '\u2018': "'", '\u2019': "'",
                    '\u201c': '"', '\u201d': '"', '\u2026': '...', '\u25cf': '•'
                }
                result = str(text)
                for a, b in replacements.items():
                    result = result.replace(a, b)
                return result.encode('latin-1', 'replace').decode('latin-1')
    
            # Fetch department data
            query = """
                SELECT 
                    s.school_name, s.address, s.phone, s.email, s.logo_path,
                    d.department_name, COALESCE(t.full_name, 'N/A') as head,
                    d.description, COUNT(te.id) as teacher_count, d.is_active
                FROM departments d
                LEFT JOIN schools s ON d.school_id = s.id
                LEFT JOIN teachers t ON d.department_head_id = t.id
                LEFT JOIN teachers te ON te.department_id = d.id AND te.is_active = 1
                GROUP BY d.id
                ORDER BY s.school_name, d.department_name
            """
            self.cursor.execute(query)
            rows = self.cursor.fetchall()
    
            if not rows:
                QMessageBox.information(self, "No Data", "No departments found to export.")
                return
    
            # Extract school info
            row0 = rows[0]
            school_name = row0['school_name'] or "CBCentra School"
            school_address = row0['address'] or "P.O. Box 12345"
            school_phone = row0['phone'] or "Tel: +254 700 000000"
            school_email = row0['email'] or "info@cbcentra.edu"
            default_logo = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                "static", "images", "logo.png"
            )
            school_logo = row0['logo_path'] or default_logo
    
            # Summary data
            total_departments = len(rows)
            active_departments = sum(1 for r in rows if r['is_active'])
            total_staff = sum(r['teacher_count'] for r in rows)
    
            # Get user name
            user_name = "Unknown"
            if self.user_session:
                user_name = self.user_session.get('full_name', 'Unknown')
    
            # Define PDF class
            class DeptPDF(FPDF):
                def __init__(self):
                    super().__init__(orientation='P', unit='mm', format='A4')
                    self.set_margins(15, 15, 15)  # 1.5 cm
                    self.set_auto_page_break(auto=True, margin=15)
    
                def header(self):
                    # === Logo on far left ===
                    if os.path.exists(school_logo):
                        try:
                            self.image(school_logo, 15, 10, 25)  # x=15, y=10, w=25
                        except Exception as e:
                            print(f"Logo load error: {e}")
    
                    # === School Info: Start at x=16 for more space ===
                    x = 16  # Pushed 50mm from logo start (=16)
                    y = 10
                    self.set_xy(x, y)
                    self.set_font("Arial", "B", 14)
                    self.cell(0, 6, safe_encode(school_name), ln=True, align='C')  # Centered
                    self.set_x(x)
                    self.set_font("Arial", "", 10)
                    self.cell(0, 4, safe_encode(school_address), ln=True, align='C')
                    self.set_x(x)
                    self.cell(0, 4, safe_encode(school_phone), ln=True, align='C')
                    self.set_x(x)
                    self.cell(0, 4, safe_encode(school_email), ln=True, align='C')
    
                    # === Report Title - Perfectly Centered on Page ===
                    self.ln(10)
                    self.set_font("Arial", "B", 16)
                    self.set_text_color(0, 0, 0)
                    self.cell(0, 8, "DEPARTMENTS REPORT", 0, 1, "C")  # Full page center
    
                    # === Generated Info - Also Centered ===
                    self.set_font("Arial", "I", 10)
                    self.cell(0, 6, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 1, "C")
                    self.cell(0, 6, f"Generated by: {user_name}", 0, 1, "C")
                    self.ln(5)
    
                def footer(self):
                    self.set_y(-15)
                    self.set_font("Arial", "I", 8)
                    self.set_text_color(128, 128, 128)
                    self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "C")
    
                def section_header(self, title):
                    """Plain section header – no background color"""
                    self.ln(3)
                    self.set_font("Arial", "B", 12)
                    self.set_text_color(0, 0, 0)
                    self.cell(0, 8, safe_encode(title), 0, 1, "L")
                    self.ln(2)
    
            # Generate PDF
            pdf = DeptPDF()
            pdf.add_page()
    
            # === SUMMARY ===
            pdf.section_header("Summary")
            pdf.set_font("Arial", "", 10)
            col_w = [95, 95]
    
            pdf.cell(col_w[0], 6, "Total Departments:", 1, 0, "L")
            pdf.cell(col_w[1], 6, str(total_departments), 1, 1, "L")
    
            pdf.cell(col_w[0], 6, "Active Departments:", 1, 0, "L")
            pdf.cell(col_w[1], 6, str(active_departments), 1, 1, "L")
    
            pdf.cell(col_w[0], 6, "Total Staff Members:", 1, 0, "L")
            pdf.cell(col_w[1], 6, str(total_staff), 1, 1, "L")
            pdf.ln(5)
    
            # === DEPARTMENTS LIST ===
            pdf.section_header("Departments List")
    
            # Table header with #e6edf3
            pdf.set_fill_color(230, 237, 243)  # #e6edf3
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", "B", 9)
            col_widths = [50, 45, 45, 30, 20]
            headers = ["Department", "Head", "Staff", "Status", "Active"]
    
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, safe_encode(h), border=1, align="C", fill=True)
            pdf.ln()
    
            # Table rows
            pdf.set_font("Arial", "", 8)
            pdf.set_text_color(0, 0, 0)
            for i, row in enumerate(rows):
                dept = safe_encode(row['department_name'] or 'N/A')
                head = safe_encode(row['head'] or 'N/A')
                count = str(row['teacher_count'])
                status = 'Active' if row['is_active'] else 'Inactive'
                active = 'Yes' if row['is_active'] else 'No'
    
                dept = dept[:25] + "..." if len(dept) > 25 else dept
                head = head[:25] + "..." if len(head) > 25 else head
    
                # Alternating row colors
                if i % 2 == 0:
                    pdf.set_fill_color(248, 249, 250)
                else:
                    pdf.set_fill_color(255, 255, 255)
    
                pdf.cell(col_widths[0], 6, dept, border=1, fill=True)
                pdf.cell(col_widths[1], 6, head, border=1, fill=True)
                pdf.cell(col_widths[2], 6, count, border=1, fill=True, align="C")
                pdf.cell(col_widths[3], 6, status, border=1, fill=True, align="C")
                pdf.cell(col_widths[4], 6, active, border=1, fill=True, align="C")
                pdf.ln()
    
            # Footer note
            pdf.ln(5)
            pdf.set_font("Arial", "I", 8)
            pdf.set_text_color(150, 150, 150)
            pdf.cell(0, 6, safe_encode(f"Report generated by: CBCentra School Management System."), ln=True, align="C")
    
            # Output to bytes
            pdf_bytes = pdf.output(dest='S').encode('latin-1')
    
            # Open in built-in viewer
            from utils.pdf_utils import view_pdf
            view_pdf(pdf_bytes, parent=self)
    
            print("✅ Departments PDF generated successfully.")
    
        except ImportError:
            QMessageBox.critical(
                self, "Missing Library",
                "Please install fpdf2: pip install fpdf2"
            )
        except Exception as e:
            error_msg = str(e)
            print(f"❌ PDF Export Error: {error_msg}")
            QMessageBox.critical(
                self, "Export Failed",
                f"Failed to generate PDF: {error_msg}"
            )
        
    def export_department_summary(self):
        """Export department summary with statistics"""
        try:
            # Get summary data
            self.cursor.execute("""
                SELECT 
                    s.school_name,
                    COUNT(d.id) as total_departments,
                    COUNT(CASE WHEN d.is_active = 1 THEN 1 END) as active_departments,
                    COUNT(CASE WHEN d.is_active = 0 THEN 1 END) as inactive_departments,
                    COUNT(CASE WHEN d.department_head_id IS NOT NULL THEN 1 END) as departments_with_heads,
                    SUM(teacher_counts.teacher_count) as total_staff
                FROM schools s
                LEFT JOIN departments d ON s.id = d.school_id
                LEFT JOIN (
                    SELECT department_id, COUNT(*) as teacher_count 
                    FROM teachers 
                    WHERE is_active = 1 
                    GROUP BY department_id
                ) teacher_counts ON d.id = teacher_counts.department_id
                GROUP BY s.id, s.school_name
                ORDER BY s.school_name
            """)
            summary_rows = self.cursor.fetchall()
    
            if not summary_rows:
                QMessageBox.information(self, "No Data", "No summary data found to export.")
                return
    
            # Prepare summary data
            headers = [
                "School Name", "Total Departments", "Active Departments", 
                "Inactive Departments", "Departments with Heads", "Total Staff"
            ]
            
            summary_data = []
            for row in summary_rows:
                summary_data.append([
                    row['school_name'] or 'N/A',
                    str(row['total_departments'] or 0),
                    str(row['active_departments'] or 0),
                    str(row['inactive_departments'] or 0),
                    str(row['departments_with_heads'] or 0),
                    str(row['total_staff'] or 0)
                ])
    
            # Add totals row
            totals = [
                sum(int(row[1]) for row in summary_data),  # Total departments
                sum(int(row[2]) for row in summary_data),  # Active departments
                sum(int(row[3]) for row in summary_data),  # Inactive departments
                sum(int(row[4]) for row in summary_data),  # Departments with heads
                sum(int(row[5]) for row in summary_data)   # Total staff
            ]
            
            summary_data.append([
                "TOTAL", str(totals[0]), str(totals[1]), 
                str(totals[2]), str(totals[3]), str(totals[4])
            ])
    
            # Export using enhanced method
            self.export_with_green_header(
                data=summary_data,
                headers=headers,
                filename_prefix="departments_summary",
                title="DEPARTMENTS SUMMARY REPORT"
            )
    
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export summary: {e}")
    
    
    # Add the force refresh method mentioned in the refresh button
    def force_refresh_data(self):
        """Force refresh by reconnecting to database"""
        try:
            # Force commit any pending transactions and refresh connection
            self.db_connection.commit()
            
            # Optional: Reconnect cursor for fresh data
            if hasattr(self, 'cursor'):
                self.cursor.close()
                self.cursor = self.db_connection.cursor(dictionary=True)
            
            # Reload all data
            self.load_schools()
            self.load_teachers()
            self.load_departments()
            
            QMessageBox.information(self, "Refreshed", "Data refreshed successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Refresh Error", f"Failed to refresh data: {e}")