# ui/parents_form.py
import sys
import os
import csv
from datetime import datetime
from typing import Optional, Dict, Any
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QMessageBox, QFileDialog, QTabWidget, QGroupBox, QFormLayout,
    QTextEdit, QCheckBox, QMenu, QComboBox, QScrollArea, QFrame, QDialog,
    QSplitter, QProgressBar, QSpinBox, QDateEdit, QApplication
)
from PySide6.QtCore import Qt, QDate, Signal, QThread, QTimer
from PySide6.QtGui import QPixmap, QIcon, QFont, QAction, QCursor

from ui.audit_base_form import AuditBaseForm
from models.models import get_db_connection
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import mysql.connector
from mysql.connector import Error
# Add these imports at the top of the file with other matplotlib imports
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure


class DatabaseWorker(QThread):
    """Background thread for database operations"""
    data_loaded = Signal(list)
    error_occurred = Signal(str)
    progress_updated = Signal(int)
    
    def __init__(self, query, params=None):
        super().__init__()
        self.query = query
        self.params = params or ()
    
    def run(self):
        try:
            connection = get_db_connection()
            if not connection:
                self.error_occurred.emit("Failed to connect to database")
                return
                
            cursor = connection.cursor(buffered=True)
            cursor.execute(self.query, self.params)
            results = cursor.fetchall()
            
            cursor.close()
            connection.close()
            
            self.data_loaded.emit(results)
        except Exception as e:
            self.error_occurred.emit(str(e))


class ParentDetailsPopup(QDialog):
    """Enhanced popup to view detailed parent information and linked students"""
    def __init__(self, parent, parent_id, user_session=None):
        super().__init__(parent)
        self.parent_id = parent_id
        self.user_session = user_session
        self.db_connection = None
        self.cursor = None

        self.setWindowTitle(f"Parent Details - ID: {parent_id}")
        self.resize(1100, 750)
        self.setModal(True)
        
        # Enhanced styling
        self.setStyleSheet("""
            QDialog { 
                background-color: #f8f9fa; 
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QLabel { 
                font-weight: bold; 
                color: #495057; 
                padding: 2px;
            }
            QTextEdit, QLineEdit { 
                background-color: white; 
                border: 1px solid #dee2e6; 
                border-radius: 4px;
                padding: 5px;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)

        self.setup_ui()
        self.load_parent_details()

    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Header
        header_frame = QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                          stop: 0 #4a90e2, stop: 1 #357abd);
                border-radius: 8px;
            }
        """)
        header_layout = QVBoxLayout(header_frame)
        
        title = QLabel("Parent Detailed Information")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: white; padding: 15px;")
        header_layout.addWidget(title)
        
        layout.addWidget(header_frame)

        # Main content with splitter
        splitter = QSplitter(Qt.Vertical)
        
        # Parent info section
        parent_group = QGroupBox("Parent Information")
        parent_layout = QFormLayout()
        
        # Info labels with enhanced styling
        self.full_name_lbl = QLineEdit()
        self.full_name_lbl.setReadOnly(True)
        self.relation_lbl = QLineEdit()
        self.relation_lbl.setReadOnly(True)
        self.email_lbl = QLineEdit()
        self.email_lbl.setReadOnly(True)
        self.phone_lbl = QLineEdit()
        self.phone_lbl.setReadOnly(True)
        self.address_lbl = QTextEdit()
        self.address_lbl.setMaximumHeight(80)
        self.address_lbl.setReadOnly(True)
        self.is_payer_lbl = QLineEdit()
        self.is_payer_lbl.setReadOnly(True)
        self.is_emergency_lbl = QLineEdit()
        self.is_emergency_lbl.setReadOnly(True)
        self.created_at_lbl = QLineEdit()
        self.created_at_lbl.setReadOnly(True)

        parent_layout.addRow("Full Name:", self.full_name_lbl)
        parent_layout.addRow("Relation:", self.relation_lbl)
        parent_layout.addRow("Email:", self.email_lbl)
        parent_layout.addRow("Phone:", self.phone_lbl)
        parent_layout.addRow("Address:", self.address_lbl)
        parent_layout.addRow("Fee Payer:", self.is_payer_lbl)
        parent_layout.addRow("Emergency Contact:", self.is_emergency_lbl)
        parent_layout.addRow("Created:", self.created_at_lbl)

        parent_group.setLayout(parent_layout)
        splitter.addWidget(parent_group)

        # Students section
        students_group = QGroupBox("Linked Students")
        students_layout = QVBoxLayout()
        
        self.students_count_lbl = QLabel("Students: 0")
        self.students_count_lbl.setStyleSheet("font-size: 14px; color: #6c757d;")
        students_layout.addWidget(self.students_count_lbl)
        
        self.students_table = QTableWidget()
        self.students_table.setColumnCount(6)
        self.students_table.setHorizontalHeaderLabels([
            "Reg No", "Full Name", "Grade", "Class Year", "Status", "Enrollment Date"
        ])
        self.students_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.students_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.students_table.setAlternatingRowColors(True)
        self.students_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        students_layout.addWidget(self.students_table)
        students_group.setLayout(students_layout)
        splitter.addWidget(students_group)
        
        # Set splitter proportions
        splitter.setSizes([300, 400])
        layout.addWidget(splitter)

        # Enhanced button layout
        button_layout = QHBoxLayout()
        
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.load_parent_details)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745; color: white; border: none; padding: 10px 20px;
                border-radius: 6px; font-weight: bold; min-width: 100px;
            }
            QPushButton:hover { background-color: #218838; }
            QPushButton:pressed { background-color: #1e7e34; }
        """)
        
        export_btn = QPushButton("Export Details")
        export_btn.clicked.connect(self.export_parent_details)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff; color: white; border: none; padding: 10px 20px;
                border-radius: 6px; font-weight: bold; min-width: 100px;
            }
            QPushButton:hover { background-color: #0069d9; }
        """)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d; color: white; border: none; padding: 10px 20px;
                border-radius: 6px; font-weight: bold; min-width: 100px;
            }
            QPushButton:hover { background-color: #5a6268; }
        """)
        
        button_layout.addWidget(refresh_btn)
        button_layout.addWidget(export_btn)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)

    def load_parent_details(self):
        try:
            self.db_connection = get_db_connection()
            if not self.db_connection:
                QMessageBox.critical(self, "Error", "Failed to connect to database")
                return
            self.cursor = self.db_connection.cursor(buffered=True)

            # Load parent info with creation date
            query = """
                SELECT full_name, relation, email, phone, address1, address2,
                       is_payer, is_emergency_contact, created_at
                FROM parents
                WHERE id = %s AND is_active = TRUE
            """
            self.cursor.execute(query, (self.parent_id,))
            parent = self.cursor.fetchone()

            if not parent:
                QMessageBox.warning(self, "Not Found", "Parent not found or inactive.")
                return

            name, relation, email, phone, addr1, addr2, is_payer, is_emergency, created_at = parent

            self.full_name_lbl.setText(name or "N/A")
            self.relation_lbl.setText(relation or "N/A")
            self.email_lbl.setText(email or "N/A")
            self.phone_lbl.setText(phone or "N/A")

            address = ", ".join(filter(None, [addr1, addr2])) or "N/A"
            self.address_lbl.setPlainText(address)

            self.is_payer_lbl.setText("Yes" if is_payer else "No")
            self.is_emergency_lbl.setText("Yes" if is_emergency else "No")
            self.created_at_lbl.setText(str(created_at) if created_at else "N/A")

            # Load linked students with enhanced query
            student_query = """
                SELECT s.regNo, s.full_name, s.grade_applied_for, s.class_year,
                       CASE WHEN s.is_active THEN 'Active' ELSE 'Inactive' END,
                       s.enrollment_date
                FROM student_parent sp
                JOIN students s ON sp.student_id = s.id
                WHERE sp.parent_id = %s AND s.is_active = TRUE
                ORDER BY s.surname, s.first_name
            """
            self.cursor.execute(student_query, (self.parent_id,))
            students = self.cursor.fetchall()

            self.students_count_lbl.setText(f"Students: {len(students)}")
            self.students_table.setRowCount(len(students))
            
            for row_idx, student in enumerate(students):
                for col_idx, data in enumerate(student):
                    item = QTableWidgetItem(str(data) if data else "N/A")
                    self.students_table.setItem(row_idx, col_idx, item)

        except Exception as e:
            print(f"Error loading parent details: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load parent: {str(e)}")

    def export_parent_details(self):
        """Export parent details to PDF"""
        try:
            if not self.parent_id:
                return
                
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export Parent Details", 
                f"parent_{self.parent_id}_details.pdf", 
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
                
            # Create PDF with parent and student details
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, f"Parent Details - ID: {self.parent_id}", 0, 1, "C")
            pdf.ln(10)
            
            # Add parent information
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 8, "Parent Information", 0, 1)
            pdf.set_font("Arial", "", 10)
            
            # Get current data from form
            details = [
                ("Full Name", self.full_name_lbl.text()),
                ("Relation", self.relation_lbl.text()),
                ("Email", self.email_lbl.text()),
                ("Phone", self.phone_lbl.text()),
                ("Address", self.address_lbl.toPlainText()),
                ("Fee Payer", self.is_payer_lbl.text()),
                ("Emergency Contact", self.is_emergency_lbl.text())
            ]
            
            for label, value in details:
                pdf.cell(50, 6, f"{label}:", 0, 0)
                pdf.cell(0, 6, str(value), 0, 1)
            
            pdf.ln(5)
            
            # Add students information
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 8, f"Linked Students ({self.students_table.rowCount()})", 0, 1)
            
            if self.students_table.rowCount() > 0:
                pdf.set_font("Arial", "", 9)
                # Table headers
                headers = ["Reg No", "Name", "Grade", "Class", "Status"]
                col_widths = [30, 60, 30, 30, 25]
                
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 6, header, 1, 0, "C")
                pdf.ln()
                
                # Table data
                for row in range(min(self.students_table.rowCount(), 20)):  # Limit rows
                    for col in range(5):  # Only first 5 columns
                        item = self.students_table.item(row, col)
                        text = item.text() if item else "N/A"
                        pdf.cell(col_widths[col], 6, text[:15], 1, 0)
                    pdf.ln()
            
            pdf.output(file_path)
            QMessageBox.information(self, "Success", "Parent details exported successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {str(e)}")

    def closeEvent(self, event):
        if self.cursor:
            self.cursor.close()
        if self.db_connection and self.db_connection.is_connected():
            self.db_connection.close()
        event.accept()


class ParentsForm(AuditBaseForm):
    """Enhanced Parents Management Form with improved UX"""
    parent_selected = Signal(int)

    def __init__(self, parent=None, user_session=None):
        super().__init__(parent, user_session)
        self.user_session = user_session
        self.current_parent_id = None
        self.db_connection = None
        self.cursor = None
        self.db_worker = None
        self.search_timer = QTimer()
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.perform_search)

        # Connect to database
        try:
            self.db_connection = get_db_connection()
            if not self.db_connection:
                raise Exception("Failed to connect to database")
            self.cursor = self.db_connection.cursor(buffered=True)
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Connection failed: {str(e)}")
            return

        self.setup_ui()
        #self.load_parents()

    # 1. Update your ParentsForm setup_ui method to initialize the menu
    def setup_ui(self):
        layout = QVBoxLayout(self)
        self.setWindowTitle("Parents Management System")
        self.resize(1400, 900)
    
        # Tab widget
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)
    
        self.form_tab = QWidget()
        self.list_tab = QWidget()
        self.analytics_tab = QWidget()
    
        self.tabs.addTab(self.form_tab, "Parent Form")
        self.tabs.addTab(self.list_tab, "Parents List")
        self.tabs.addTab(self.analytics_tab, "Analytics")
    
        self.setup_form_tab()
        self.setup_list_tab()
        self.setup_analytics_tab()
        self._last_error = False
        
        # Initialize the refresh menu HERE - this is the key fix
        self.create_refresh_menu()
        
        # Update stats periodically and load data without message
        self.update_statistics()
        self.load_parents()  # No success message on startup
    
    # 2. Create a separate method to initialize the refresh menu
    def create_refresh_menu(self):
        """Create the refresh menu (called during initialization)"""
        self.refresh_menu = QMenu(self)
    
        # Action 1: Refresh All Data
        refresh_all_action = QAction("🔄 Refresh All Data", self)
        refresh_all_action.triggered.connect(self.refresh_all_data)
        self.refresh_menu.addAction(refresh_all_action)
    
        # Action 2: Refresh Student Counts Only
        refresh_counts_action = QAction("🔢 Refresh Student Counts Only", self)
        refresh_counts_action.triggered.connect(self.refresh_student_counts)
        self.refresh_menu.addAction(refresh_counts_action)
    
        # Action 3: Cleanup Orphaned Relationships
        cleanup_action = QAction("🧹 Cleanup Orphaned Relationships", self)
        cleanup_action.triggered.connect(self.cleanup_orphaned_relationships)
        self.refresh_menu.addAction(cleanup_action)
    
        # Separator
        self.refresh_menu.addSeparator()
    
        # Action 4: Enable Auto-Refresh (Toggle)
        auto_refresh_action = QAction("⏱ Enable Auto-Refresh", self)
        auto_refresh_action.setCheckable(True)
        auto_refresh_action.toggled.connect(self.toggle_auto_refresh)
        self.refresh_menu.addAction(auto_refresh_action)
    
    # 3. Remove/update the add_refresh_controls method since we're not using the button
    def add_refresh_controls(self):
        """Legacy method - refresh menu is now created in create_refresh_menu"""
        # This method can be removed or just return None
        return None

    def setup_form_tab(self):
        layout = QVBoxLayout(self.form_tab)
        
        # Progress bar for operations
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Main splitter
        splitter = QSplitter(Qt.Horizontal)
        
        # Form section
        form_scroll = QScrollArea()
        form_scroll.setWidgetResizable(True)
        form_content = QWidget()
        form_layout = QVBoxLayout(form_content)

        # Personal Information Group
        personal_group = QGroupBox("Personal Information")
        personal_layout = QFormLayout()

        self.first_name_entry = QLineEdit()
        self.first_name_entry.setPlaceholderText("Enter first name")
        self.surname_entry = QLineEdit()
        self.surname_entry.setPlaceholderText("Enter surname")
        self.full_name_entry = QLineEdit()
        self.full_name_entry.setReadOnly(True)
        self.full_name_entry.setStyleSheet("background-color: #f8f9fa;")
        
        self.relation_combo = QComboBox()
        self.relation_combo.addItems([
            "", "Father", "Mother", "Guardian", "Uncle", "Aunt", "Grandfather", 
            "Grandmother", "Brother", "Sister", "Other"
        ])
        self.relation_combo.setEditable(True)
        
        personal_layout.addRow("First Name *", self.first_name_entry)
        personal_layout.addRow("Surname *", self.surname_entry)
        personal_layout.addRow("Full Name", self.full_name_entry)
        personal_layout.addRow("Relation", self.relation_combo)

        personal_group.setLayout(personal_layout)
        form_layout.addWidget(personal_group)

        # Contact Information Group
        contact_group = QGroupBox("Contact Information")
        contact_layout = QFormLayout()

        self.email_entry = QLineEdit()
        self.email_entry.setPlaceholderText("Enter email address")
        self.phone_entry = QLineEdit()
        self.phone_entry.setPlaceholderText("Enter phone number")
        
        self.address1_entry = QTextEdit()
        self.address1_entry.setMaximumHeight(60)
        self.address1_entry.setPlaceholderText("Enter primary address")
        self.address2_entry = QTextEdit()
        self.address2_entry.setMaximumHeight(60)
        self.address2_entry.setPlaceholderText("Enter secondary address (optional)")

        contact_layout.addRow("Email", self.email_entry)
        contact_layout.addRow("Phone *", self.phone_entry)
        contact_layout.addRow("Address Line 1", self.address1_entry)
        contact_layout.addRow("Address Line 2", self.address2_entry)

        contact_group.setLayout(contact_layout)
        form_layout.addWidget(contact_group)

        # Settings Group
        settings_group = QGroupBox("Settings")
        settings_layout = QFormLayout()

        self.is_payer_check = QCheckBox("This parent is responsible for fee payments")
        self.is_emergency_check = QCheckBox("This parent is an emergency contact")
        self.is_active_check = QCheckBox("This parent record is active")
        self.is_active_check.setChecked(True)

        settings_layout.addRow(self.is_payer_check)
        settings_layout.addRow(self.is_emergency_check)
        settings_layout.addRow(self.is_active_check)

        settings_group.setLayout(settings_layout)
        form_layout.addWidget(settings_group)

        form_scroll.setWidget(form_content)
        splitter.addWidget(form_scroll)

        # Quick actions panel
        actions_frame = QFrame()
        actions_frame.setMaximumWidth(250)
        actions_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        actions_layout = QVBoxLayout(actions_frame)
        
        actions_title = QLabel("Quick Actions")
        actions_title.setStyleSheet("font-weight: bold; font-size: 14px; margin-bottom: 10px;")
        actions_layout.addWidget(actions_title)

        # Action buttons with different colors
        self.save_btn = QPushButton("Save Parent")
        self.save_btn.setStyleSheet("background-color: #28a745; color: white;")
        self.save_btn.clicked.connect(self.save_parent)
        
        self.update_btn = QPushButton("Update Parent")
        self.update_btn.setStyleSheet("background-color: #ffc107; color: black;")
        self.update_btn.clicked.connect(self.update_parent)
        self.update_btn.setEnabled(False)
        
        self.clear_btn = QPushButton("Clear Form")
        self.clear_btn.setStyleSheet("background-color: #6c757d; color: white;")
        self.clear_btn.clicked.connect(self.clear_form)
        
        self.delete_btn = QPushButton("Delete Parent")
        self.delete_btn.setStyleSheet("background-color: #dc3545; color: white;")
        self.delete_btn.clicked.connect(self.delete_parent)
        self.delete_btn.setEnabled(False)

        for btn in [self.save_btn, self.update_btn, self.clear_btn, self.delete_btn]:
            btn.setMinimumHeight(40)
            actions_layout.addWidget(btn)

        actions_layout.addStretch()
        splitter.addWidget(actions_frame)
        
        # Set splitter proportions
        splitter.setSizes([800, 250])
        layout.addWidget(splitter)

        # Auto-update full name
        self.first_name_entry.textChanged.connect(self.update_full_name)
        self.surname_entry.textChanged.connect(self.update_full_name)

    # Updated setup_list_tab method with refresh button removed
    def setup_list_tab(self):
        layout = QVBoxLayout(self.list_tab)
    
        # Enhanced search section
        search_frame = QFrame()
        search_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 15px;
                margin-bottom: 10px;
            }
        """)
        search_layout = QHBoxLayout(search_frame)
    
        search_layout.addWidget(QLabel("Search:"))
        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText("Search by name, phone, email, or relation...")
        self.search_entry.textChanged.connect(self.on_search_text_changed)
        search_layout.addWidget(self.search_entry)
    
        # Filter by relation
        relation_filter = QComboBox()
        relation_filter.addItems(["All Relations", "Father", "Mother", "Guardian", "Other"])
        relation_filter.currentTextChanged.connect(self.apply_filters)
        self.relation_filter = relation_filter
        search_layout.addWidget(QLabel("Relation:"))
        search_layout.addWidget(relation_filter)
    
        # Status filter
        status_filter = QComboBox()
        status_filter.addItems(["All Status", "Active", "Inactive"])
        status_filter.currentTextChanged.connect(self.apply_filters)
        self.status_filter = status_filter
        search_layout.addWidget(QLabel("Status:"))
        search_layout.addWidget(status_filter)
    
        layout.addWidget(search_frame)
    
        # Action buttons WITHOUT refresh options button
        action_layout = QHBoxLayout()
        
        buttons_data = [
            ("New Parent", self.new_parent, "#28a745"),
            ("Edit Selected", self.edit_selected_parent, "#ffc107"),
            ("View Details", self.view_parent_details, "#17a2b8"),
            ("Export Excel", self.export_to_excel, "#198754"),
            ("Generate PDF", self.generate_pdf_report, "#dc3545")
            # Removed refresh options button since it's now in ribbon
        ]
        
        for btn_text, btn_func, btn_color in buttons_data:
            btn = QPushButton(btn_text)
            btn.setStyleSheet(f"background-color: {btn_color}; color: white; min-height: 35px; border-radius: 6px; font-weight: bold;")
            btn.clicked.connect(btn_func)
            action_layout.addWidget(btn)
    
        # Add integrity check button
        integrity_btn = QPushButton("Check Integrity")
        integrity_btn.setStyleSheet("background-color: #fd7e14; color: white; min-height: 35px; border-radius: 6px; font-weight: bold;")
        integrity_btn.clicked.connect(self.validate_parent_student_integrity)
        action_layout.addWidget(integrity_btn)
    
        action_layout.addStretch()
        layout.addLayout(action_layout)
    
        # Enhanced table (rest of the method stays the same)
        self.parents_table = QTableWidget()
        self.parents_table.setColumnCount(10)
        self.parents_table.setHorizontalHeaderLabels([
            "ID", "Name", "Relation", "Phone", "Email", "Students", 
            "Fee Payer", "Emergency", "Status", "Created"
        ])
        
        # Set column widths
        header = self.parents_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)  # ID
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # Name
        header.setSectionResizeMode(2, QHeaderView.Fixed)   # Relation
        header.setSectionResizeMode(3, QHeaderView.Fixed)   # Phone
        header.setSectionResizeMode(4, QHeaderView.Stretch) # Email
        
        self.parents_table.setColumnWidth(0, 60)   # ID
        self.parents_table.setColumnWidth(2, 100)  # Relation
        self.parents_table.setColumnWidth(3, 120)  # Phone
        self.parents_table.setColumnWidth(5, 80)   # Students
        self.parents_table.setColumnWidth(6, 80)   # Fee Payer
        self.parents_table.setColumnWidth(7, 80)   # Emergency
        self.parents_table.setColumnWidth(8, 80)   # Status
        self.parents_table.setColumnWidth(9, 120)  # Created
        
        self.parents_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.parents_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.parents_table.setAlternatingRowColors(True)
        self.parents_table.setSortingEnabled(True)
        self.parents_table.cellClicked.connect(self.on_parent_select)
        self.parents_table.cellDoubleClicked.connect(self.view_parent_details)
    
        # Context menu
        self.parents_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.parents_table.customContextMenuRequested.connect(self.show_context_menu)
    
        layout.addWidget(self.parents_table)
    
    
    def update_full_name(self):
        """Auto-update full name when first name or surname changes"""
        first = self.first_name_entry.text().strip()
        last = self.surname_entry.text().strip()
        full_name = f"{first} {last}".strip()
        self.full_name_entry.setText(full_name)

    def on_search_text_changed(self):
        """Handle search text changes with debounce"""
        self.search_timer.stop()
        self.search_timer.start(500)  # 500ms delay

    def perform_search(self):
        """Perform the actual search"""
        self.apply_filters()

    def apply_filters(self):
        """Apply search and filter criteria with accurate student counts"""
        search_term = self.search_entry.text().strip()
        relation_filter = self.relation_filter.currentText()
        status_filter = self.status_filter.currentText()
        
        try:
            # Enhanced query with accurate student counting
            query = """
                SELECT p.id, p.full_name, p.relation, p.phone, p.email,
                       COALESCE(student_counts.student_count, 0) as student_count,
                       CASE WHEN p.is_payer THEN 'Yes' ELSE 'No' END as is_payer,
                       CASE WHEN p.is_emergency_contact THEN 'Yes' ELSE 'No' END as is_emergency,
                       CASE WHEN p.is_active THEN 'Active' ELSE 'Inactive' END as status,
                       DATE_FORMAT(p.created_at, '%Y-%m-%d') as created_date
                FROM parents p
                LEFT JOIN (
                    SELECT sp.parent_id, COUNT(DISTINCT s.id) as student_count
                    FROM student_parent sp
                    INNER JOIN students s ON sp.student_id = s.id
                    WHERE s.is_active = TRUE
                    GROUP BY sp.parent_id
                ) student_counts ON p.id = student_counts.parent_id
                WHERE 1=1
            """
            params = []
            
            # Add search filter
            if search_term:
                query += """ AND (p.full_name LIKE %s OR p.phone LIKE %s 
                               OR p.email LIKE %s OR p.relation LIKE %s)"""
                like_term = f"%{search_term}%"
                params.extend([like_term, like_term, like_term, like_term])
            
            # Add relation filter
            if relation_filter != "All Relations":
                query += " AND p.relation = %s"
                params.append(relation_filter)
            
            # Add status filter
            if status_filter == "Active":
                query += " AND p.is_active = TRUE"
            elif status_filter == "Inactive":
                query += " AND p.is_active = FALSE"
            
            query += " ORDER BY p.full_name LIMIT 200"
            
            self.cursor.execute(query, params)
            parents = self.cursor.fetchall()
            
            self.populate_parents_table(parents)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Filter failed: {str(e)}")
            
    def load_parents(self, show_success_message=False):
        """Load all parents with enhanced information and proper count calculation"""
        try:
            self.show_loading(True)
            query = """
                SELECT p.id, p.full_name, p.relation, p.phone, p.email,
                       COALESCE(student_counts.student_count, 0) as student_count,
                       CASE WHEN p.is_payer THEN 'Yes' ELSE 'No' END as is_payer,
                       CASE WHEN p.is_emergency_contact THEN 'Yes' ELSE 'No' END as is_emergency,
                       CASE WHEN p.is_active THEN 'Active' ELSE 'Inactive' END as status,
                       DATE_FORMAT(p.created_at, '%Y-%m-%d') as created_date
                FROM parents p
                LEFT JOIN (
                    SELECT sp.parent_id, COUNT(DISTINCT s.id) as student_count
                    FROM student_parent sp
                    INNER JOIN students s ON sp.student_id = s.id
                    WHERE s.is_active = TRUE
                    GROUP BY sp.parent_id
                ) student_counts ON p.id = student_counts.parent_id
                WHERE p.is_active = TRUE
                ORDER BY p.full_name
                LIMIT 200
            """
            self.db_worker = DatabaseWorker(query)
            self.db_worker.data_loaded.connect(self.populate_parents_table)
            self.db_worker.error_occurred.connect(self.handle_database_error)
            
            def on_finished():
                self.show_loading(False)
                # Only show success message if explicitly requested (manual refresh)
                if show_success_message and (not hasattr(self, '_last_error') or not self._last_error):
                    QMessageBox.information(self, "Success", "Data refreshed successfully!")
                elif hasattr(self, '_last_error') and self._last_error:
                    self._last_error = False  # Reset
    
            self.db_worker.finished.connect(on_finished)
            self.db_worker.start()
        except Exception as e:
            self.show_loading(False)
            QMessageBox.critical(self, "Error", f"Failed to load parents: {str(e)}")
    
    def refresh_all_data(self):
        """Manually refresh all data with success message"""
        self.load_parents(show_success_message=True)

    def populate_parents_table(self, parents):
        """Populate the parents table with data"""
        self.parents_table.setRowCount(len(parents))
        self.parents_table.setSortingEnabled(False)  # Disable during population
        
        for row_idx, parent in enumerate(parents):
            for col_idx, data in enumerate(parent):
                if data is None:
                    text = "N/A"
                elif col_idx == 5:  # Student count
                    text = str(data) if data > 0 else "0"
                else:
                    text = str(data)
                    
                item = QTableWidgetItem(text)
                
                # Color coding for status
                if col_idx == 8:  # Status column
                    if text == "Active":
                        item.setBackground(Qt.green)
                        item.setForeground(Qt.white)
                    else:
                        item.setBackground(Qt.red)
                        item.setForeground(Qt.white)
                
                self.parents_table.setItem(row_idx, col_idx, item)
        
        self.parents_table.setSortingEnabled(True)
        self.update_statistics()

    def show_loading(self, show):
        """Show/hide loading indicator"""
        self.progress_bar.setVisible(show)
        if show:
            self.progress_bar.setRange(0, 0)  # Indeterminate progress
            QApplication.setOverrideCursor(QCursor(Qt.WaitCursor))
        else:
            QApplication.restoreOverrideCursor()

    def handle_database_error(self, error_message):
        """Handle database errors"""
        self.show_loading(False)
        self._last_error = True
        QMessageBox.critical(self, "Database Error", error_message)

    def show_context_menu(self, position):
        """Show context menu for table items"""
        if self.current_parent_id:
            menu = QMenu()
            
            view_action = QAction("View Details", self)
            view_action.triggered.connect(self.view_parent_details)
            menu.addAction(view_action)

            edit_action = QAction("Edit", self)
            edit_action.triggered.connect(self.edit_selected_parent)
            menu.addAction(edit_action)
            
            link_action = QAction("Link Student", self)
            link_action.triggered.connect(self.link_student_to_parent)
            menu.addAction(link_action)

            menu.addSeparator()
            
            delete_action = QAction("Delete", self)
            delete_action.triggered.connect(self.delete_parent)
            menu.addAction(delete_action)

            menu.exec_(self.parents_table.mapToGlobal(position))

    def on_parent_select(self, row, col):
        """Handle parent selection"""
        item = self.parents_table.item(row, 0)
        if item:
            self.current_parent_id = int(item.text())
            self.update_btn.setEnabled(True)
            self.delete_btn.setEnabled(True)

    def new_parent(self):
        """Create new parent"""
        self.tabs.setCurrentWidget(self.form_tab)
        self.clear_form()

    def view_parent_details(self, parent_id=None):
        """View detailed parent information"""
        pid = parent_id or self.current_parent_id
        if not pid:
            QMessageBox.warning(self, "Select", "Please select a parent first.")
            return
        
        try:
            dialog = ParentDetailsPopup(self, pid, self.user_session)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open details: {str(e)}")

    def edit_selected_parent(self):
        """Edit selected parent"""
        if not self.current_parent_id:
            QMessageBox.warning(self, "Select", "Please select a parent to edit.")
            return
        
        self.tabs.setCurrentWidget(self.form_tab)
        self.load_parent_details(self.current_parent_id)

    def load_parent_details(self, parent_id):
        """Load parent details into form"""
        try:
            query = """
                SELECT first_name, surname, full_name, relation, email, phone,
                       address1, address2, is_payer, is_emergency_contact, is_active
                FROM parents WHERE id = %s
            """
            self.cursor.execute(query, (parent_id,))
            parent = self.cursor.fetchone()
            
            if not parent:
                QMessageBox.warning(self, "Not Found", "Parent not found.")
                return

            self.clear_form()
            self.current_parent_id = parent_id

            self.first_name_entry.setText(parent[0] or "")
            self.surname_entry.setText(parent[1] or "")
            self.relation_combo.setCurrentText(parent[3] or "")
            self.email_entry.setText(parent[4] or "")
            self.phone_entry.setText(parent[5] or "")
            self.address1_entry.setPlainText(parent[6] or "")
            self.address2_entry.setPlainText(parent[7] or "")
            self.is_payer_check.setChecked(bool(parent[8]))
            self.is_emergency_check.setChecked(bool(parent[9]))
            self.is_active_check.setChecked(bool(parent[10]))
            
            # Enable update/delete buttons
            self.update_btn.setEnabled(True)
            self.delete_btn.setEnabled(True)
            self.save_btn.setText("Save as New")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load parent: {str(e)}")

    def save_parent(self):
        """Save new parent"""
        if not self.validate_form():
            return

        first = self.first_name_entry.text().strip()
        last = self.surname_entry.text().strip()
        phone = self.phone_entry.text().strip()
        full_name = f"{first} {last}".strip()

        try:
            self.show_loading(True)
            
            # Check for duplicates
            self.cursor.execute(
                "SELECT id FROM parents WHERE phone = %s AND is_active = TRUE", 
                (phone,)
            )
            if self.cursor.fetchone():
                QMessageBox.warning(self, "Duplicate", "A parent with this phone number already exists.")
                self.show_loading(False)
                return

            query = """
                INSERT INTO parents (
                    school_id, first_name, surname, full_name, relation,
                    email, phone, address1, address2, is_payer,
                    is_emergency_contact, is_active
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            values = (
                1,  # Default school_id
                first, last, full_name, self.relation_combo.currentText(),
                self.email_entry.text().strip(), phone,
                self.address1_entry.toPlainText().strip(), 
                self.address2_entry.toPlainText().strip(),
                self.is_payer_check.isChecked(), 
                self.is_emergency_check.isChecked(),
                self.is_active_check.isChecked()
            )
            
            self.cursor.execute(query, values)
            self.current_parent_id = self.cursor.lastrowid
            self.db_connection.commit()
            
            QMessageBox.information(self, "Success", "Parent saved successfully!")
            self.load_parents()
            self.clear_form()
            
        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Save failed: {str(e)}")
        finally:
            self.show_loading(False)

    def update_parent(self):
        """Update existing parent"""
        if not self.current_parent_id:
            QMessageBox.warning(self, "Error", "No parent selected for update.")
            return

        if not self.validate_form():
            return

        first = self.first_name_entry.text().strip()
        last = self.surname_entry.text().strip()
        full_name = f"{first} {last}".strip()

        try:
            self.show_loading(True)
            
            query = """
                UPDATE parents SET
                first_name=%s, surname=%s, full_name=%s, relation=%s,
                email=%s, phone=%s, address1=%s, address2=%s,
                is_payer=%s, is_emergency_contact=%s, is_active=%s,
                updated_at=CURRENT_TIMESTAMP
                WHERE id=%s
            """
            values = (
                first, last, full_name, self.relation_combo.currentText(),
                self.email_entry.text().strip(), self.phone_entry.text().strip(),
                self.address1_entry.toPlainText().strip(),
                self.address2_entry.toPlainText().strip(),
                self.is_payer_check.isChecked(), 
                self.is_emergency_check.isChecked(),
                self.is_active_check.isChecked(), 
                self.current_parent_id
            )
            
            self.cursor.execute(query, values)
            self.db_connection.commit()
            
            QMessageBox.information(self, "Success", "Parent updated successfully!")
            self.load_parents()
            
        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Update failed: {str(e)}")
        finally:
            self.show_loading(False)

    def validate_form(self):
        """Validate form data"""
        if not self.first_name_entry.text().strip():
            QMessageBox.warning(self, "Validation", "First name is required.")
            self.first_name_entry.setFocus()
            return False
            
        if not self.surname_entry.text().strip():
            QMessageBox.warning(self, "Validation", "Surname is required.")
            self.surname_entry.setFocus()
            return False
            
        if not self.phone_entry.text().strip():
            QMessageBox.warning(self, "Validation", "Phone number is required.")
            self.phone_entry.setFocus()
            return False
            
        # Basic email validation
        email = self.email_entry.text().strip()
        if email and "@" not in email:
            QMessageBox.warning(self, "Validation", "Please enter a valid email address.")
            self.email_entry.setFocus()
            return False
            
        return True

    def clear_form(self):
        """Clear all form fields"""
        self.first_name_entry.clear()
        self.surname_entry.clear()
        self.full_name_entry.clear()
        self.relation_combo.setCurrentIndex(0)
        self.email_entry.clear()
        self.phone_entry.clear()
        self.address1_entry.clear()
        self.address2_entry.clear()
        self.is_payer_check.setChecked(False)
        self.is_emergency_check.setChecked(False)
        self.is_active_check.setChecked(True)
        
        self.current_parent_id = None
        self.update_btn.setEnabled(False)
        self.delete_btn.setEnabled(False)
        self.save_btn.setText("Save Parent")

    def delete_parent(self):
        """Delete selected parent with confirmation"""
        if not self.current_parent_id:
            QMessageBox.warning(self, "Select", "Please select a parent to delete.")
            return

        try:
            # Check if parent has linked students
            self.cursor.execute(
                "SELECT COUNT(*) FROM student_parent WHERE parent_id = %s", 
                (self.current_parent_id,)
            )
            student_count = self.cursor.fetchone()[0]
            
            if student_count > 0:
                reply = QMessageBox.question(
                    self, "Confirm Delete",
                    f"This parent is linked to {student_count} student(s). "
                    "Deleting will remove these links. Continue?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return

            # Final confirmation
            reply = QMessageBox.question(
                self, "Confirm Delete",
                "Are you sure you want to delete this parent? This action cannot be undone.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.show_loading(True)
                
                # Soft delete (set is_active = FALSE)
                self.cursor.execute(
                    "UPDATE parents SET is_active = FALSE WHERE id = %s", 
                    (self.current_parent_id,)
                )
                self.db_connection.commit()
                
                QMessageBox.information(self, "Success", "Parent deleted successfully.")
                self.load_parents()
                self.clear_form()
                
        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Delete failed: {str(e)}")
        finally:
            self.show_loading(False)

    def update_statistics(self):
        """Update statistics in header and analytics tab"""
        try:
            # Get statistics
            stats_query = """
                SELECT 
                    COUNT(*) as total_parents,
                    SUM(CASE WHEN is_active = TRUE THEN 1 ELSE 0 END) as active_parents,
                    SUM(CASE WHEN is_payer = TRUE AND is_active = TRUE THEN 1 ELSE 0 END) as fee_payers,
                    SUM(CASE WHEN is_emergency_contact = TRUE AND is_active = TRUE THEN 1 ELSE 0 END) as emergency_contacts
                FROM parents
            """
            self.cursor.execute(stats_query)
            stats = self.cursor.fetchone()
            
            if stats:
                total, active, payers, emergency = stats           
                
                # Update analytics tab if it exists
                if hasattr(self, 'total_parents_lbl'):
                    self.total_parents_lbl.value_label.setText(str(total))
                    self.active_parents_lbl.value_label.setText(str(active))
                    self.fee_payers_lbl.value_label.setText(str(payers))
                    self.emergency_contacts_lbl.value_label.setText(str(emergency))
                
                # Update relations breakdown
                self.update_relations_breakdown()
                
        except Exception as e:
            print(f"Error updating statistics: {e}")

    
    def update_relations_breakdown(self):
        """Update relations breakdown table"""
        try:
            relations_query = """
                SELECT 
                    COALESCE(relation, 'Not Specified') as relation,
                    COUNT(*) as count,
                    ROUND((COUNT(*) * 100.0 / (SELECT COUNT(*) FROM parents WHERE is_active = TRUE)), 1) as percentage
                FROM parents 
                WHERE is_active = TRUE
                GROUP BY relation
                ORDER BY count DESC
            """
            self.cursor.execute(relations_query)
            relations = self.cursor.fetchall()
            
            if hasattr(self, 'relations_table'):
                self.relations_table.setRowCount(len(relations))
                for row_idx, (relation, count, percentage) in enumerate(relations):
                    self.relations_table.setItem(row_idx, 0, QTableWidgetItem(str(relation)))
                    self.relations_table.setItem(row_idx, 1, QTableWidgetItem(str(count)))
                    self.relations_table.setItem(row_idx, 2, QTableWidgetItem(f"{percentage}%"))
                    
        except Exception as e:
            print(f"Error updating relations breakdown: {e}")
        
    def refresh_student_counts(self):
        """Manually refresh student counts for all visible parents"""
        try:
            self.show_loading(True)
            
            # Get all parent IDs currently displayed
            parent_ids = []
            for row in range(self.parents_table.rowCount()):
                item = self.parents_table.item(row, 0)  # ID column
                if item:
                    parent_ids.append(int(item.text()))
            
            if not parent_ids:
                return
            
            # Create placeholders for IN clause
            placeholders = ','.join(['%s'] * len(parent_ids))
            
            # Query to get updated counts
            count_query = f"""
                SELECT sp.parent_id, COUNT(DISTINCT s.id) as student_count
                FROM student_parent sp
                INNER JOIN students s ON sp.student_id = s.id
                WHERE s.is_active = TRUE AND sp.parent_id IN ({placeholders})
                GROUP BY sp.parent_id
            """
            
            self.cursor.execute(count_query, parent_ids)
            updated_counts = dict(self.cursor.fetchall())
            
            # Update the table
            for row in range(self.parents_table.rowCount()):
                parent_id_item = self.parents_table.item(row, 0)
                if parent_id_item:
                    parent_id = int(parent_id_item.text())
                    count = updated_counts.get(parent_id, 0)
                    
                    # Update the count column (index 5)
                    count_item = QTableWidgetItem(str(count))
                    self.parents_table.setItem(row, 5, count_item)
            
            QMessageBox.information(self, "Success", "Student counts refreshed successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to refresh counts: {str(e)}")
        finally:
            self.show_loading(False)
    
    def cleanup_orphaned_relationships(self):
        """Remove orphaned parent-student relationships"""
        try:
            reply = QMessageBox.question(
                self, "Cleanup Relationships",
                "This will remove relationships where:\n"
                "- Student is inactive/deleted\n"
                "- Parent is inactive/deleted\n"
                "- Duplicate relationships exist\n\n"
                "Continue?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            self.show_loading(True)
            
            # Remove relationships with inactive students
            cleanup_queries = [
                """DELETE sp FROM student_parent sp 
                   LEFT JOIN students s ON sp.student_id = s.id 
                   WHERE s.id IS NULL OR s.is_active = FALSE""",
                
                """DELETE sp FROM student_parent sp 
                   LEFT JOIN parents p ON sp.parent_id = p.id 
                   WHERE p.id IS NULL OR p.is_active = FALSE""",
                
                # Remove duplicate relationships
                """DELETE sp1 FROM student_parent sp1
                   INNER JOIN student_parent sp2 
                   WHERE sp1.id > sp2.id 
                   AND sp1.student_id = sp2.student_id 
                   AND sp1.parent_id = sp2.parent_id"""
            ]
            
            total_deleted = 0
            for query in cleanup_queries:
                self.cursor.execute(query)
                total_deleted += self.cursor.rowcount
            
            self.db_connection.commit()
            
            QMessageBox.information(
                self, "Cleanup Complete", 
                f"Removed {total_deleted} orphaned/duplicate relationships."
            )
            
            # Refresh the display
            self.load_parents()
            
        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Cleanup failed: {str(e)}")
        finally:
            self.show_loading(False)
    
    def setup_auto_refresh(self):
        """Setup automatic refresh of student counts"""
        # Add a timer to periodically refresh counts
        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self.check_for_updates)
        self.refresh_timer.start(30000)  # Check every 30 seconds
        
    def check_for_updates(self):
        """Check if parent-student relationships have been modified"""
        try:
            # Get the latest modification timestamp from student_parent table
            self.cursor.execute("""
                SELECT MAX(GREATEST(
                    COALESCE(sp.created_at, '1970-01-01'),
                    COALESCE(sp.updated_at, '1970-01-01'),
                    COALESCE(s.updated_at, '1970-01-01'),
                    COALESCE(p.updated_at, '1970-01-01')
                )) as last_modified
                FROM student_parent sp
                LEFT JOIN students s ON sp.student_id = s.id
                LEFT JOIN parents p ON sp.parent_id = p.id
            """)
            
            result = self.cursor.fetchone()
            if result and result[0]:
                last_modified = result[0]
                
                # Check if we need to refresh (compare with stored timestamp)
                if not hasattr(self, 'last_refresh_time') or last_modified > self.last_refresh_time:
                    self.refresh_student_counts()
                    self.last_refresh_time = last_modified
                    
        except Exception as e:
            print(f"Auto-refresh check failed: {e}")

    
    def toggle_auto_refresh(self, enabled):
        """Toggle automatic refresh on/off"""
        if enabled:
            if not hasattr(self, 'refresh_timer'):
                self.setup_auto_refresh()
            else:
                self.refresh_timer.start(30000)
        else:
            if hasattr(self, 'refresh_timer'):
                self.refresh_timer.stop()
    
    def validate_parent_student_integrity(self):
        """Validate data integrity between parents and students"""
        try:
            integrity_issues = []
            
            # Check for parents with non-existent students
            self.cursor.execute("""
                SELECT p.id, p.full_name, COUNT(sp.student_id) as linked_count,
                       COUNT(s.id) as active_count
                FROM parents p
                LEFT JOIN student_parent sp ON p.id = sp.parent_id
                LEFT JOIN students s ON sp.student_id = s.id AND s.is_active = TRUE
                WHERE p.is_active = TRUE
                GROUP BY p.id, p.full_name
                HAVING linked_count > active_count
            """)
            
            problematic_parents = self.cursor.fetchall()
            
            if problematic_parents:
                for parent_id, parent_name, linked, active in problematic_parents:
                    integrity_issues.append(
                        f"Parent '{parent_name}' (ID: {parent_id}) has {linked} links but only {active} active students"
                    )
            
            # Check for duplicate relationships
            self.cursor.execute("""
                SELECT sp1.parent_id, sp1.student_id, COUNT(*) as duplicate_count
                FROM student_parent sp1
                GROUP BY sp1.parent_id, sp1.student_id
                HAVING COUNT(*) > 1
            """)
            
            duplicates = self.cursor.fetchall()
            if duplicates:
                for parent_id, student_id, count in duplicates:
                    integrity_issues.append(
                        f"Duplicate relationship: Parent {parent_id} -> Student {student_id} ({count} times)"
                    )
            
            if integrity_issues:
                issues_text = "\n".join(integrity_issues[:10])  # Show first 10 issues
                if len(integrity_issues) > 10:
                    issues_text += f"\n... and {len(integrity_issues) - 10} more issues"
                
                QMessageBox.warning(
                    self, "Data Integrity Issues Found",
                    f"Found {len(integrity_issues)} integrity issues:\n\n{issues_text}\n\n"
                    "Use 'Cleanup Orphaned Relationships' to fix these issues."
                )
            else:
                QMessageBox.information(self, "Integrity Check", "No integrity issues found!")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Integrity check failed: {str(e)}")

    def export_to_excel(self):
        """Export parents data to Excel with enhanced formatting"""
        try:
            from openpyxl.utils import get_column_letter
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export to Excel", 
                f"parents_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
                
            self.show_loading(True)
            
            # Get comprehensive data
            query = """
                SELECT p.id, p.first_name, p.surname, p.full_name, p.relation, 
                       p.email, p.phone, p.address1, p.address2,
                       CASE WHEN p.is_payer THEN 'Yes' ELSE 'No' END,
                       CASE WHEN p.is_emergency_contact THEN 'Yes' ELSE 'No' END,
                       CASE WHEN p.is_active THEN 'Active' ELSE 'Inactive' END,
                       COUNT(DISTINCT sp.student_id) as student_count,
                       p.created_at, p.updated_at
                FROM parents p
                LEFT JOIN student_parent sp ON p.id = sp.parent_id
                GROUP BY p.id
                ORDER BY p.full_name
            """
            self.cursor.execute(query)
            data = self.cursor.fetchall()
            
            if not data:
                QMessageBox.information(self, "Info", "No data to export.")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Parents Data"
            
            # Headers
            headers = [
                "ID", "First Name", "Surname", "Full Name", "Relation", "Email", 
                "Phone", "Address 1", "Address 2", "Fee Payer", "Emergency Contact", 
                "Status", "Linked Students", "Created", "Updated"
            ]
            ws.append(headers)

            # Data
            for row in data:
                ws.append(list(row))

            # Formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'].font = header_font
                ws[f'{col_letter}1'].fill = header_fill
                ws[f'{col_letter}1'].alignment = Alignment(horizontal="center")
                
                # Auto-adjust column width
                max_length = len(header)
                for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            wb.save(file_path)
            QMessageBox.information(self, "Success", "Data exported successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {str(e)}")
        finally:
            self.show_loading(False)

    def generate_pdf_report(self):
        """Generate comprehensive PDF report"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save PDF Report", 
                f"parents_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
                
            self.show_loading(True)
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "PARENTS MANAGEMENT REPORT", 0, 1, "C")
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 5, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 1, "C")
            pdf.ln(10)

            # Summary statistics
            self.cursor.execute("""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN is_active = TRUE THEN 1 ELSE 0 END) as active,
                    SUM(CASE WHEN is_payer = TRUE THEN 1 ELSE 0 END) as payers,
                    SUM(CASE WHEN is_emergency_contact = TRUE THEN 1 ELSE 0 END) as emergency_contacts
                FROM parents
            """)
            stats = self.cursor.fetchone()
            
            if stats:
                total, active, payers, emergency = stats
                pdf.set_font("Arial", "B", 12)
                pdf.cell(0, 8, "SUMMARY STATISTICS", 0, 1)
                pdf.set_font("Arial", "", 10)
                
                stats_data = [
                    ("Total Parents:", str(total)),
                    ("Active Parents:", str(active)),
                    ("Fee Payers:", str(payers)),
                    ("Emergency Contacts:", str(emergency))
                ]
                
                for label, value in stats_data:
                    pdf.cell(60, 6, label, 1)
                    pdf.cell(40, 6, value, 1)
                    pdf.ln()
                
                pdf.ln(10)

            # Relations breakdown
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 8, "RELATIONS BREAKDOWN", 0, 1)
            pdf.set_font("Arial", "", 10)
            
            self.cursor.execute("""
                SELECT 
                    COALESCE(relation, 'Not Specified') as relation,
                    COUNT(*) as count
                FROM parents 
                WHERE is_active = TRUE
                GROUP BY relation
                ORDER BY count DESC
                LIMIT 10
            """)
            relations = self.cursor.fetchall()
            
            if relations:
                pdf.cell(80, 6, "Relation", 1, 0, "C")
                pdf.cell(40, 6, "Count", 1, 1, "C")
                
                for relation, count in relations:
                    pdf.cell(80, 6, str(relation), 1)
                    pdf.cell(40, 6, str(count), 1)
                    pdf.ln()
                
                pdf.ln(5)

            # Recent parents (last 20)
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 8, "RECENTLY ADDED PARENTS", 0, 1)
            pdf.set_font("Arial", "", 9)
            
            self.cursor.execute("""
                SELECT full_name, relation, phone, 
                       DATE_FORMAT(created_at, '%Y-%m-%d') as created_date
                FROM parents 
                WHERE is_active = TRUE
                ORDER BY created_at DESC
                LIMIT 20
            """)
            recent_parents = self.cursor.fetchall()
            
            if recent_parents:
                # Table headers
                headers = ["Name", "Relation", "Phone", "Created"]
                col_widths = [60, 30, 40, 30]
                
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 6, header, 1, 0, "C")
                pdf.ln()
                
                # Data rows
                for parent in recent_parents:
                    for i, data in enumerate(parent):
                        text = str(data) if data else "N/A"
                        # Truncate long names
                        if i == 0 and len(text) > 25:
                            text = text[:22] + "..."
                        pdf.cell(col_widths[i], 6, text, 1)
                    pdf.ln()

            pdf.output(file_path)
            QMessageBox.information(self, "Success", "PDF report generated successfully!")
            
            # Ask to open file
            reply = QMessageBox.question(
                self, "Open File", 
                "Would you like to open the generated PDF report?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                import os
                import subprocess
                import platform
                
                try:
                    if platform.system() == 'Windows':
                        os.startfile(file_path)
                    elif platform.system() == 'Darwin':  # macOS
                        subprocess.call(['open', file_path])
                    else:  # Linux
                        subprocess.call(['xdg-open', file_path])
                except Exception as e:
                    QMessageBox.information(self, "Info", f"Report saved to: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"PDF generation failed: {str(e)}")
        finally:
            self.show_loading(False)

    def setup_analytics_tab(self):
        """Setup enhanced analytics tab with statistics, charts, and visualizations"""
        # Main scroll area
        analytics_scroll = QScrollArea()
        analytics_scroll.setWidgetResizable(True)
        analytics_scroll.setFrameShape(QFrame.NoFrame)
        analytics_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        analytics_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        # Container widget
        analytics_container = QWidget()
        main_layout = QVBoxLayout(analytics_container)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        analytics_scroll.setWidget(analytics_container)
        
        # Set scroll area as main layout
        self.analytics_tab.setLayout(QVBoxLayout())
        self.analytics_tab.layout().addWidget(analytics_scroll)
        self.analytics_tab.layout().setContentsMargins(0, 0, 0, 0)
        
        # Title and refresh button
        header_layout = QHBoxLayout()
        title = QLabel("Parents Analytics Dashboard")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #1f538d; padding: 10px;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        refresh_analytics_btn = QPushButton("Refresh Analytics")
        refresh_analytics_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                font-weight: bold;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #218838; }
        """)
        refresh_analytics_btn.clicked.connect(self.refresh_parents_analytics)
        header_layout.addWidget(refresh_analytics_btn)
        
        main_layout.addLayout(header_layout)
        
        # === OVERVIEW STATISTICS CARDS ===
        stats_container = QWidget()
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)
        
        # Total Parents Card
        self.total_parents_card = self.create_stats_card("Total Parents", "0", "#007bff")
        stats_layout.addWidget(self.total_parents_card)
        
        # Active Parents Card
        self.active_parents_card = self.create_stats_card("Active Parents", "0", "#28a745")
        stats_layout.addWidget(self.active_parents_card)
        
        # Fee Payers Card
        self.fee_payers_card = self.create_stats_card("Fee Payers", "0", "#ffc107")
        stats_layout.addWidget(self.fee_payers_card)
        
        # Emergency Contacts Card
        self.emergency_contacts_card = self.create_stats_card("Emergency Contacts", "0", "#dc3545")
        stats_layout.addWidget(self.emergency_contacts_card)
        
        main_layout.addWidget(stats_container)
        
        # === CHARTS SECTION ===
        charts_container = QWidget()
        charts_layout = QHBoxLayout(charts_container)
        charts_layout.setSpacing(20)
        
        # Left side - Relation Distribution
        left_chart_group = QGroupBox("Distribution by Relation Type")
        left_chart_group.setMinimumHeight(400)
        left_chart_layout = QVBoxLayout()
        left_chart_group.setLayout(left_chart_layout)
        
        # Create matplotlib figure for relation distribution
        self.relation_figure = Figure(figsize=(8, 6))
        self.relation_canvas = FigureCanvas(self.relation_figure)
        self.relation_canvas.setMinimumHeight(350)
        left_chart_layout.addWidget(self.relation_canvas)
        charts_layout.addWidget(left_chart_group, 1)
        
        # Right side - Student Count Distribution  
        right_chart_group = QGroupBox("Distribution by Parent-Student Count")
        right_chart_group.setMinimumHeight(400)
        right_chart_layout = QVBoxLayout()
        right_chart_group.setLayout(right_chart_layout)
        
        # Create matplotlib figure for student count distribution
        self.student_count_figure = Figure(figsize=(8, 6))
        self.student_count_canvas = FigureCanvas(self.student_count_figure)
        self.student_count_canvas.setMinimumHeight(350)
        right_chart_layout.addWidget(self.student_count_canvas)
        charts_layout.addWidget(right_chart_group, 1)
        
        main_layout.addWidget(charts_container)
        
        # === DETAILED TABLES ===
        tables_container = QWidget()
        tables_layout = QHBoxLayout(tables_container)
        tables_layout.setSpacing(20)
        
        # Relation breakdown table
        relation_table_group = QGroupBox("Relation Type Breakdown")
        relation_table_group.setMinimumHeight(250)
        relation_table_layout = QVBoxLayout()
        relation_table_group.setLayout(relation_table_layout)
        
        self.relation_stats_table = QTableWidget()
        self.relation_stats_table.setColumnCount(4)
        self.relation_stats_table.setHorizontalHeaderLabels(["Relation Type", "Count", "Percentage", "Avg Students"])
        self.relation_stats_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.relation_stats_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.relation_stats_table.setMinimumHeight(250)
        self.relation_stats_table.setAlternatingRowColors(True)
        self.relation_stats_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
            }
        """)
        relation_table_layout.addWidget(self.relation_stats_table)
        tables_layout.addWidget(relation_table_group, 1)
        
        # Student count breakdown table
        student_count_table_group = QGroupBox("Student-Parent Count Distribution")
        student_count_table_group.setMinimumHeight(250)
        student_count_table_layout = QVBoxLayout()
        student_count_table_group.setLayout(student_count_table_layout)
        
        self.student_count_stats_table = QTableWidget()
        self.student_count_stats_table.setColumnCount(3)
        self.student_count_stats_table.setHorizontalHeaderLabels(["Students", "Parents", "Percentage"])
        self.student_count_stats_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.student_count_stats_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.student_count_stats_table.setMinimumHeight(250)
        self.student_count_stats_table.setAlternatingRowColors(True)
        self.student_count_stats_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
            }
        """)
        student_count_table_layout.addWidget(self.student_count_stats_table)
        tables_layout.addWidget(student_count_table_group, 1)
        
        main_layout.addWidget(tables_container)
        
        # Add stretch to push content to top and allow scrolling
        main_layout.addStretch()
        
        # Load initial analytics data
        QTimer.singleShot(500, self.refresh_parents_analytics)
    
    def create_stats_card(self, title, value, color):
        """Create a statistics card widget for parents analytics"""
        card = QFrame()
        card.setFrameStyle(QFrame.Box)
        card.setMinimumHeight(120)
        card.setMinimumWidth(200)
        card.setStyleSheet(f"""
            QFrame {{
                border: 2px solid {color};
                border-radius: 10px;
                background-color: white;
                padding: 15px;
            }}
            QLabel {{
                border: none;
                background: transparent;
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setAlignment(Qt.AlignCenter)
        layout.setSpacing(8)
        
        # Title
        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(f"font-size: 14px; font-weight: bold; color: {color}; margin-bottom: 5px;")
        title_label.setWordWrap(True)
        layout.addWidget(title_label)
        
        # Value
        value_label = QLabel(value)
        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #333;")
        value_label.setWordWrap(True)
        layout.addWidget(value_label)
        
        # Store value label for updates
        card.value_label = value_label
        
        return card
    
    def refresh_parents_analytics(self):
        """Refresh all parents analytics data"""
        try:
            # Show loading cursor
            QApplication.setOverrideCursor(Qt.WaitCursor)
            
            # Load all analytics components
            self.load_parents_overview_stats()
            self.load_relation_distribution()
            self.load_student_count_distribution()
            self.update_parents_charts()
            
            # Show success popup only if manually triggered (not initial load)
            if hasattr(self, '_analytics_initial_load_complete'):
                QApplication.restoreOverrideCursor()
                QMessageBox.information(self, "Refresh Complete", 
                                      "Parents analytics data has been refreshed successfully!")
            else:
                self._analytics_initial_load_complete = True
                QApplication.restoreOverrideCursor()
            
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "Error", f"Failed to load parents analytics: {str(e)}")
    
    def load_parents_overview_stats(self):
        """Load overview statistics for parents"""
        try:
            # Total parents
            self.cursor.execute("SELECT COUNT(*) FROM parents")
            total_parents = self.cursor.fetchone()[0]
            
            # Active parents
            self.cursor.execute("SELECT COUNT(*) FROM parents WHERE is_active = TRUE")
            total_active = self.cursor.fetchone()[0]
            
            # Fee payers
            self.cursor.execute("SELECT COUNT(*) FROM parents WHERE is_payer = TRUE AND is_active = TRUE")
            fee_payers = self.cursor.fetchone()[0]
            
            # Emergency contacts
            self.cursor.execute("SELECT COUNT(*) FROM parents WHERE is_emergency_contact = TRUE AND is_active = TRUE")
            emergency_contacts = self.cursor.fetchone()[0]
            
            # Update cards
            self.total_parents_card.value_label.setText(str(total_parents))
            self.active_parents_card.value_label.setText(str(total_active))
            self.fee_payers_card.value_label.setText(str(fee_payers))
            self.emergency_contacts_card.value_label.setText(str(emergency_contacts))
            
            # Force UI update
            self.total_parents_card.value_label.update()
            self.active_parents_card.value_label.update()
            self.fee_payers_card.value_label.update()
            self.emergency_contacts_card.value_label.update()
            
        except Exception as e:
            print(f"Error loading parents overview stats: {e}")
    
    def load_relation_distribution(self):
        """Load distribution by relation type"""
        try:
            query = """
                SELECT 
                    COALESCE(relation, 'Not Specified') as relation,
                    COUNT(*) as total_count,
                    ROUND((COUNT(*) * 100.0 / (SELECT COUNT(*) FROM parents WHERE is_active = TRUE)), 1) as percentage,
                    ROUND(AVG(student_counts.student_count), 1) as avg_students
                FROM parents p
                LEFT JOIN (
                    SELECT sp.parent_id, COUNT(DISTINCT s.id) as student_count
                    FROM student_parent sp
                    INNER JOIN students s ON sp.student_id = s.id
                    WHERE s.is_active = TRUE
                    GROUP BY sp.parent_id
                ) student_counts ON p.id = student_counts.parent_id
                WHERE p.is_active = TRUE
                GROUP BY p.relation
                ORDER BY total_count DESC
            """
            self.cursor.execute(query)
            relation_data = self.cursor.fetchall()
            
            # Update table
            self.relation_stats_table.setRowCount(len(relation_data))
            
            for row_idx, (relation, count, percentage, avg_students) in enumerate(relation_data):
                self.relation_stats_table.setItem(row_idx, 0, QTableWidgetItem(str(relation)))
                self.relation_stats_table.setItem(row_idx, 1, QTableWidgetItem(str(count)))
                self.relation_stats_table.setItem(row_idx, 2, QTableWidgetItem(f"{percentage}%"))
                self.relation_stats_table.setItem(row_idx, 3, QTableWidgetItem(str(avg_students)))
            
            self.relation_stats_data = relation_data
            
        except Exception as e:
            print(f"Error loading relation distribution: {e}")
    
    def load_student_count_distribution(self):
        """Load distribution by student count"""
        try:
            query = """
                SELECT 
                    student_counts.student_count,
                    COUNT(*) as parent_count,
                    ROUND((COUNT(*) * 100.0 / (SELECT COUNT(*) FROM parents WHERE is_active = TRUE)), 1) as percentage
                FROM parents p
                LEFT JOIN (
                    SELECT sp.parent_id, COUNT(DISTINCT s.id) as student_count
                    FROM student_parent sp
                    INNER JOIN students s ON sp.student_id = s.id
                    WHERE s.is_active = TRUE
                    GROUP BY sp.parent_id
                ) student_counts ON p.id = student_counts.parent_id
                WHERE p.is_active = TRUE
                GROUP BY student_counts.student_count
                ORDER BY student_counts.student_count
            """
            self.cursor.execute(query)
            student_count_data = self.cursor.fetchall()
            
            # Update table
            self.student_count_stats_table.setRowCount(len(student_count_data))
            
            for row_idx, (student_count, parent_count, percentage) in enumerate(student_count_data):
                student_count_display = str(student_count) if student_count is not None else "0"
                self.student_count_stats_table.setItem(row_idx, 0, QTableWidgetItem(student_count_display))
                self.student_count_stats_table.setItem(row_idx, 1, QTableWidgetItem(str(parent_count)))
                self.student_count_stats_table.setItem(row_idx, 2, QTableWidgetItem(f"{percentage}%"))
            
            self.student_count_stats_data = student_count_data
            
        except Exception as e:
            print(f"Error loading student count distribution: {e}")
    
    def update_parents_charts(self):
        """Update both parents charts with current data"""
        self.update_relation_chart()
        self.update_student_count_chart()
    
    def update_relation_chart(self):
        """Update the relation distribution chart"""
        try:
            self.relation_figure.clear()
            ax = self.relation_figure.add_subplot(111)
            
            if hasattr(self, 'relation_stats_data') and self.relation_stats_data:
                relations = [item[0] for item in self.relation_stats_data]
                counts = [item[1] for item in self.relation_stats_data]
                
                # Create bar chart
                bars = ax.bar(relations, counts, color=['#4472C4', '#E15759', '#76B7B2', '#F28E2B', '#59A14F'])
                
                ax.set_xlabel('Relation Type', fontsize=12, fontweight='bold')
                ax.set_ylabel('Number of Parents', fontsize=12, fontweight='bold')
                ax.set_title('Parent Distribution by Relation Type', fontsize=14, fontweight='bold', pad=20)
                ax.tick_params(axis='x', rotation=45)
                ax.grid(True, alpha=0.3)
                
                # Add value labels on bars
                for bar in bars:
                    height = bar.get_height()
                    if height > 0:
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{int(height)}', ha='center', va='bottom', 
                               fontsize=9, fontweight='bold')
            else:
                # Show placeholder when no data
                ax.text(0.5, 0.5, 'No Data Available\nAdd parents to see charts', 
                       horizontalalignment='center', verticalalignment='center',
                       transform=ax.transAxes, fontsize=14, color='gray')
                ax.set_title('Parent Distribution by Relation Type', fontsize=14, fontweight='bold')
            
            self.relation_figure.tight_layout()
            self.relation_canvas.draw()
            
        except Exception as e:
            print(f"Error updating relation chart: {e}")
    
    def update_student_count_chart(self):
        """Update the student count distribution chart"""
        try:
            self.student_count_figure.clear()
            ax = self.student_count_figure.add_subplot(111)
            
            if hasattr(self, 'student_count_stats_data') and self.student_count_stats_data:
                student_counts = [item[0] if item[0] is not None else 0 for item in self.student_count_stats_data]
                parent_counts = [item[1] for item in self.student_count_stats_data]
                
                # Create bar chart
                bars = ax.bar([str(x) for x in student_counts], parent_counts, color='#4E79A7')
                
                ax.set_xlabel('Number of Students', fontsize=12, fontweight='bold')
                ax.set_ylabel('Number of Parents', fontsize=12, fontweight='bold')
                ax.set_title('Distribution by Parent-Student Count', fontsize=14, fontweight='bold', pad=20)
                ax.grid(True, alpha=0.3)
                
                # Add value labels on bars
                for bar in bars:
                    height = bar.get_height()
                    if height > 0:
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{int(height)}', ha='center', va='bottom', 
                               fontsize=9, fontweight='bold')
            else:
                # Show placeholder when no data
                ax.text(0.5, 0.5, 'No Data Available\nAdd parents to see charts', 
                       horizontalalignment='center', verticalalignment='center',
                       transform=ax.transAxes, fontsize=14, color='gray')
                ax.set_title('Distribution by Student Count', fontsize=14, fontweight='bold')
            
            self.student_count_figure.tight_layout()
            self.student_count_canvas.draw()
            
        except Exception as e:
            print(f"Error updating student count chart: {e}")

    def closeEvent(self, event):
        """Clean up database connections on close"""
        try:
            if hasattr(self, 'db_worker') and self.db_worker and self.db_worker.isRunning():
                self.db_worker.terminate()
                self.db_worker.wait()
            
            if self.cursor:
                self.cursor.close()
            if self.db_connection and self.db_connection.is_connected():
                self.db_connection.close()
        except Exception as e:
            print(f"Error closing database connections: {e}")
        finally:
            event.accept()


# Additional utility classes for enhanced functionality

class StudentLinkDialog(QDialog):
    """Dialog to link students to parents"""
    def __init__(self, parent, parent_id, parent_name):
        super().__init__(parent)
        self.parent_id = parent_id
        self.parent_name = parent_name
        self.db_connection = None
        self.cursor = None
        
        self.setWindowTitle(f"Link Students to {parent_name}")
        self.resize(800, 600)
        self.setModal(True)
        
        self.setup_ui()
        self.load_available_students()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Header
        header_label = QLabel(f"Select students to link with: {self.parent_name}")
        header_label.setStyleSheet("font-size: 14px; font-weight: bold; padding: 10px;")
        layout.addWidget(header_label)
        
        # Search
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search Students:"))
        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText("Search by name or registration number...")
        self.search_entry.textChanged.connect(self.filter_students)
        search_layout.addWidget(self.search_entry)
        layout.addLayout(search_layout)
        
        # Students table
        self.students_table = QTableWidget()
        self.students_table.setColumnCount(5)
        self.students_table.setHorizontalHeaderLabels([
            "Select", "Reg No", "Full Name", "Grade", "Current Parent"
        ])
        self.students_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.students_table)
        
        # Relation type
        relation_layout = QHBoxLayout()
        relation_layout.addWidget(QLabel("Relation Type:"))
        self.relation_combo = QComboBox()
        self.relation_combo.addItems([
            "Primary Guardian", "Secondary Guardian", "Emergency Contact", 
            "Fee Payer", "Authorized Pickup", "Other"
        ])
        relation_layout.addWidget(self.relation_combo)
        relation_layout.addStretch()
        layout.addLayout(relation_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        link_btn = QPushButton("Link Selected Students")
        link_btn.clicked.connect(self.link_selected_students)
        link_btn.setStyleSheet("background-color: #28a745; color: white; padding: 10px;")
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("background-color: #6c757d; color: white; padding: 10px;")
        
        button_layout.addWidget(link_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)

    def load_available_students(self):
        """Load students that can be linked to this parent"""
        try:
            from models.models import get_db_connection
            self.db_connection = get_db_connection()
            self.cursor = self.db_connection.cursor(buffered=True)
            
            # Get students not already linked to this parent
            query = """
                SELECT s.id, s.regNo, s.full_name, s.grade_applied_for,
                       COALESCE(p2.full_name, 'No Parent') as current_parent
                FROM students s
                LEFT JOIN student_parent sp2 ON s.id = sp2.student_id AND sp2.is_primary_contact = TRUE
                LEFT JOIN parents p2 ON sp2.parent_id = p2.id
                WHERE s.is_active = TRUE 
                AND s.id NOT IN (
                    SELECT student_id FROM student_parent WHERE parent_id = %s
                )
                ORDER BY s.full_name
                LIMIT 100
            """
            
            self.cursor.execute(query, (self.parent_id,))
            students = self.cursor.fetchall()
            
            self.populate_students_table(students)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load students: {str(e)}")

    def populate_students_table(self, students):
        """Populate the students table with checkboxes"""
        self.students_table.setRowCount(len(students))
        
        for row_idx, student in enumerate(students):
            # Checkbox
            checkbox = QCheckBox()
            self.students_table.setCellWidget(row_idx, 0, checkbox)
            
            # Student data
            for col_idx, data in enumerate(student[1:], 1):  # Skip ID
                item = QTableWidgetItem(str(data) if data else "N/A")
                item.setData(Qt.UserRole, student[0])  # Store student ID
                self.students_table.setItem(row_idx, col_idx, item)

    def filter_students(self):
        """Filter students based on search text"""
        search_text = self.search_entry.text().lower()
        for row in range(self.students_table.rowCount()):
            show_row = False
            for col in range(1, self.students_table.columnCount()):
                item = self.students_table.item(row, col)
                if item and search_text in item.text().lower():
                    show_row = True
                    break
            self.students_table.setRowHidden(row, not show_row)

    def link_selected_students(self):
        """Link selected students to the parent"""
        try:
            selected_students = []
            for row in range(self.students_table.rowCount()):
                checkbox = self.students_table.cellWidget(row, 0)
                if checkbox and checkbox.isChecked():
                    reg_item = self.students_table.item(row, 1)
                    name_item = self.students_table.item(row, 2)
                    if reg_item and name_item:
                        student_id = reg_item.data(Qt.UserRole)
                        selected_students.append((student_id, name_item.text()))
            
            if not selected_students:
                QMessageBox.warning(self, "No Selection", "Please select at least one student to link.")
                return
            
            # Confirm linking
            student_names = ", ".join([name for _, name in selected_students])
            reply = QMessageBox.question(
                self, "Confirm Linking",
                f"Link the following students to {self.parent_name}?\n\n{student_names}",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Perform linking
            relation_type = self.relation_combo.currentText()
            for student_id, student_name in selected_students:
                link_query = """
                    INSERT INTO student_parent (student_id, parent_id, relation_type, is_primary_contact)
                    VALUES (%s, %s, %s, %s)
                """
                is_primary = (relation_type == "Primary Guardian")
                self.cursor.execute(link_query, (student_id, self.parent_id, relation_type, is_primary))
            
            self.db_connection.commit()
            
            QMessageBox.information(
                self, "Success", 
                f"Successfully linked {len(selected_students)} student(s) to {self.parent_name}!"
            )
            self.accept()
            
        except Exception as e:
            if self.db_connection:
                self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Failed to link students: {str(e)}")

    def closeEvent(self, event):
        """Clean up database connections"""
        if self.cursor:
            self.cursor.close()
        if self.db_connection and self.db_connection.is_connected():
            self.db_connection.close()
        event.accept()


# Enhanced Parents Form with student linking capability
class EnhancedParentsForm(ParentsForm):
    """Extended ParentsForm with additional features"""
    
    def link_student_to_parent(self):
        """Enhanced student linking with proper dialog"""
        if not self.current_parent_id:
            QMessageBox.warning(self, "Select", "Please select a parent first.")
            return
        
        try:
            # Get parent name
            self.cursor.execute("SELECT full_name FROM parents WHERE id = %s", (self.current_parent_id,))
            result = self.cursor.fetchone()
            if not result:
                QMessageBox.warning(self, "Error", "Selected parent not found.")
                return
                
            parent_name = result[0]
            
            # Open linking dialog
            dialog = StudentLinkDialog(self, self.current_parent_id, parent_name)
            if dialog.exec() == QDialog.Accepted:
                # Refresh the parents list to show updated student counts
                self.load_parents()
                QMessageBox.information(
                    self, "Success", 
                    "Students linked successfully! The parents list has been refreshed."
                )
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open student linking: {str(e)}")

    def export_parent_student_relationships(self):
        """Export detailed parent-student relationships"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export Relationships", 
                f"parent_student_relationships_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
                
            self.show_loading(True)
            
            # Get relationship data
            query = """
                SELECT 
                    p.id as parent_id,
                    p.full_name as parent_name,
                    p.relation as parent_relation,
                    p.phone as parent_phone,
                    p.email as parent_email,
                    s.regNo as student_reg,
                    s.full_name as student_name,
                    s.grade_applied_for as student_grade,
                    sp.relation_type,
                    CASE WHEN sp.is_primary_contact THEN 'Yes' ELSE 'No' END as is_primary,
                    sp.created_at as relationship_created
                FROM parents p
                INNER JOIN student_parent sp ON p.id = sp.parent_id
                INNER JOIN students s ON sp.student_id = s.id
                WHERE p.is_active = TRUE AND s.is_active = TRUE
                ORDER BY p.full_name, s.full_name
            """
            
            self.cursor.execute(query)
            data = self.cursor.fetchall()
            
            if not data:
                QMessageBox.information(self, "Info", "No parent-student relationships found.")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Parent-Student Relationships"
            
            headers = [
                "Parent ID", "Parent Name", "Parent Relation", "Parent Phone", "Parent Email",
                "Student Reg", "Student Name", "Student Grade", "Relationship Type", 
                "Is Primary", "Relationship Created"
            ]
            ws.append(headers)

            for row in data:
                ws.append(list(row))

            # Format the worksheet
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2F5597", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'].font = header_font
                ws[f'{col_letter}1'].fill = header_fill
                ws[f'{col_letter}1'].alignment = Alignment(horizontal="center")
                
                # Auto-adjust column width
                max_length = len(header)
                for row in ws.iter_rows(min_col=col_num, max_col=col_num, min_row=2):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            wb.save(file_path)
            QMessageBox.information(self, "Success", "Relationships data exported successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {str(e)}")
        finally:
            self.show_loading(False)


    def closeEvent(self, event):
        """Handle close event to clean up database connection"""
        try:
            if hasattr(self, 'cursor'):
                self.cursor.close()
            if hasattr(self, 'db_connection'):
                self.db_connection.close()
        except:
            pass
        event.accept()

    def __del__(self):
        """Close database connection when object is destroyed"""
        try:
            if hasattr(self, 'cursor'):
                self.cursor.close()
            if hasattr(self, 'db_connection'):
                self.db_connection.close()
        except:
            pass