# ui/students_form.py
import sys
import os
import csv
import traceback
from datetime import datetime
from typing import Optional, Dict, Any, List

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QMessageBox,
    QFileDialog, QScrollArea, QFrame, QGroupBox, QGridLayout, QComboBox,
    QFormLayout, QTabWidget, QMenu, QCheckBox, QDateEdit, QTextEdit, QApplication
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QPixmap, QIcon, QFont, QAction 
from PySide6.QtWidgets import QProgressDialog, QVBoxLayout, QHBoxLayout, QTextEdit, QPushButton, QLabel

# Import from your existing structure
from ui.audit_base_form import AuditBaseForm
from utils.permissions import has_permission
from models.models import get_db_connection  # Centralized DB connection
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import platform
import subprocess
import random
import pandas as pd
from datetime import datetime
# Define required permissions
STUDENT_PERMISSIONS = {
    'create': 'create_student',
    'edit': 'edit_student',
    'delete': 'delete_student',
    'view': 'view_student',
    'import': 'import_students',
    'manage_parents': 'manage_student_parents'
}
    

class StudentDetailsPopup(QDialog):
    """Popup window to view detailed student information with parent management"""
    def __init__(self, parent, student_id, user_session=None):
        super().__init__(parent)
        self.student_id = student_id
        self.user_session = user_session
        self.db_connection = None
        self.cursor = None
        self.setWindowTitle("Student Details")
        self.resize(1000, 700)
        self.setModal(True)
        self.setup_ui()
        self.load_student_details()

    def setup_ui(self):
        """Setup the UI components with three-column layout and centered photo"""
        layout = QVBoxLayout()
        self.setLayout(layout)
    
        # Title
        title = QLabel("Student Details")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #1f538d; padding: 10px;")
        layout.addWidget(title)
    
        # Scroll area for content
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        layout.addWidget(scroll)
    
        # Content widget
        content = QWidget()
        form_layout = QVBoxLayout(content)
        form_layout.setSpacing(15)
        form_layout.setContentsMargins(15, 15, 15, 15)
        scroll.setWidget(content)
    
        # === TOP: Three-Column Layout (Left | Photo | Right) ===
        top_layout = QHBoxLayout()
        form_layout.addLayout(top_layout)
    
        # --- LEFT COLUMN: Personal & Academic Info ---
        left_layout = QVBoxLayout()
        left_layout.setSpacing(10)
        left_layout.addWidget(QLabel("<b>Personal & Academic Information</b>"), 0, Qt.AlignTop)
        left_layout.addSpacing(10)
    
        # Grid for student info
        info_grid = QGridLayout()
        info_grid.setSpacing(8)
        info_grid.addWidget(QLabel("Full Name:"), 0, 0)
        info_grid.addWidget(self.full_name_lbl, 0, 1)
        info_grid.addWidget(QLabel("Reg No:"), 1, 0)
        info_grid.addWidget(self.reg_no_lbl, 1, 1)
        info_grid.addWidget(QLabel("Sex:"), 2, 0)
        info_grid.addWidget(self.sex_lbl, 2, 1)
        info_grid.addWidget(QLabel("Date of Birth:"), 3, 0)
        info_grid.addWidget(self.dob_lbl, 3, 1)
        info_grid.addWidget(QLabel("Email:"), 4, 0)
        info_grid.addWidget(self.email_lbl, 4, 1)
        info_grid.addWidget(QLabel("Grade Applied For:"), 5, 0)
        info_grid.addWidget(self.grade_lbl, 5, 1)
        info_grid.addWidget(QLabel("Class Year:"), 6, 0)
        info_grid.addWidget(self.class_year_lbl, 6, 1)
        info_grid.addWidget(QLabel("Enrollment Date:"), 7, 0)
        info_grid.addWidget(self.enrollment_lbl, 7, 1)
        info_grid.addWidget(QLabel("Religion:"), 8, 0)
        info_grid.addWidget(self.religion_lbl, 8, 1)
        info_grid.addWidget(QLabel("Citizenship:"), 9, 0)
        info_grid.addWidget(self.citizenship_lbl, 9, 1)
        info_grid.addWidget(QLabel("Last School:"), 10, 0)
        info_grid.addWidget(self.last_school_lbl, 10, 1)
        info_grid.addWidget(QLabel("Status:"), 11, 0)
        info_grid.addWidget(self.status_lbl, 11, 1)
    
        left_layout.addLayout(info_grid)
        left_layout.addStretch()
        top_layout.addLayout(left_layout, 4)
    
        # --- MIDDLE COLUMN: Centered Photo ---
        photo_layout = QVBoxLayout()
        photo_layout.setAlignment(Qt.AlignCenter)
        photo_layout.setSpacing(10)
    
        photo_group = QGroupBox("Photo")
        photo_group.setAlignment(Qt.AlignCenter)
        photo_inner_layout = QVBoxLayout()
        self.photo_label = QLabel("No Photo")
        self.photo_label.setAlignment(Qt.AlignCenter)
        self.photo_label.setFixedSize(140, 170)
        self.photo_label.setStyleSheet("border: 1px solid #ccc; background: #f9f9f9;")
        photo_inner_layout.addWidget(self.photo_label, alignment=Qt.AlignCenter)
        photo_group.setLayout(photo_inner_layout)
        photo_layout.addWidget(photo_group)
    
        top_layout.addLayout(photo_layout, 2)
    
        # --- RIGHT COLUMN: Medical & Allergies ---
        right_layout = QVBoxLayout()
        right_layout.setSpacing(10)
        right_layout.addWidget(QLabel("<b>Medical Information</b>"), 0, Qt.AlignTop)
        right_layout.addSpacing(10)
    
        # Medical Conditions
        medical_group = QGroupBox("Medical Conditions")
        medical_group_layout = QVBoxLayout()
        medical_group_layout.addWidget(self.medical_lbl)
        medical_group.setLayout(medical_group_layout)
        right_layout.addWidget(medical_group)
    
        # Allergies
        allergies_group = QGroupBox("Allergies")
        allergies_group_layout = QVBoxLayout()
        allergies_group_layout.addWidget(self.allergies_lbl)
        allergies_group.setLayout(allergies_group_layout)
        right_layout.addWidget(allergies_group)
    
        right_layout.addStretch()
        top_layout.addLayout(right_layout, 4)
    
        # === BOTTOM: Parents Table (Full Width) ===
        parents_group = QGroupBox("Parents / Guardians")
        parents_layout = QVBoxLayout()
        parents_group.setLayout(parents_layout)
    
        # Make table larger
        self.parents_table.setMinimumHeight(180)
        parents_layout.addWidget(self.parents_table)
    
        form_layout.addWidget(parents_group)
    
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #c82333; }
        """)
        layout.addWidget(close_btn, alignment=Qt.AlignCenter)

    def __init__(self, parent, student_id, user_session=None):
        super().__init__(parent)
        self.student_id = student_id
        self.user_session = user_session
        self.db_connection = None
        self.cursor = None

        # Initialize UI widgets
        self.full_name_lbl = QLabel()
        self.reg_no_lbl = QLabel()
        self.sex_lbl = QLabel()
        self.dob_lbl = QLabel()
        self.email_lbl = QLabel()
        self.grade_lbl = QLabel()
        self.class_year_lbl = QLabel()
        self.enrollment_lbl = QLabel()
        self.religion_lbl = QLabel()
        self.citizenship_lbl = QLabel()
        self.last_school_lbl = QLabel()
        self.status_lbl = QLabel()
        self.medical_lbl = QTextEdit()
        self.medical_lbl.setReadOnly(True)
        self.medical_lbl.setMaximumHeight(100)
        self.allergies_lbl = QTextEdit()
        self.allergies_lbl.setReadOnly(True)
        self.allergies_lbl.setMaximumHeight(100)
        self.parents_table = QTableWidget()
        self.parents_table.setColumnCount(6)
        self.parents_table.setHorizontalHeaderLabels([
            "Name", "Relation", "Phone", "Email", "Fee Payer", "Emergency Contact"
        ])
        self.parents_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.parents_table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.setWindowTitle("Student Details")
        self.resize(1000, 700)
        self.setModal(True)
        self.setup_ui()
        self.load_student_details()

    def load_student_details(self):
        """Load student details from database"""
        try:
            self.db_connection = get_db_connection()
            if not self.db_connection:
                QMessageBox.critical(self, "Error", "Failed to connect to database")
                return
            self.cursor = self.db_connection.cursor(buffered=True)

            query = '''
                SELECT s.first_name, s.surname, s.sex, s.date_of_birth, s.email,
                       s.grade_applied_for, s.class_year, s.enrollment_date, s.regNo,
                       s.religion, s.citizenship, s.last_school, s.medical_conditions, 
                       s.allergies, s.is_active, s.photo_path
                FROM students s
                WHERE s.id = %s
            '''
            self.cursor.execute(query, (self.student_id,))
            student = self.cursor.fetchone()

            if not student:
                QMessageBox.warning(self, "Error", "Student not found.")
                return

            # Format full name
            first_name = student[0] or ""
            surname = student[1] or ""
            full_name = f"{first_name} {surname}".strip()
            self.full_name_lbl.setText(full_name or "Unnamed Student")

            # Set other fields
            self.reg_no_lbl.setText(student[8] or "N/A")
            self.sex_lbl.setText(student[2] or "N/A")
            self.dob_lbl.setText(student[3].strftime("%Y-%m-%d") if student[3] else "N/A")
            self.email_lbl.setText(student[4] or "N/A")
            self.grade_lbl.setText(student[5] or "N/A")
            self.class_year_lbl.setText(student[6] or "N/A")
            self.enrollment_lbl.setText(student[7].strftime("%Y-%m-%d") if student[7] else "N/A")
            self.religion_lbl.setText(student[9] or "N/A")
            self.citizenship_lbl.setText(student[10] or "N/A")
            self.last_school_lbl.setText(student[11] or "N/A")
            self.status_lbl.setText("Active" if student[14] else "Inactive")
            self.medical_lbl.setPlainText(student[12] or "None")
            self.allergies_lbl.setPlainText(student[13] or "None")

            # Load photo
            if student[15] and os.path.exists(student[15]):
                pixmap = QPixmap(student[15])
                if not pixmap.isNull():
                    scaled = pixmap.scaled(150, 180, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.photo_label.setPixmap(scaled)
                else:
                    self.photo_label.setText("Invalid Image")
            else:
                self.photo_label.setText("No Photo")

            self.load_parents_info()

        except Exception as e:
            print(f"Error loading student details: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load student details: {str(e)}")

    def load_parents_info(self):
        """Load parent information for the student"""
        try:
            parent_query = '''
                SELECT p.id, p.full_name, sp.relation_type, p.phone, p.email,
                       sp.is_fee_payer, sp.is_emergency_contact
                FROM student_parent sp
                JOIN parents p ON sp.parent_id = p.id
                WHERE sp.student_id = %s AND p.is_active = TRUE
                ORDER BY sp.is_primary_contact DESC
            '''
            self.cursor.execute(parent_query, (self.student_id,))
            parents = self.cursor.fetchall()

            self.parents_table.setRowCount(len(parents))
            for row_idx, parent in enumerate(parents):
                parent_id = parent[0]  # Hidden ID
                for col_idx, data in enumerate(parent[1:]):  # Skip ID
                    cell = QTableWidgetItem(str(data) if data is not None else "N/A")
                    if isinstance(data, bool):
                        cell.setText("Yes" if data else "No")
                    self.parents_table.setItem(row_idx, col_idx, cell)
                    # Store parent_id in first cell
                    if col_idx == 0:
                        self.parents_table.item(row_idx, col_idx).setData(Qt.UserRole, parent_id)

        except Exception as e:
            print(f"Error loading parents: {e}")
            QMessageBox.warning(self, "Warning", "Could not load parent information.")

    def show_parent_context_menu(self, position):
        """Show context menu for parent table"""
        row = self.parents_table.rowAt(position.y())
        if row < 0:
            return

        menu = QMenu(self)
        remove_action = QAction("Remove Parent", self)
        remove_action.triggered.connect(lambda: self.remove_parent(row))
        menu.addAction(remove_action)

        menu.exec_(self.parents_table.mapToGlobal(position))

    def remove_parent(self, row):
        """Remove selected parent from student"""
        try:
            parent_item = self.parents_table.item(row, 0)
            if not parent_item:
                return
            parent_id = parent_item.data(Qt.UserRole)
            parent_name = parent_item.text()
    
            reply = QMessageBox.question(
                self,
                "Confirm Removal",
                f"Are you sure you want to remove '{parent_name}' as a parent/guardian?\n"
                "This cannot be undone.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                return
    
            # Delete from junction table
            delete_query = "DELETE FROM student_parent WHERE student_id = %s AND parent_id = %s"
            self.cursor.execute(delete_query, (self.student_id, parent_id))
            self.db_connection.commit()
    
            # Log audit
            if self.user_session:
                self.log_audit(
                    action="REMOVE_PARENT_LINK",
                    table_name="student_parent",
                    record_id=self.student_id,
                    old_value=f"Removed parent: {parent_name} (ID: {parent_id})",
                    new_value=None,
                    user_id=self.user_session.user_id,
                    school_id=self.user_session.school_id
                )
    
            self.load_parents_info()
            QMessageBox.information(self, "Success", f"Parent '{parent_name}' removed successfully.")
    
        except Exception as e:
            print(f"Error removing parent: {e}")
            QMessageBox.critical(self, "Error", f"Failed to remove parent: {str(e)}")

    def closeEvent(self, event):
        """Clean up database connection"""
        if self.cursor:
            self.cursor.close()
        if self.db_connection and self.db_connection.is_connected():
            self.db_connection.close()
        event.accept()


class StudentsForm(AuditBaseForm):
    """Main students management form"""

    def __init__(self, parent=None, user_session=None):
        super().__init__(parent, user_session)
        self.user_session = user_session
        self.current_student_id = None
        self.photo_path = None
        self.db_connection = None
        self.cursor = None

        # Use centralized DB connection
        try:
            self.db_connection = get_db_connection()
            if not self.db_connection:
                raise Exception("Failed to establish database connection")
            self.cursor = self.db_connection.cursor(buffered=True)
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Failed to connect: {str(e)}")
            return

        self.setup_ui()
        self.load_initial_data()
        self.apply_button_permissions()

    def setup_ui(self):
        """Setup the user interface"""
        self.setWindowTitle("Students Management")
        self.resize(1200, 800)

        # Main layout
        main_layout = QVBoxLayout()
        self.setLayout(main_layout)

        # Create tab widget
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # Create tabs
        self.form_tab = QWidget()
        self.list_tab = QWidget()

        self.tabs.addTab(self.form_tab, "Student Form")
        self.tabs.addTab(self.list_tab, "Students List")

        self.setup_form_tab()
        self.setup_list_tab()
        
    def setup_form_tab(self):
        """Setup the Student Form tab with two-column layout and scroll bar"""
        # Create main scroll area for the entire form tab
        form_scroll = QScrollArea()
        form_scroll.setWidgetResizable(True)
        form_scroll.setFrameShape(QFrame.NoFrame)
        
        # Create container widget for scroll area
        form_container = QWidget()
        main_container_layout = QVBoxLayout(form_container)
        form_scroll.setWidget(form_container)
        
        # Set the scroll area as the main layout for the form tab
        self.form_tab.setLayout(QVBoxLayout())
        self.form_tab.layout().addWidget(form_scroll)
        self.form_tab.layout().setContentsMargins(0, 0, 0, 0)
        
        # Add some styling for better spacing
        form_container.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ccc;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QLabel {
                font-weight: bold;
                color: #333;
            }
        """)
    
        # Main form layout inside the container
        form_layout = QVBoxLayout()
        form_layout.setContentsMargins(15, 15, 15, 15)
        form_layout.setSpacing(15)
        main_container_layout.addLayout(form_layout)
    
        # Main horizontal splitter
        main_splitter = QHBoxLayout()
        main_splitter.setSpacing(20)
        form_layout.addLayout(main_splitter)
    
        # === LEFT SIDE: Photo, Student & Academic Info ===
        left_layout = QVBoxLayout()
        left_layout.setSpacing(15)
        main_splitter.addLayout(left_layout, 6)
    
        # --- Photo Section ---
        photo_group = QGroupBox("Student Photo")
        photo_layout = QGridLayout()
        photo_layout.setContentsMargins(10, 15, 10, 10)
        photo_layout.setSpacing(10)
        photo_group.setLayout(photo_layout)
    
        self.photo_label = QLabel("No Photo")
        self.photo_label.setAlignment(Qt.AlignCenter)
        self.photo_label.setFixedSize(150, 180)
        self.photo_label.setStyleSheet("""
            border: 2px dashed #ccc;
            background-color: #f0f0f0;
            color: #888;
            font-size: 12px;
        """)
        photo_layout.addWidget(self.photo_label, 0, 0, 3, 1, Qt.AlignCenter)
    
        self.upload_btn = QPushButton("Upload")
        self.upload_btn.setMaximumWidth(100)
        self.upload_btn.clicked.connect(self.upload_photo)
        photo_layout.addWidget(self.upload_btn, 0, 1)
    
        self.remove_photo_btn = QPushButton("Remove")
        self.remove_photo_btn.setMaximumWidth(100)
        self.remove_photo_btn.clicked.connect(self.remove_photo)
        photo_layout.addWidget(self.remove_photo_btn, 1, 1)
    
        left_layout.addWidget(photo_group)
    
        # --- Student Personal Info ---
        personal_group = QGroupBox("Personal Information")
        personal_layout = QGridLayout()
        personal_layout.setContentsMargins(10, 15, 10, 10)
        personal_layout.setSpacing(10)
        personal_layout.setVerticalSpacing(12)
        personal_layout.setHorizontalSpacing(15)
        personal_group.setLayout(personal_layout)
    
        self.first_name_entry = QLineEdit()
        self.surname_entry = QLineEdit()
        self.sex_combo = QComboBox()
        self.sex_combo.addItems(["", "Male", "Female"])
        self.dob_edit = QDateEdit()
        self.dob_edit.setCalendarPopup(True)
        self.dob_edit.setDate(QDate.currentDate().addYears(-10))  # Default to 10 years ago
        self.email_entry = QLineEdit()
        self.reg_no_entry = QLineEdit()
        self.religion_entry = QLineEdit()
        self.citizenship_entry = QLineEdit()
        # === NEW: Add full_name_entry ===
        self.full_name_entry = QLineEdit()
        self.full_name_entry.setPlaceholderText("Full name will be auto-generated from First and Surname")
        self.full_name_entry.setStyleSheet("font-style: italic; color: #666;")
        self.full_name_entry.setReadOnly(True)  # Recommended: auto-generated
        # Connect first name and surname changes to update full name
        self.first_name_entry.textChanged.connect(self.update_full_name)
        self.surname_entry.textChanged.connect(self.update_full_name)
    
        personal_layout.addWidget(QLabel("First Name *"), 0, 0)
        personal_layout.addWidget(self.first_name_entry, 0, 1)
        personal_layout.addWidget(QLabel("Surname *"), 0, 2)
        personal_layout.addWidget(self.surname_entry, 0, 3)

        # --- Full Name (Row 1) - Full width ---
        personal_layout.addWidget(QLabel("Full Name *"), 1, 0)
        personal_layout.addWidget(self.full_name_entry, 1, 1, 1, 3)  # Span columns 1-3
    
        personal_layout.addWidget(QLabel("Sex"), 2, 0)
        personal_layout.addWidget(self.sex_combo, 2, 1)
        personal_layout.addWidget(QLabel("Date of Birth"), 2, 2)
        personal_layout.addWidget(self.dob_edit, 2, 3)
    
        personal_layout.addWidget(QLabel("Email"), 3, 0)
        personal_layout.addWidget(self.email_entry, 3, 1, 1, 3)
    
        # With this updated code:
        personal_layout.addWidget(QLabel("Reg No *"), 4, 0)
        
        # Create a horizontal layout for reg no and generate button
        reg_no_layout = QHBoxLayout()
        reg_no_layout.setSpacing(5)
        reg_no_layout.addWidget(self.reg_no_entry)
        
        self.generate_reg_btn = QPushButton("Generate")
        self.generate_reg_btn.setMaximumWidth(80)
        self.generate_reg_btn.setToolTip("Click to generate Registration Number")
        self.generate_reg_btn.clicked.connect(self.generate_reg_no)
        reg_no_layout.addWidget(self.generate_reg_btn)
        
        # Add the layout to the grid
        personal_layout.addLayout(reg_no_layout, 4, 1)

        personal_layout.addWidget(QLabel("Religion"), 4, 2)
        personal_layout.addWidget(self.religion_entry, 4, 3)
    
        personal_layout.addWidget(QLabel("Citizenship"), 5, 0)
        personal_layout.addWidget(self.citizenship_entry, 5, 1, 1, 3)
    
        # Add stretch to columns for better spacing
        personal_layout.setColumnStretch(1, 1)
        personal_layout.setColumnStretch(3, 1)
    
        left_layout.addWidget(personal_group)
    
        # --- Academic Information ---
        academic_group = QGroupBox("Academic Information")
        academic_layout = QGridLayout()
        academic_layout.setContentsMargins(10, 15, 10, 10)
        academic_layout.setSpacing(10)
        academic_layout.setVerticalSpacing(12)
        academic_layout.setHorizontalSpacing(15)
        academic_group.setLayout(academic_layout)
    
        self.grade_combo = QComboBox()
        self.class_year_entry = QLineEdit()
        self.enrollment_date = QDateEdit()
        self.enrollment_date.setCalendarPopup(True)
        self.enrollment_date.setDate(QDate.currentDate())
        self.last_school_entry = QLineEdit()
    
        academic_layout.addWidget(QLabel("Grade Applied For *"), 0, 0)
        academic_layout.addWidget(self.grade_combo, 0, 1)
        academic_layout.addWidget(QLabel("Class Year"), 0, 2)
        academic_layout.addWidget(self.class_year_entry, 0, 3)
    
        academic_layout.addWidget(QLabel("Enrollment Date"), 1, 0)
        academic_layout.addWidget(self.enrollment_date, 1, 1)
        academic_layout.addWidget(QLabel("Last School"), 1, 2)
        academic_layout.addWidget(self.last_school_entry, 1, 3)
    
        # Add stretch to columns for better spacing
        academic_layout.setColumnStretch(1, 1)
        academic_layout.setColumnStretch(3, 1)
    
        left_layout.addWidget(academic_group)
        left_layout.addStretch()
    
        # === RIGHT SIDE: Medical & Parent Info ===
        right_layout = QVBoxLayout()
        right_layout.setSpacing(15)
        main_splitter.addLayout(right_layout, 4)
    
        # --- Medical Information ---
        medical_group = QGroupBox("Medical Information")
        medical_layout = QVBoxLayout()
        medical_layout.setContentsMargins(10, 15, 10, 10)
        medical_layout.setSpacing(8)
        medical_group.setLayout(medical_layout)
    
        self.medical_text = QTextEdit()
        self.medical_text.setMaximumHeight(100)
        self.medical_text.setPlaceholderText("List any medical conditions (e.g., asthma, diabetes)")
        medical_layout.addWidget(QLabel("Medical Conditions"))
        medical_layout.addWidget(self.medical_text)
    
        self.allergies_text = QTextEdit()
        self.allergies_text.setMaximumHeight(100)
        self.allergies_text.setPlaceholderText("List known allergies (e.g., peanuts, penicillin)")
        medical_layout.addWidget(QLabel("Allergies"))
        medical_layout.addWidget(self.allergies_text)
    
        right_layout.addWidget(medical_group)
    
        # --- Parent/Guardian Information ---
        parent_group = QGroupBox("Parent / Guardian")
        parent_layout = QVBoxLayout()
        parent_layout.setContentsMargins(10, 15, 10, 10)
        parent_layout.setSpacing(10)
        parent_group.setLayout(parent_layout)
    
        search_row = QHBoxLayout()
        search_row.setSpacing(10)
        self.parent_search_entry = QLineEdit()
        self.parent_search_entry.setPlaceholderText("Search parent...")
        search_parent_btn = QPushButton("Search")
        search_row.addWidget(self.parent_search_entry)
        search_row.addWidget(search_parent_btn)
        parent_layout.addLayout(search_row)
    
        # Add scroll area for parent table
        parent_scroll = QScrollArea()
        parent_scroll.setWidgetResizable(True)
        parent_scroll.setMinimumHeight(200)
        parent_scroll.setMaximumHeight(300)
        
        parent_table_container = QWidget()
        parent_table_layout = QVBoxLayout(parent_table_container)
        parent_table_layout.setContentsMargins(0, 0, 0, 0)
        
        self.parents_table = QTableWidget()
        self.parents_table.setColumnCount(5)
        self.parents_table.setHorizontalHeaderLabels(["Name", "Relation", "Phone", "Email", "Actions"])
        self.parents_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.parents_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        parent_table_layout.addWidget(self.parents_table)
        parent_scroll.setWidget(parent_table_container)
        parent_layout.addWidget(parent_scroll)
    
        self.add_parent_btn = QPushButton("+ Add Parent")
        self.add_parent_btn.setStyleSheet("font-weight: bold; color: green;")
        self.add_parent_btn.clicked.connect(self.add_parent_to_student)
        parent_layout.addWidget(self.add_parent_btn)
    
        right_layout.addWidget(parent_group)
    
        # --- Active Status & Action Buttons ---
        bottom_row = QHBoxLayout()
        bottom_row.setSpacing(15)
        
        self.is_active_check = QCheckBox("Is Active")
        self.is_active_check.setChecked(True)
        bottom_row.addWidget(self.is_active_check)
        
        # Save Button
        self.save_btn = QPushButton("Save Student")
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.save_btn.clicked.connect(self.save_student)
        bottom_row.addWidget(self.save_btn)
        
        # Update Button (hidden by default)
        self.update_btn = QPushButton("Update Student")
        self.update_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        self.update_btn.clicked.connect(self.update_student)
        self.update_btn.hide()  # Hidden initially
        bottom_row.addWidget(self.update_btn)
        
        bottom_row.addStretch()
        form_layout.addLayout(bottom_row)

    def setup_list_tab(self):
        """Setup the list tab for viewing students with scroll bar"""
        # Create main scroll area for the list tab
        list_scroll = QScrollArea()
        list_scroll.setWidgetResizable(True)
        list_scroll.setFrameShape(QFrame.NoFrame)
        
        # Create container widget for scroll area
        list_container = QWidget()
        container_layout = QVBoxLayout(list_container)
        container_layout.setContentsMargins(15, 15, 15, 15)
        container_layout.setSpacing(15)
        
        list_scroll.setWidget(list_container)
        
        # Set the scroll area as the main layout for the list tab
        self.list_tab.setLayout(QVBoxLayout())
        self.list_tab.layout().addWidget(list_scroll)
        self.list_tab.layout().setContentsMargins(0, 0, 0, 0)
    
        # Search section
        search_layout = QHBoxLayout()
        search_layout.setSpacing(10)
        search_layout.addWidget(QLabel("Search Students:"))
    
        self.list_search_entry = QLineEdit()
        self.list_search_entry.setPlaceholderText("Search by name, reg no, or parent...")
        self.list_search_entry.textChanged.connect(self.search_students_list)
        search_layout.addWidget(self.list_search_entry)
    
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.refresh_students)
        search_layout.addWidget(refresh_btn)
    
        clear_list_search_btn = QPushButton("Clear")
        clear_list_search_btn.clicked.connect(self.clear_list_search)
        search_layout.addWidget(clear_list_search_btn)
    
        container_layout.addLayout(search_layout)
    
        # Action buttons
        # Action buttons
        action_layout = QHBoxLayout()
        action_layout.setSpacing(10)
        
        # Edit Selected Button
        self.edit_selected_btn = QPushButton("Edit Selected")
        self.edit_selected_btn.setStyleSheet("QPushButton { background-color: #28a745; color: white; }")
        self.edit_selected_btn.clicked.connect(self.edit_selected_student)
        action_layout.addWidget(self.edit_selected_btn)
    
        view_btn = QPushButton("View Details")
        view_btn.setStyleSheet("QPushButton { background-color: #17a2b8; color: white; }")
        view_btn.clicked.connect(self.view_student_details)
        action_layout.addWidget(view_btn)
    
        export_excel_btn = QPushButton("Export to Excel")
        export_excel_btn.setStyleSheet("QPushButton { background-color: #28a745; color: white; }")
        export_excel_btn.clicked.connect(self.export_students_data)
        action_layout.addWidget(export_excel_btn)
    
        export_pdf_btn = QPushButton("Generate PDF")
        export_pdf_btn.setStyleSheet("QPushButton { background-color: #dc3545; color: white; }")
        export_pdf_btn.clicked.connect(self.generate_pdf_report)
        action_layout.addWidget(export_pdf_btn)
        
        self.import_btn = QPushButton("Import Students")
        self.import_btn.setStyleSheet("QPushButton { background-color: #ffc107; color: black; font-weight: bold; }")
        self.import_btn.clicked.connect(self.import_students)
        action_layout.addWidget(self.import_btn)

        # Add Delete button
        # Delete button
        self.delete_btn_list = QPushButton("Delete Selected")
        self.delete_btn_list.setStyleSheet("QPushButton { background-color: #dc3545; color: white; }")
        self.delete_btn_list.clicked.connect(self.delete_student)
        action_layout.addWidget(self.delete_btn_list)

        action_layout.addStretch()
        container_layout.addLayout(action_layout)
    
        # Students table with its own scroll area
        table_scroll = QScrollArea()
        table_scroll.setWidgetResizable(True)
        table_scroll.setMinimumHeight(400)
        
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)
        
        self.students_table = QTableWidget()
        self.students_table.setColumnCount(10)
        self.students_table.setHorizontalHeaderLabels([
            "ID", "Reg No", "Name", "Sex", "Grade", "Class Year",
            "Status", "Email", "Enrollment", "Parent"
        ])
    
        header = self.students_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
    
        self.students_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.students_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.students_table.cellClicked.connect(self.on_student_select)
        self.students_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.students_table.customContextMenuRequested.connect(self.show_context_menu)
    
        table_layout.addWidget(self.students_table)
        table_scroll.setWidget(table_container)
        container_layout.addWidget(table_scroll)

    def show_context_menu(self, position):
        """Show context menu for table"""
        if self.current_student_id:
            context_menu = QMenu()

            view_action = QAction("View Details", self)
            view_action.triggered.connect(self.view_student_details)
            context_menu.addAction(view_action)

            edit_action = QAction("Edit Student", self)
            edit_action.triggered.connect(self.edit_selected_student)
            context_menu.addAction(edit_action)

            context_menu.addSeparator()

            delete_action = QAction("Delete Student", self)
            delete_action.triggered.connect(self.delete_student)
            context_menu.addAction(delete_action)

            context_menu.exec_(self.students_table.mapToGlobal(position))

    def load_initial_data(self):
        """Load initial data for dropdowns and tables"""
        if not self.cursor:
            return
        self.load_grades()
        self.load_students()

    def apply_button_permissions(self):
        """Enable/disable buttons based on user permissions"""
        can_create = has_permission(self.user_session, STUDENT_PERMISSIONS['create'])
        can_edit = has_permission(self.user_session, STUDENT_PERMISSIONS['edit'])
        can_delete = has_permission(self.user_session, STUDENT_PERMISSIONS['delete'])
        can_import = has_permission(self.user_session, STUDENT_PERMISSIONS['import'])
        can_manage_parents = has_permission(self.user_session, STUDENT_PERMISSIONS['manage_parents'])
    
        # --- Form Tab Buttons ---
        self.save_btn.setEnabled(can_create)
        self.update_btn.setEnabled(can_edit)
        self.add_parent_btn.setEnabled(can_manage_parents)
    
        # --- List Tab Buttons ---
        self.import_btn.setEnabled(can_import)
        self.delete_btn_list.setEnabled(can_delete)  # ✅ Correct name
        self.edit_selected_btn.setEnabled(can_edit)
    
        # Tooltips
        self.save_btn.setToolTip("Add new student" if can_create else "Requires: create_student")
        self.update_btn.setToolTip("Edit student" if can_edit else "Requires: edit_student")
        self.add_parent_btn.setToolTip("Link parent" if can_manage_parents else "Requires: manage_student_parents")
        self.import_btn.setToolTip("Import students" if can_import else "Requires: import_students")
        self.delete_btn_list.setToolTip("Delete student" if can_delete else "Requires: delete_student")
        self.edit_selected_btn.setToolTip("Edit student" if can_edit else "Requires: edit_student")

    def load_grades(self):
        """Load grades for dropdown"""
        try:
            self.cursor.execute("""
                SELECT DISTINCT grade_applied_for 
                FROM students 
                WHERE grade_applied_for IS NOT NULL AND is_active = TRUE
                ORDER BY grade_applied_for
            """)
            grades = [row[0] for row in self.cursor.fetchall()]
            default_grades = ["S1", "S2", "S3", "S4", "S5", "S6"]

            self.grade_combo.clear()
            self.grade_combo.addItems(grades or default_grades)

        except Exception as e:
            print(f"Error loading grades: {e}")
            self.grade_combo.clear()
            self.grade_combo.addItems(default_grades)

    def toggle_save_update_buttons(self):
        """Show Save or Update button based on whether we're editing"""
        if self.current_student_id:
            self.save_btn.hide()
            self.update_btn.show()
        else:
            self.save_btn.show()
            self.update_btn.hide()


    def load_students(self):
        """Load students into the table"""
        try:
            query = '''
                SELECT s.id, s.regNo, s.first_name, s.surname, s.sex, 
                       s.grade_applied_for, s.class_year, s.is_active, 
                       s.email, s.enrollment_date,
                       GROUP_CONCAT(DISTINCT p.full_name SEPARATOR ', ') as parents
                FROM students s
                LEFT JOIN student_parent sp ON s.id = sp.student_id
                LEFT JOIN parents p ON sp.parent_id = p.id AND p.is_active = TRUE
                WHERE s.is_active = TRUE
                GROUP BY s.id
                ORDER BY s.surname, s.first_name
                LIMIT 200
            '''
            self.cursor.execute(query)
            students = self.cursor.fetchall()
            
            self.update_students_table(students)
            
        except Exception as e:
            print(f"Error loading students: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load students: {str(e)}")

    def update_students_table(self, students):
        """Update the students table with data"""
        self.students_table.setRowCount(len(students))
        
        for row_idx, student in enumerate(students):
            # ID
            self.students_table.setItem(row_idx, 0, QTableWidgetItem(str(student[0])))
            
            # Reg No
            self.students_table.setItem(row_idx, 1, QTableWidgetItem(student[1] or ""))
            
            # Full Name
            full_name = f"{student[2] or ''} {student[3] or ''}".strip()
            self.students_table.setItem(row_idx, 2, QTableWidgetItem(full_name))
            
            # Sex
            self.students_table.setItem(row_idx, 3, QTableWidgetItem(student[4] or ""))
            
            # Grade
            self.students_table.setItem(row_idx, 4, QTableWidgetItem(student[5] or ""))
            
            # Class Year
            self.students_table.setItem(row_idx, 5, QTableWidgetItem(student[6] or ""))
            
            # Status
            status = "Active" if student[7] else "Inactive"
            self.students_table.setItem(row_idx, 6, QTableWidgetItem(status))
            
            # Email
            self.students_table.setItem(row_idx, 7, QTableWidgetItem(student[8] or ""))
            
            # Enrollment Date
            enrollment_date = str(student[9]) if student[9] else ""
            self.students_table.setItem(row_idx, 8, QTableWidgetItem(enrollment_date))
            
            # Parents
            self.students_table.setItem(row_idx, 9, QTableWidgetItem(student[10] or ""))

    def update_full_name(self):
        """Update full name when first name or surname changes"""
        first_name = self.first_name_entry.text().strip()
        surname = self.surname_entry.text().strip()
        full_name = f"{first_name} {surname}".strip()
        self.full_name_entry.setText(full_name)

    def upload_photo(self):
        """Upload student photo"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Select Student Photo",
            "",
            "Image Files (*.jpg *.jpeg *.png *.bmp *.gif)"
        )
        
        if file_path:
            try:
                # Load and display the photo
                pixmap = QPixmap(file_path)
                if not pixmap.isNull():
                    # Scale the image to fit the label
                    scaled_pixmap = pixmap.scaled(
                        150, 150, 
                        Qt.KeepAspectRatio, 
                        Qt.SmoothTransformation
                    )
                    self.photo_label.setPixmap(scaled_pixmap)
                    self.photo_path = file_path
                else:
                    QMessageBox.warning(self, "Error", "Failed to load image file")
                    
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to load image: {str(e)}")

    def remove_photo(self):
        """Remove student photo"""
        self.photo_label.clear()
        self.photo_label.setText("No Photo")
        self.photo_path = None

    def get_school_prefix(self):
        """Get school prefix from database based on school name"""
        try:
            # Query your schools table - adjust column names as needed
            query = """
                SELECT school_name FROM schools 
                WHERE id = %s AND is_active = TRUE
                LIMIT 1
            """
            
            # Get school_id from user session or use default
            school_id = getattr(self.user_session, 'school_id', 1) if self.user_session else 1
            
            self.cursor.execute(query, (school_id,))
            result = self.cursor.fetchone()
            
            if result and result[0]:
                school_name = result[0].strip().upper()
                
                # Extract first letters of each word
                words = school_name.split()
                if len(words) >= 2:
                    # Take first letter of each word (up to 4)
                    prefix_chars = []
                    for word in words:
                        if word and len(prefix_chars) < 4:
                            prefix_chars.append(word[0])
                    
                    # If we need more characters, take from first word
                    if len(prefix_chars) < 4 and words:
                        first_word = words[0]
                        for i in range(len(prefix_chars), min(4, len(first_word))):
                            prefix_chars.append(first_word[i])
                    
                    # Still need more? Repeat first letter
                    while len(prefix_chars) < 4:
                        prefix_chars.append(words[0][0])
                    
                    return ''.join(prefix_chars[:4])
                else:
                    # Single word - take first 4 characters
                    first_word = words[0] if words else "SCHL"
                    if len(first_word) >= 4:
                        return first_word[:4]
                    else:
                        return (first_word + first_word[0] * 4)[:4]
            
            return "SCHL"  # Default fallback
            
        except Exception as e:
            print(f"Error fetching school prefix: {e}")
            return "SCHL"

    def generate_reg_no(self):
        """Generate registration number in format PPPPYYXXXX-ST where PPPP is school prefix"""
        try:
            # Get dynamic school prefix from database
            school_prefix = self.get_school_prefix()
            
            # Get current year (last 2 digits)
            year_suffix = datetime.now().year % 100
            year_str = f"{year_suffix:02d}"
            
            # Generate random number between 1000 and 9999
            rand_num = random.randint(1000, 9999)
            
            # Format: SCHOOLPREFIX + YEAR + RANDOM + SUFFIX
            reg_no = f"{school_prefix}{year_str}{rand_num}-ST"
            self.reg_no_entry.setText(reg_no)
            
        except Exception as e:
            print(f"Error generating registration number: {e}")
            # Fallback with current year
            year_suffix = datetime.now().year % 100
            fallback_reg_no = f"SCHL{year_suffix:02d}0000-ST"
            self.reg_no_entry.setText(fallback_reg_no)

    def add_parent_to_student(self):
        if not has_permission(self.user_session, STUDENT_PERMISSIONS['manage_parents']):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to link parents.")
            return
        """Open dialog to add or select parent for the student"""
        if not self.current_student_id:
            QMessageBox.warning(self, "No Student", "Please save the student first before adding a parent.")
            return
    
        dialog = QDialog(self)
        dialog.setWindowTitle("Link Parent / Guardian")
        dialog.resize(900, 550)
        dialog.setModal(True)
    
        layout = QVBoxLayout(dialog)
    
        # Search bar
        search_layout = QHBoxLayout()
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("Search parents by name, phone, or email...")
        search_btn = QPushButton("Search")
        search_layout.addWidget(search_edit)
        search_layout.addWidget(search_btn)
        layout.addLayout(search_layout)
    
        # Parents table
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Name", "Relation", "Phone", "Email", "Select"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(table)
    
        # Role checkboxes
        role_group = QGroupBox("Parent Roles")
        role_layout = QGridLayout()
        role_group.setLayout(role_layout)
    
        is_primary_check = QCheckBox("Primary Contact")
        is_primary_check.setChecked(True)
        is_payer_check = QCheckBox("Fee Payer")
        is_payer_check.setChecked(True)
        is_emergency_check = QCheckBox("Emergency Contact")
        is_emergency_check.setChecked(True)
    
        relation_edit = QLineEdit("Guardian")
        relation_edit.setPlaceholderText("e.g., Father, Mother, Guardian")
    
        role_layout.addWidget(QLabel("Relation:"), 0, 0)
        role_layout.addWidget(relation_edit, 0, 1)
        role_layout.addWidget(is_primary_check, 1, 0)
        role_layout.addWidget(is_payer_check, 1, 1)
        role_layout.addWidget(is_emergency_check, 2, 0, 1, 2)
    
        layout.addWidget(role_group)
    
        # Buttons
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Link Parent")
        cancel_btn = QPushButton("Cancel")
        btn_layout.addStretch()
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
    
        # Load parents
        def load_parents():
            search_term = search_edit.text().strip()
            table.setRowCount(0)
            try:
                if search_term:
                    query = """
                        SELECT id, full_name, relation, phone, email
                        FROM parents
                        WHERE is_active = TRUE
                        AND (full_name LIKE %s OR phone LIKE %s OR email LIKE %s)
                        ORDER BY full_name
                    """
                    params = [f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"]
                else:
                    query = """
                        SELECT id, full_name, relation, phone, email
                        FROM parents
                        WHERE is_active = TRUE
                        ORDER BY full_name
                    """
                    params = []
    
                self.cursor.execute(query, params)
                parents = self.cursor.fetchall()
    
                table.setRowCount(len(parents))
                for row_idx, parent in enumerate(parents):
                    for col_idx, data in enumerate(parent[1:]):  # Skip ID
                        table.setItem(row_idx, col_idx, QTableWidgetItem(str(data) if data else "N/A"))
    
                    # Select button
                    select_btn = QPushButton("Select")
                    select_btn.setProperty("parent_id", parent[0])
                    select_btn.setProperty("full_name", parent[1])
                    select_btn.setProperty("relation", parent[2])
                    select_btn.clicked.connect(lambda _, btn=select_btn: on_parent_selected(btn))
                    table.setCellWidget(row_idx, 4, select_btn)
    
            except Exception as e:
                QMessageBox.critical(dialog, "Error", f"Failed to load parents: {str(e)}")
    
        def on_parent_selected(button):
            self.selected_parent_id = button.property("parent_id")
            self.selected_parent_name = button.property("full_name")
            relation_edit.setText(button.property("relation") or "Guardian")
            QMessageBox.information(dialog, "Selected", f"Parent selected: {self.selected_parent_name}")
    
        search_btn.clicked.connect(load_parents)
        load_parents()  # Load on open
    
        # Handle save
        def on_save():
            if not hasattr(self, 'selected_parent_id'):
                QMessageBox.warning(dialog, "No Selection", "Please select a parent first.")
                return
    
            relation = relation_edit.text().strip()
            is_primary = is_primary_check.isChecked()
            is_payer = is_payer_check.isChecked()
            is_emergency = is_emergency_check.isChecked()
    
            if not relation:
                QMessageBox.warning(dialog, "Input Required", "Please enter the relation type.")
                return
    
            try:
                # Prevent duplicate links
                self.cursor.execute("""
                    SELECT id FROM student_parent
                    WHERE student_id = %s AND parent_id = %s
                """, (self.current_student_id, self.selected_parent_id))
                if self.cursor.fetchone():
                    QMessageBox.warning(dialog, "Already Linked", "This parent is already linked to the student.")
                    return
    
                # If marking as primary, remove primary from others
                if is_primary:
                    self.cursor.execute("""
                        UPDATE student_parent SET is_primary_contact = FALSE
                        WHERE student_id = %s AND is_primary_contact = TRUE
                    """, (self.current_student_id,))
    
                # Insert new relationship
                self.cursor.execute("""
                    INSERT INTO student_parent (
                        student_id, parent_id, relation_type,
                        is_primary_contact, is_fee_payer, is_emergency_contact
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (
                    self.current_student_id,
                    self.selected_parent_id,
                    relation,
                    is_primary,
                    is_payer,
                    is_emergency
                ))
                self.db_connection.commit()
    
                QMessageBox.information(dialog, "Success", "Parent linked successfully!")
                dialog.accept()
                self.load_student_parents()  # Refresh parent list in form
            except Exception as e:
                self.db_connection.rollback()
                QMessageBox.critical(dialog, "Error", f"Failed to link parent: {str(e)}")
    
        save_btn.clicked.connect(on_save)
        cancel_btn.clicked.connect(dialog.reject)
    
        dialog.exec()

    def search_students(self):
        """Search students from form tab"""
        search_term = self.search_entry.text().strip()
        
        if not search_term:
            self.load_students()
            return
        
        self.search_students_common(search_term)

    def search_students_list(self):
        """Search students from list tab"""
        search_term = self.list_search_entry.text().strip()
        
        if not search_term:
            self.load_students()
            return
        
        self.search_students_common(search_term)

    def refresh_students(self):
        """Refresh student list and show confirmation"""
        self.load_students()
        QMessageBox.information(self, "Refresh Complete", 
                              "Student data has been loaded and refreshed successfully!")

    def search_students_common(self, search_term):
        """Common search functionality"""
        try:
            query = '''
                SELECT s.id, s.regNo, s.first_name, s.surname, s.sex, 
                       s.grade_applied_for, s.class_year, s.is_active, 
                       s.email, s.enrollment_date,
                       GROUP_CONCAT(DISTINCT p.full_name SEPARATOR ', ') as parents
                FROM students s
                LEFT JOIN student_parent sp ON s.id = sp.student_id
                LEFT JOIN parents p ON sp.parent_id = p.id AND p.is_active = TRUE
                WHERE s.is_active = TRUE
                AND (s.first_name LIKE %s OR s.surname LIKE %s OR s.regNo LIKE %s 
                     OR p.full_name LIKE %s)
                GROUP BY s.id
                ORDER BY s.surname, s.first_name
                LIMIT 100
            '''
            like_term = f"%{search_term}%"
            self.cursor.execute(query, (like_term, like_term, like_term, like_term))
            students = self.cursor.fetchall()
            
            self.update_students_table(students)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Search failed: {str(e)}")

    def clear_search(self):
        """Clear search in form tab"""
        self.search_entry.clear()
        self.load_students()

    def clear_list_search(self):
        """Clear search in list tab"""
        self.list_search_entry.clear()
        self.load_students()

    def on_student_select(self, row, col):
        """Handle student selection from table"""
        try:
            item = self.students_table.item(row, 0)  # Get ID from first column
            if item:
                self.current_student_id = int(item.text())
                print(f"Selected student ID: {self.current_student_id}")
        except Exception as e:
            print(f"Error selecting student: {e}")

    def edit_selected_student(self):
        """Load selected student for editing"""
        if not self.current_student_id:
            QMessageBox.warning(self, "No Selection", "Please select a student from the table first.")
            return
        
        self.load_student_details(self.current_student_id)
        self.tabs.setCurrentIndex(0)  # Switch to form tab
        
        # Get student name for confirmation
        student_name = self.full_name_entry.text().strip()
        QMessageBox.information(self, "Student Loaded", 
                              f"Student '{student_name}' loaded for editing.")

    def load_student_details(self, student_id):
        """Load student details into form for editing"""
        try:
            query = '''
                SELECT first_name, surname, sex, date_of_birth, email, 
                       grade_applied_for, class_year, enrollment_date, regNo, 
                       religion, citizenship, last_school, medical_conditions, 
                       allergies, is_active, photo_path
                FROM students 
                WHERE id = %s
            '''
            self.cursor.execute(query, (student_id,))
            student = self.cursor.fetchone()
            
            if not student:
                QMessageBox.warning(self, "Error", "Student not found")
                return
            
            # Clear form first
            self.clear_form()
            
            # Set current student ID
            self.current_student_id = student_id
            
            # Populate form fields
            self.first_name_entry.setText(student[0] or "")
            self.surname_entry.setText(student[1] or "")
            self.sex_combo.setCurrentText(student[2] or "Male")
            
            # Handle date of birth
            if student[3]:
                if isinstance(student[3], str):
                    dob = QDate.fromString(student[3], "yyyy-MM-dd")
                else:
                    dob = QDate(student[3])
                self.dob_edit.setDate(dob)
            
            self.email_entry.setText(student[4] or "")
            self.grade_combo.setCurrentText(student[5] or "")
            self.class_year_entry.setText(student[6] or "")
            
            # Handle enrollment date
            if student[7]:
                if isinstance(student[7], str):
                    enrollment = QDate.fromString(student[7], "yyyy-MM-dd")
                else:
                    enrollment = QDate(student[7])
                self.enrollment_date.setDate(enrollment)
            
            self.reg_no_entry.setText(student[8] or "")
            self.religion_entry.setText(student[9] or "")
            self.citizenship_entry.setText(student[10] or "")
            self.last_school_entry.setText(student[11] or "")
            self.medical_text.setPlainText(student[12] or "")
            self.allergies_text.setPlainText(student[13] or "")
            self.is_active_check.setChecked(bool(student[14]))
            
            # Load photo if exists
            if student[15] and os.path.exists(student[15]):
                pixmap = QPixmap(student[15])
                scaled_pixmap = pixmap.scaled(150, 150, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.photo_label.setPixmap(scaled_pixmap)
                self.photo_path = student[15]
            
            # Load associated parent
            self.load_student_parents()
            # Toggle buttons
            self.toggle_save_update_buttons()
            
        except Exception as e:
            print(f"Error loading student details: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load student details: {str(e)}")

    def load_student_parents(self):
        """Load all parents linked to the current student"""
        if not self.current_student_id:
            self.parents_table.setRowCount(0)
            return
        try:
            query = """
                SELECT p.id, p.full_name, sp.relation_type, p.phone, p.email
                FROM student_parent sp
                JOIN parents p ON sp.parent_id = p.id
                WHERE sp.student_id = %s AND p.is_active = TRUE
                ORDER BY sp.is_primary_contact DESC
            """
            self.cursor.execute(query, (self.current_student_id,))
            parents = self.cursor.fetchall()
    
            self.parents_table.setRowCount(len(parents))
            for row_idx, parent in enumerate(parents):
                parent_id = parent[0]  # Store parent_id
                # Name, Relation, Phone, Email
                for col_idx, data in enumerate(parent[1:]):
                    self.parents_table.setItem(row_idx, col_idx, QTableWidgetItem(str(data) if data else "N/A"))
    
                # Action Button (Remove)
                remove_btn = QPushButton("Remove")
                remove_btn.setProperty("parent_id", parent_id)
                remove_btn.setProperty("parent_name", parent[1])
                remove_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #dc3545;
                        color: white;
                        border: none;
                        padding: 3px 10px;
                        border-radius: 3px;
                        font-size: 11px;
                    }
                    QPushButton:hover {
                        background-color: #c82333;
                    }
                """)
                remove_btn.clicked.connect(lambda _, btn=remove_btn: self.remove_parent_link(btn))
                self.parents_table.setCellWidget(row_idx, 4, remove_btn)  # Actions column
    
        except Exception as e:
            print(f"Error loading student parents: {e}")
            QMessageBox.warning(self, "Warning", "Could not load parent list.")

    def remove_parent_link(self, button):
        """Remove the link between student and parent"""
        parent_id = button.property("parent_id")
        parent_name = button.property("parent_name")
    
        reply = QMessageBox.question(
            self,
            "Confirm Removal",
            f"Are you sure you want to remove '{parent_name}' as a parent/guardian?\n"
            "This will not delete the parent, only the link.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
    
        # Prevent removing last parent
        self.cursor.execute("SELECT COUNT(*) FROM student_parent WHERE student_id = %s", (self.current_student_id,))
        count = self.cursor.fetchone()[0]
        if count <= 1:
            QMessageBox.warning(self, "Cannot Remove", "A student must have at least one parent linked.")
            return
    
        try:
            # Delete from junction table
            query = "DELETE FROM student_parent WHERE student_id = %s AND parent_id = %s"
            self.cursor.execute(query, (self.current_student_id, parent_id))
            self.db_connection.commit()
    
            # ✅ Log audit using correct method
            if self.user_session:
                self.log_audit_action(
                    action="REMOVE_PARENT_LINK",
                    table_name="student_parent",
                    record_id=self.current_student_id,
                    description=f"Removed parent: {parent_name} (ID: {parent_id})"
                )
    
            # Refresh the table
            self.load_student_parents()
    
            QMessageBox.information(self, "Success", f"Parent '{parent_name}' removed successfully.")
    
        except Exception as e:
            self.db_connection.rollback()
            print(f"Error removing parent link: {e}")
            QMessageBox.critical(self, "Error", f"Failed to remove parent: {str(e)}")
        
    def view_student_details(self):
        """Show student details popup"""
        if not has_permission(self.user_session, STUDENT_PERMISSIONS['view']):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to view student details.")
            return
        if not self.current_student_id:
            QMessageBox.warning(self, "No Selection", "Please select a student from the table first.")
            return
        try:
            popup = StudentDetailsPopup(self, self.current_student_id, self.user_session)
            popup.exec_()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to show student details: {str(e)}")

    def save_student(self):
        """Save new student to database"""
        # 🔒 Check permission first
        if not has_permission(self.user_session, STUDENT_PERMISSIONS['create']):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to add students.")
            return
    
        # ✅ Validate form
        if not self.validate_form():
            return
    
        # ❌ Prevent saving again if editing
        if self.current_student_id:
            QMessageBox.information(self, "Info", "Student already exists. Use Update instead.")
            return
    
        try:
            # ✅ Prepare data
            first_name = self.first_name_entry.text().strip()
            surname = self.surname_entry.text().strip()
            full_name = f"{first_name} {surname}".strip()
            reg_no = self.reg_no_entry.text().strip()
            dob = self.dob_edit.date().toPython()
            enrollment = self.enrollment_date.date().toPython()
    
            # ✅ Insert student
            query = '''
                INSERT INTO students (
                    school_id, first_name, surname, full_name, sex, date_of_birth,
                    email, grade_applied_for, class_year, enrollment_date, regNo,
                    religion, citizenship, last_school, medical_conditions, allergies,
                    is_active, photo_path
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            '''
            school_id = 1  # Default school ID - modify as needed
            values = (
                school_id, first_name, surname, full_name, self.sex_combo.currentText(),
                dob, self.email_entry.text().strip(), self.grade_combo.currentText(),
                self.class_year_entry.text().strip(), enrollment, reg_no,
                self.religion_entry.text().strip(), self.citizenship_entry.text().strip(),
                self.last_school_entry.text().strip(), self.medical_text.toPlainText(),
                self.allergies_text.toPlainText(), self.is_active_check.isChecked(),
                self.photo_path
            )
            self.cursor.execute(query, values)
            student_id = self.cursor.lastrowid  # ✅ Now available
            self.current_student_id = student_id
    
            # ✅ Commit first
            self.db_connection.commit()
    
            # ✅ Log audit AFTER successful commit
            if self.user_session:
                self.log_audit_action(
                    action="CREATE",
                    table_name="students",
                    record_id=student_id,
                    description=f"Added student: {full_name} (RegNo: {reg_no})"
                )
    
            # ✅ Success feedback
            QMessageBox.information(self, "Success", "Student saved successfully!")
            self.clear_form()
            self.load_students()
    
        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Failed to save student: {str(e)}")
        
    def update_student(self):
        """Update existing student"""
        # 🔒 Permission check
        if not has_permission(self.user_session, STUDENT_PERMISSIONS['edit']):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to edit students.")
            return
    
        if not self.current_student_id:
            QMessageBox.warning(self, "Error", "No student selected for update")
            return
    
        if not self.validate_form():
            return
    
        try:
            # Prepare data
            first_name = self.first_name_entry.text().strip()
            surname = self.surname_entry.text().strip()
            full_name = f"{first_name} {surname}".strip()
            reg_no = self.reg_no_entry.text().strip()
            dob = self.dob_edit.date().toPython()
            enrollment = self.enrollment_date.date().toPython()
    
            # Update student
            query = '''
                UPDATE students SET
                    first_name = %s, surname = %s, full_name = %s, sex = %s,
                    date_of_birth = %s, email = %s, grade_applied_for = %s,
                    class_year = %s, enrollment_date = %s, regNo = %s,
                    religion = %s, citizenship = %s, last_school = %s,
                    medical_conditions = %s, allergies = %s, is_active = %s,
                    photo_path = %s
                WHERE id = %s
            '''
            values = (
                first_name, surname, full_name, self.sex_combo.currentText(),
                dob, self.email_entry.text().strip(), self.grade_combo.currentText(),
                self.class_year_entry.text().strip(), enrollment, reg_no,
                self.religion_entry.text().strip(), self.citizenship_entry.text().strip(),
                self.last_school_entry.text().strip(), self.medical_text.toPlainText(),
                self.allergies_text.toPlainText(), self.is_active_check.isChecked(),
                self.photo_path, self.current_student_id
            )
            self.cursor.execute(query, values)
            self.db_connection.commit()
    
            # ✅ Audit log
            if self.user_session:
                self.log_audit_action(
                    action="UPDATE",
                    table_name="students",
                    record_id=self.current_student_id,
                    description=f"Updated student: {full_name} (RegNo: {reg_no})"
                )
    
            QMessageBox.information(self, "Success", "Student updated successfully!")
            self.load_students()
            self.clear_form()
    
        except Exception as e:
            self.db_connection.rollback()
            QMessageBox.critical(self, "Error", f"Failed to update student: {str(e)}")
        
    def delete_student(self):
        """Delete (soft delete) selected student"""
        # 🔒 Permission check
        if not has_permission(self.user_session, STUDENT_PERMISSIONS['delete']):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to delete students.")
            return
    
        if not self.current_student_id:
            QMessageBox.warning(self, "Error", "No student selected for deletion")
            return
    
        # Get student name for confirmation
        first_name = self.first_name_entry.text().strip()
        surname = self.surname_entry.text().strip()
        full_name = f"{first_name} {surname}".strip()
        reg_no = self.reg_no_entry.text().strip()
    
        # Confirmation dialog
        reply = QMessageBox.question(
            self, "Confirm Deletion",
            f"Are you sure you want to delete '{full_name}' (RegNo: {reg_no})?\n"
            "This will mark the student as inactive but preserve their records.\n"
            "This action can be reversed later.\n\n"
            "This cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
    
        if reply == QMessageBox.Yes:
            try:
                # Soft delete
                query = "UPDATE students SET is_active = FALSE WHERE id = %s"
                self.cursor.execute(query, (self.current_student_id,))
                self.db_connection.commit()
    
                # ✅ Audit log
                if self.user_session:
                    self.log_audit_action(
                        action="DELETE",
                        table_name="students",
                        record_id=self.current_student_id,
                        description=f"Deactivated student: {full_name} (RegNo: {reg_no})"
                    )
    
                QMessageBox.information(self, "Success", "Student deleted successfully!")
                self.clear_form()
                self.load_students()
    
            except Exception as e:
                self.db_connection.rollback()
                QMessageBox.critical(self, "Error", f"Failed to delete student: {str(e)}")

    def link_student_parent(self, student_id, parent_id, relation_type=None, is_primary=True, is_fee_payer=True, is_emergency_contact=True):
        """Link student to parent with customizable roles"""
        try:
            if not relation_type:
                self.cursor.execute("SELECT relation_type FROM parents WHERE id = %s", (parent_id,))
                result = self.cursor.fetchone()
                relation_type = result[0] if result and result[0] else "Guardian"
            
            query = '''
                INSERT INTO student_parent 
                (student_id, parent_id, relation_type, is_primary_contact, is_fee_payer, is_emergency_contact)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                relation_type = VALUES(relation_type),
                is_primary_contact = VALUES(is_primary_contact),
                is_fee_payer = VALUES(is_fee_payer),
                is_emergency_contact = VALUES(is_emergency_contact)
            '''
            
            self.cursor.execute(query, (
                student_id, parent_id, relation_type,
                is_primary, is_fee_payer, is_emergency_contact
            ))
            
        except Exception as e:
            print(f"Error linking student to parent: {e}")

    def validate_form(self):
        """Validate form data"""
        if not self.first_name_entry.text().strip():
            QMessageBox.warning(self, "Validation Error", "First name is required")
            self.first_name_entry.setFocus()
            return False
        
        if not self.surname_entry.text().strip():
            QMessageBox.warning(self, "Validation Error", "Surname is required")
            self.surname_entry.setFocus()
            return False
        
        if not self.reg_no_entry.text().strip():
            QMessageBox.warning(self, "Validation Error", "Registration number is required")
            self.reg_no_entry.setFocus()
            return False
        
        # Check if registration number already exists (for new students or different student)
        reg_no = self.reg_no_entry.text().strip()
        if self.current_student_id:
            query = "SELECT id FROM students WHERE regNo = %s AND id != %s AND is_active = TRUE"
            self.cursor.execute(query, (reg_no, self.current_student_id))
        else:
            query = "SELECT id FROM students WHERE regNo = %s AND is_active = TRUE"
            self.cursor.execute(query, (reg_no,))
        
        if self.cursor.fetchone():
            QMessageBox.warning(self, "Validation Error", 
                              "Registration number already exists for another student")
            self.reg_no_entry.setFocus()
            return False
        
        return True

    def clear_form(self):
        """Clear all form fields"""
        # Clear text fields
        self.first_name_entry.clear()
        self.surname_entry.clear()
        self.full_name_entry.clear()
        self.full_name_entry.clear()
        self.email_entry.clear()
        self.reg_no_entry.clear()
        self.religion_entry.clear()
        self.citizenship_entry.clear()
        self.last_school_entry.clear()
        self.class_year_entry.clear()
        self.medical_text.clear()
        self.allergies_text.clear()
        self.parent_search_entry.clear()
        
        # Reset dropdowns
        self.sex_combo.setCurrentIndex(0)
        self.grade_combo.setCurrentIndex(0)
        
        # Reset dates
        self.dob_edit.setDate(QDate.currentDate().addYears(-10))
        self.enrollment_date.setDate(QDate.currentDate())
        
        # Reset photo
        self.remove_photo()
        
        # Reset checkboxes
        self.is_active_check.setChecked(True)
        
        # Reset current student ID
        self.current_student_id = None

        # Toggle buttons
        self.toggle_save_update_buttons()

    def export_students_data(self):
        """Export students data using shared export_with_green_header method"""
        try:
            # Fetch data from database with comprehensive student information
            self.cursor.execute('''
                SELECT 
                    s.regNo, s.first_name, s.surname, s.full_name, s.sex,
                    s.date_of_birth, s.email, s.grade_applied_for, s.class_year,
                    s.enrollment_date, s.religion, s.citizenship, s.last_school,
                    s.medical_conditions, s.allergies,
                    CASE WHEN s.is_active = 1 THEN 'Yes' ELSE 'No' END,
                    GROUP_CONCAT(DISTINCT CONCAT(p.full_name, ' (', sp.relation_type, ')') 
                                ORDER BY sp.is_primary_contact DESC SEPARATOR '; ') as parents,
                    GROUP_CONCAT(DISTINCT p.phone ORDER BY sp.is_primary_contact DESC SEPARATOR '; ') as parent_phones,
                    GROUP_CONCAT(DISTINCT p.email ORDER BY sp.is_primary_contact DESC SEPARATOR '; ') as parent_emails,
                    sch.school_name
                FROM students s
                LEFT JOIN student_parent sp ON s.id = sp.student_id
                LEFT JOIN parents p ON sp.parent_id = p.id AND p.is_active = TRUE
                LEFT JOIN schools sch ON s.school_id = sch.id
                WHERE s.is_active = TRUE
                GROUP BY s.id
                ORDER BY s.surname, s.first_name
            ''')
            students = self.cursor.fetchall()
    
            if not students:
                QMessageBox.information(self, "No Data", "No student data found to export.")
                return
    
            # Define headers (must match SELECT order exactly)
            headers = [
                "Registration No", "First Name", "Surname", "Full Name", "Sex",
                "Date of Birth", "Email", "Grade Applied For", "Class Year",
                "Enrollment Date", "Religion", "Citizenship", "Last School",
                "Medical Conditions", "Allergies", "Active Status",
                "Parents/Guardians", "Parent Phone Numbers", "Parent Email Addresses",
                "School Name"
            ]
    
            # Get school name for title
            school_info = self.get_school_info()
            title = f"{school_info['name']} - STUDENTS DATA"
    
            # Use shared export method
            self.export_with_green_header(
                data=students,
                headers=headers,
                filename_prefix="students_export",
                title=title
            )
    
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export student data:\n{str(e)}")
    

    def generate_pdf_report(self):
        """Generate professional PDF report for selected student with school branding"""
        if not self.current_student_id:
            QMessageBox.warning(self, "No Selection", "Please select a student first.")
            return
        
        try:
            # Get school information
            school_query = """
                SELECT school_name, address, phone, email, logo_path 
                FROM schools WHERE id = %s LIMIT 1
            """
            school_id = getattr(self.user_session, 'school_id', 1) if self.user_session else 1
            self.cursor.execute(school_query, (school_id,))
            school_info = self.cursor.fetchone()
            
            # Use empty strings if school info not available
            school_name = school_info[0] if school_info and school_info[0] else ""
            school_address = school_info[1] if school_info and school_info[1] else ""
            school_phone = school_info[2] if school_info and school_info[2] else ""
            school_email = school_info[3] if school_info and school_info[3] else ""
            
            # Get default logo path
            default_logo = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                "static", "images", "logo.png"
            )
            school_logo = school_info[4] if school_info and school_info[4] else default_logo
    
            # Get student data
            query = '''
                SELECT s.first_name, s.surname, s.sex, s.date_of_birth, s.email,
                       s.grade_applied_for, s.class_year, s.enrollment_date, s.regNo,
                       s.religion, s.citizenship, s.last_school, s.medical_conditions, 
                       s.allergies, s.photo_path,
                       GROUP_CONCAT(DISTINCT CONCAT(p.full_name, ' (', sp.relation_type, ')') 
                                   SEPARATOR ', ') as parents
                FROM students s
                LEFT JOIN student_parent sp ON s.id = sp.student_id
                LEFT JOIN parents p ON sp.parent_id = p.id AND p.is_active = TRUE
                WHERE s.id = %s
                GROUP BY s.id
            '''
            
            self.cursor.execute(query, (self.current_student_id,))
            student = self.cursor.fetchone()
            
            if not student:
                QMessageBox.warning(self, "Error", "Student data not found")
                return
            
            # Generate PDF bytes and use internal viewer
            pdf_bytes = self.generate_student_profile_pdf_bytes(student, school_info, school_logo)
            
            # Use system's built-in PDF viewer
            try:
                from utils.pdf_utils import view_pdf
                view_pdf(pdf_bytes, parent=self)
            except ImportError:
                # Fallback - offer to save file if viewer not available
                self.save_pdf_fallback(pdf_bytes, student)
                        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate PDF: {str(e)}")
            print(f"PDF generation error: {traceback.format_exc()}")
    
    def generate_student_profile_pdf_bytes(self, student, school_info, school_logo):
        """Generate student profile PDF and return PDF bytes"""
        import tempfile
        
        # Create PDF with custom class for better formatting
        class StudentPDF(FPDF):
            def __init__(self, student_photo=None):
                super().__init__(orientation='P', unit='mm', format='A4')
                self.set_margins(15, 15, 15)
                self.set_auto_page_break(auto=False)  # Manual page breaks for better control
                self.student_photo = student_photo
            
            def header(self):
                # School logo (left side)
                if os.path.exists(school_logo):
                    try:
                        self.image(school_logo, 15, 10, 25)
                    except:
                        pass  # Skip if logo can't be loaded
                
                # Student photo (right side of header)
                if self.student_photo and os.path.exists(self.student_photo):
                    try:
                        self.image(self.student_photo, 165, 5, 30, 30)  # Shifted up to avoid line
                    except:
                        pass  # Skip if photo can't be loaded
                
                # School information (centered) - only show if available
                self.set_y(10)
                if school_info and school_info[0]:
                    self.set_font("Arial", "B", 16)
                    self.cell(0, 8, school_info[0], 0, 1, "C")
                
                if school_info and school_info[1]:
                    self.set_font("Arial", "", 10)
                    self.cell(0, 5, school_info[1], 0, 1, "C")
                
                if school_info and (school_info[2] or school_info[3]):
                    contact_info = ""
                    if school_info[2]:
                        contact_info += school_info[2]
                    if school_info[2] and school_info[3]:
                        contact_info += " | "
                    if school_info[3]:
                        contact_info += school_info[3]
                    
                    self.set_font("Arial", "", 10)
                    self.cell(0, 5, contact_info, 0, 1, "C")
                
                # Report title
                self.ln(3)
                self.set_font("Arial", "B", 14)
                self.set_text_color(70, 70, 70)  # Dark gray
                self.cell(0, 8, "STUDENT PROFILE REPORT", 0, 1, "C")
                
                # Generation info
                self.set_font("Arial", "I", 8)
                self.set_text_color(100, 100, 100)
                gen_info = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                if hasattr(self, 'user_session') and self.user_session:
                    user_name = getattr(self.user_session, 'full_name', 'System')
                    gen_info += f" | By: {user_name}"
                self.cell(0, 4, gen_info, 0, 1, "C")
                
                # Line separator
                self.line(15, self.get_y() + 2, 195, self.get_y() + 2)
                self.ln(6)
            
            def footer(self):
                self.set_y(-15)
                self.set_font("Arial", "I", 8)
                self.set_text_color(128, 128, 128)
                self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "C")
            
            def section_header(self, title, highlight=False):
                self.ln(3)  # Increased space before section
                self.set_font("Arial", "B", 11)
                if highlight:
                    self.set_fill_color(173, 216, 230)  # Light blue highlight
                    self.set_text_color(0, 0, 0)      # Black text
                else:
                    self.set_fill_color(230, 230, 230)  # Light gray background
                    self.set_text_color(70, 70, 70)     # Dark gray text
                self.cell(0, 7, title, 0, 1, "L", True)
                self.ln(3)  # Increased space after section header
            
            def add_field(self, label, value, width1=50, width2=0):
                self.set_font("Arial", "B", 10)
                self.cell(width1, 6, label, 0, 0)  # Increased height
                self.set_font("Arial", "", 10)
                if width2 == 0:
                    width2 = 195 - 15 - width1  # Calculate remaining width
                self.cell(width2, 6, str(value) if value else "N/A", 0, 1)
            
            def add_multiline_field(self, label, value):
                self.set_font("Arial", "B", 10)
                self.cell(0, 6, label, 0, 1)  # Increased height
                self.set_font("Arial", "", 10)
                text = str(value) if value else "N/A"
                self.multi_cell(0, 6, text)  # Increased line height
                self.ln(2)  # Added space after multiline fields
        
        # Create PDF instance with student photo
        photo_path = student[14]  # photo_path from query
        pdf = StudentPDF(photo_path)
        pdf.add_page()
        
        # Student Information Section
        pdf.section_header("STUDENT INFORMATION")
        
        full_name = f"{student[0] or ''} {student[1] or ''}".strip()
        pdf.add_field("Registration Number:", student[8])
        pdf.add_field("Full Name:", full_name)
        
        # Two-column layout for compact display with spacing
        current_y = pdf.get_y()
        pdf.add_field("Sex:", student[2], 50, 45)
        pdf.set_y(current_y)
        pdf.set_x(110)
        pdf.add_field("Date of Birth:", student[3].strftime("%Y-%m-%d") if student[3] else "N/A", 40, 45)
        
        current_y = pdf.get_y()
        pdf.add_field("Religion:", student[9], 50, 45)
        pdf.set_y(current_y)
        pdf.set_x(110)
        pdf.add_field("Citizenship:", student[10], 40, 45)
        
        # Academic Information Section
        pdf.section_header("ACADEMIC INFORMATION")
        
        pdf.add_field("Email:", student[4])
        pdf.add_field("Last School:", student[11])
        
        current_y = pdf.get_y()
        pdf.add_field("Grade Applied For:", student[5], 50, 45)
        pdf.set_y(current_y)
        pdf.set_x(110)
        pdf.add_field("Class Year:", student[6], 40, 45)
        
        pdf.add_field("Enrollment Date:", student[7].strftime("%Y-%m-%d") if student[7] else "N/A")
        
        # Parent/Guardian Information
        pdf.section_header("PARENT/GUARDIAN INFORMATION")
        parents = student[15] or "No parents/guardians linked"
        pdf.set_font("Arial", "", 10)
        pdf.multi_cell(0, 6, parents)  # Increased line height
        pdf.ln(3)  # Added space after section
        
        # Medical Information Section
        pdf.section_header("MEDICAL INFORMATION")
        
        pdf.set_font("Arial", "B", 10)
        pdf.cell(50, 6, "Medical Conditions:", 0, 0)  # Increased height
        pdf.set_font("Arial", "", 10)
        medical_text = student[12] or "None reported"
        pdf.multi_cell(0, 6, medical_text)  # Increased line height
        pdf.ln(2)  # Added space between fields
        
        pdf.set_font("Arial", "B", 10)
        pdf.cell(50, 6, "Allergies:", 0, 0)  # Increased height
        pdf.set_font("Arial", "", 10)
        allergies_text = student[13] or "None reported"
        pdf.multi_cell(0, 6, allergies_text)  # Increased line height
        pdf.ln(3)  # Added space after section
        
        # Declaration Section
        pdf.section_header("DECLARATION")
        pdf.set_font("Arial", "", 10)
        declaration_text = (
            "I, _____________________________________________________________, "
            "hereby declare that the information provided in this form is true "
            "and accurate to the best of my knowledge.\n\n\n"
            "Signature: ________________________________________________    "
            "Date: _________________________\n\n\n"
        )
        pdf.multi_cell(0, 4, declaration_text)  # Increased line height
        
        # FOR OFFICIAL USE ONLY Section - Highlighted
        pdf.section_header("FOR OFFICIAL USE ONLY", highlight=True)
        pdf.set_font("Arial", "", 10)
        official_text = (
            "Recommended admission to class: ____________________________________________\n\n\n"
            "On (Date): _______________________________     "
            "Signed: _________________________________\n\n"
            "                                                                                                    (Principal)"
        )
        pdf.multi_cell(0, 4, official_text)  # Increased line height
        
        # Use temporary file approach to get PDF bytes
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            temp_path = temp_file.name
        
        try:
            # Output to temporary file
            pdf.output(temp_path)
            
            # Read the file back as bytes
            with open(temp_path, 'rb') as f:
                pdf_bytes = f.read()
            
        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.unlink(temp_path)
        
        return pdf_bytes
    
    def save_pdf_fallback(self, pdf_bytes, student):
        """Fallback method to save PDF if viewer not available"""
        full_name = f"{student[0] or ''} {student[1] or ''}".strip()
        safe_name = "".join(c for c in full_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"student_profile_{safe_name}_{timestamp}.pdf"
        
        # Get save path from user
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF File", default_filename, "PDF Files (*.pdf)"
        )
        
        if file_path:
            # Save PDF bytes to file
            with open(file_path, 'wb') as f:
                f.write(pdf_bytes)
            
            # Show success message
            QMessageBox.information(
                self, 
                "Success", 
                f"Student profile report saved successfully!\nFile: {file_path}"
            )
            
            # Try to open the PDF
            try:
                if platform.system() == "Windows":
                    os.startfile(file_path)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.run(["open", file_path])
                else:  # Linux
                    subprocess.run(["xdg-open", file_path])
            except Exception as e:
                print(f"Could not open PDF: {e}")
        
    
    def import_students(self):
        if not has_permission(self.user_session, STUDENT_PERMISSIONS['import']):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to import students.")
            return
        
        """Import students from CSV or Excel file"""
        try:
            file_path, file_type = QFileDialog.getOpenFileName(
                self, 
                "Import Students File",
                "",
                "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
            )
            
            if not file_path:
                return
            
            # Show progress dialog
            progress = QProgressDialog("Importing students...", "Cancel", 0, 100, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            progress.setValue(10)
            
            # Read file based on extension
            if file_path.lower().endswith('.csv'):
                students_data = self.read_csv_file(file_path)
            elif file_path.lower().endswith(('.xlsx', '.xls')):
                students_data = self.read_excel_file(file_path)
            else:
                QMessageBox.warning(self, "Error", "Unsupported file format. Please use CSV or Excel files.")
                return
            
            if not students_data:
                QMessageBox.warning(self, "Error", "No valid data found in file.")
                return
            
            progress.setValue(30)
            
            # Validate and process data
            processed_data, errors = self.validate_import_data(students_data)
            progress.setValue(60)
            
            if errors:
                self.show_import_errors(errors, processed_data)
                return
            
            # Show preview before import
            if not self.show_import_preview(processed_data):
                QMessageBox.information(self, "Import Canceled", "Import was canceled by user.")
                return
            
            # Import to database
            success_count, error_count, import_errors = self.import_to_database(processed_data, progress)
            progress.setValue(100)
            progress.close()
            
            # Show results
            result_msg = f"Import completed!\n"
            result_msg += f"Successfully imported: {success_count} students\n"
            if error_count > 0:
                result_msg += f"Failed to import: {error_count} students\n"
                result_msg += f"Check the error log for details."
            
            QMessageBox.information(self, "Import Results", result_msg)
            
            if import_errors:
                self.show_import_errors(import_errors, [])
            
            # Refresh the students table
            self.load_students()
            
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import students: {str(e)}")
    
    def read_csv_file(self, file_path):
        """Read CSV file and return student data"""
        try:
            students_data = []
            
            with open(file_path, 'r', encoding='utf-8') as file:
                # Try to detect delimiter
                sample = file.read(1024)
                file.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
                
                reader = csv.DictReader(file, delimiter=delimiter)
                
                for row_num, row in enumerate(reader, start=2):  # Start at 2 for header row
                    if any(row.values()):  # Skip empty rows
                        row['_row_number'] = row_num
                        students_data.append(row)
            
            return students_data
            
        except Exception as e:
            QMessageBox.critical(self, "CSV Error", f"Failed to read CSV file: {str(e)}")
            return []
    
    def read_excel_file(self, file_path):
        """Read Excel file and return student data"""
        try:
            students_data = []
            
            # Try to read with pandas first
            df = pd.read_excel(file_path, engine='openpyxl')
            
            # Convert to list of dictionaries
            for index, row in df.iterrows():
                row_dict = row.to_dict()
                row_dict['_row_number'] = index + 2  # +2 for header and 0-based index
                
                # Skip empty rows
                if any(pd.notna(val) and str(val).strip() for val in row_dict.values() if not str(val).startswith('_')):
                    students_data.append(row_dict)
            
            return students_data
            
        except Exception as e:
            QMessageBox.critical(self, "Excel Error", f"Failed to read Excel file: {str(e)}")
            return []
    
    def validate_import_data(self, raw_data):
        """Validate imported data and return processed data with errors"""
        processed_data = []
        errors = []
        
        # Expected columns mapping (file_column -> db_column)
        column_mapping = {
            # Basic info
            'first_name': 'first_name',
            'firstname': 'first_name',
            'fname': 'first_name',
            'surname': 'surname',
            'lastname': 'surname',
            'lname': 'surname',
            'full_name': 'full_name',
            'name': 'full_name',
            
            # Contact info
            'email': 'email',
            'email_address': 'email',
            'sex': 'sex',
            'gender': 'sex',
            'date_of_birth': 'date_of_birth',
            'dob': 'date_of_birth',
            'birth_date': 'date_of_birth',
            
            # Academic info
            'reg_no': 'regNo',
            'registration_number': 'regNo',
            'regno': 'regNo',
            'student_id': 'regNo',
            'grade': 'grade_applied_for',
            'grade_applied_for': 'grade_applied_for',
            'class': 'grade_applied_for',
            'class_year': 'class_year',
            'year': 'class_year',
            'enrollment_date': 'enrollment_date',
            'admission_date': 'enrollment_date',
            
            # Additional info
            'religion': 'religion',
            'citizenship': 'citizenship',
            'nationality': 'citizenship',
            'last_school': 'last_school',
            'previous_school': 'last_school',
            'medical_conditions': 'medical_conditions',
            'medical': 'medical_conditions',
            'allergies': 'allergies',
            
            # Parent info
            'parent_name': 'parent_name',
            'guardian_name': 'parent_name',
            'parent_phone': 'parent_phone',
            'guardian_phone': 'parent_phone',
            'parent_email': 'parent_email',
            'guardian_email': 'parent_email',
        }
        
        for row_num, row_data in enumerate(raw_data):
            try:
                # Normalize column names (lowercase, remove spaces/underscores)
                normalized_row = {}
                for key, value in row_data.items():
                    if not key.startswith('_'):
                        clean_key = str(key).lower().replace(' ', '_').replace('-', '_')
                        normalized_row[clean_key] = value
                
                # Map to database columns
                student_data = {'_row_number': row_data.get('_row_number', row_num + 1)}
                
                for file_col, db_col in column_mapping.items():
                    if file_col in normalized_row:
                        value = normalized_row[file_col]
                        if pd.notna(value) and str(value).strip():
                            student_data[db_col] = str(value).strip()
                
                # Validate required fields
                row_errors = []
                
                # Check for required fields
                if not student_data.get('first_name') and not student_data.get('full_name'):
                    row_errors.append("Missing first name or full name")
                
                if not student_data.get('surname') and not student_data.get('full_name'):
                    row_errors.append("Missing surname or full name")
                
                # Generate full name if missing
                if not student_data.get('full_name'):
                    first_name = student_data.get('first_name', '').strip()
                    surname = student_data.get('surname', '').strip()
                    if first_name or surname:
                        student_data['full_name'] = f"{first_name} {surname}".strip()
                
                # Split full name if individual names missing
                if student_data.get('full_name') and not student_data.get('first_name'):
                    name_parts = student_data['full_name'].strip().split()
                    if len(name_parts) >= 2:
                        student_data['first_name'] = name_parts[0]
                        student_data['surname'] = ' '.join(name_parts[1:])
                    elif len(name_parts) == 1:
                        student_data['first_name'] = name_parts[0]
                        student_data['surname'] = ""
                
                # Validate email format
                if student_data.get('email'):
                    email = student_data['email']
                    if '@' not in email or '.' not in email.split('@')[-1]:
                        row_errors.append(f"Invalid email format: {email}")
                
                # Validate sex/gender
                if student_data.get('sex'):
                    sex = student_data['sex'].lower()
                    if sex in ['m', 'male', 'boy']:
                        student_data['sex'] = 'Male'
                    elif sex in ['f', 'female', 'girl']:
                        student_data['sex'] = 'Female'
                    else:
                        student_data['sex'] = 'Male'  # Default
                
                # Parse date of birth
                if student_data.get('date_of_birth'):
                    parsed_date = self.parse_date(student_data['date_of_birth'])
                    if parsed_date:
                        student_data['date_of_birth'] = parsed_date
                    else:
                        row_errors.append(f"Invalid date format: {student_data['date_of_birth']}")
                
                # Parse enrollment date
                if student_data.get('enrollment_date'):
                    parsed_date = self.parse_date(student_data['enrollment_date'])
                    if parsed_date:
                        student_data['enrollment_date'] = parsed_date
                    else:
                        student_data['enrollment_date'] = datetime.now().date()
                else:
                    student_data['enrollment_date'] = datetime.now().date()
                
                # Generate registration number if missing
                if not student_data.get('regNo'):
                    student_data['regNo'] = self.generate_import_reg_no()
                
                # Set defaults
                student_data['is_active'] = True
                student_data['school_id'] = 1  # Default school
                
                if row_errors:
                    errors.append({
                        'row': student_data['_row_number'],
                        'errors': row_errors,
                        'data': student_data
                    })
                else:
                    processed_data.append(student_data)
                    
            except Exception as e:
                errors.append({
                    'row': row_data.get('_row_number', row_num + 1),
                    'errors': [f"Processing error: {str(e)}"],
                    'data': row_data
                })
        
        return processed_data, errors
    
    def generate_import_reg_no(self):
        """Generate registration number for import"""
        try:
            school_prefix = self.get_school_prefix()
            year_suffix = datetime.now().year % 100
            rand_num = random.randint(1000, 9999)
            return f"{school_prefix}{year_suffix:02d}{rand_num}-ST"
        except:
            return f"IMP{datetime.now().year % 100:02d}{random.randint(1000, 9999)}-ST"

    def parse_date(self, date_input):
        """
        Parse various date formats and return a date object.
        Supports: 'YYYY-MM-DD', 'DD/MM/YYYY', 'MM/DD/YYYY', 'DD-MM-YYYY', etc.
        Returns None if parsing fails.
        """
        if not date_input:
            return None
    
        # Convert to string if it's a float/int (e.g., from Excel)
        if isinstance(date_input, (float, int)):
            # Excel serial dates are not handled here; assume pandas already converted
            return None
    
        str_date = str(date_input).strip()
    
        # Handle pandas.Timestamp if passed
        if 'Timestamp' in str(type(date_input)):
            try:
                return date_input.date()  # Convert to datetime.date
            except:
                return None
    
        # Common date formats to try
        formats = [
            '%Y-%m-%d',      # 2024-05-15
            '%d/%m/%Y',      # 15/05/2024
            '%m/%d/%Y',      # 05/15/2024
            '%d-%m-%Y',      # 15-05-2024
            '%m-%d-%Y',      # 05-15-2024
            '%Y/%m/%d',      # 2024/05/15
            '%d.%m.%Y',      # 15.05.2024
            '%m.%d.%Y',      # 05.15.2024
            '%B %d, %Y',     # May 15, 2024
            '%b %d, %Y',     # May 15, 2024
            '%Y-%m-%d %H:%M:%S',  # 2024-05-15 10:30:00
            '%Y-%m-%d %H:%M',     # 2024-05-15 10:30
        ]
    
        for fmt in formats:
            try:
                return datetime.strptime(str_date, fmt).date()
            except ValueError:
                continue
    
        # If none work, return None
        return None
    
    def import_to_database(self, processed_data, progress_dialog):
        """Import processed data to database"""
        success_count = 0
        error_count = 0
        import_errors = []
        total_records = len(processed_data)
    
        for idx, student_data in enumerate(processed_data):
            try:
                reg_no = student_data.get('regNo')
                if not reg_no:
                    error_count += 1
                    import_errors.append({
                        'row': student_data.get('_row_number', idx + 1),
                        'error': "Missing registration number",
                        'student': student_data.get('full_name', 'Unknown')
                    })
                    continue
    
                # 🔍 CHECK: Does an active student with this regNo already exist?
                self.cursor.execute(
                    "SELECT id FROM students WHERE regNo = %s AND is_active = TRUE",
                    (reg_no,)
                )
                if self.cursor.fetchone():
                    # 🟡 Skip duplicate
                    import_errors.append({
                        'row': student_data.get('_row_number', idx + 1),
                        'error': f"Student with regNo '{reg_no}' already exists",
                        'student': student_data.get('full_name', 'Unknown')
                    })
                    error_count += 1
                    continue  # 🔴 Skip insertion

                # Optional: Add this after regNo check
                elif student_data.get('date_of_birth') and student_data.get('full_name'):
                    self.cursor.execute("""
                        SELECT id FROM students 
                        WHERE full_name = %s AND date_of_birth = %s AND is_active = TRUE
                    """, (student_data['full_name'], student_data['date_of_birth']))
                    if self.cursor.fetchone():
                        import_errors.append({
                            'row': student_data.get('_row_number'),
                            'error': "Student with same name and DOB already exists",
                            'student': student_data['full_name']
                        })
                        error_count += 1
                        continue
    
                # ✅ No duplicate — proceed with insert
                current_time = datetime.now()
                insert_query = '''
                    INSERT INTO students (
                        school_id, first_name, surname, full_name, sex, date_of_birth,
                        religion, citizenship, email, last_school, grade_applied_for, 
                        class_year, enrollment_date, regNo, is_active, medical_conditions, 
                        allergies, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                '''
                values = (
                    student_data.get('school_id', 1),
                    student_data.get('first_name', ''),
                    student_data.get('surname', ''),
                    student_data.get('full_name', ''),
                    student_data.get('sex', 'Male'),
                    student_data.get('date_of_birth'),
                    student_data.get('religion'),
                    student_data.get('citizenship'),
                    student_data.get('email'),
                    student_data.get('last_school'),
                    student_data.get('grade_applied_for'),
                    student_data.get('class_year'),
                    student_data.get('enrollment_date'),
                    reg_no,
                    student_data.get('is_active', True),
                    student_data.get('medical_conditions'),
                    student_data.get('allergies'),
                    current_time,
                    current_time
                )
                self.cursor.execute(insert_query, values)
                student_id = self.cursor.lastrowid
    
                # Link parent if provided
                if student_data.get('parent_name'):
                    self.create_parent_from_import(student_id, student_data)
    
                success_count += 1
    
            except Exception as e:
                error_count += 1
                import_errors.append({
                    'row': student_data.get('_row_number', idx + 1),
                    'error': str(e),
                    'student': student_data.get('full_name', 'Unknown')
                })
    
            # Update progress
            progress = int((idx + 1) / total_records * 40) + 60
            if progress_dialog:
                progress_dialog.setValue(progress)
    
        # Commit all changes
        self.db_connection.commit()
        return success_count, error_count, import_errors
    
    def create_parent_from_import(self, student_id, student_data):
        """Create parent record from import data"""
        try:
            parent_name = student_data.get('parent_name', '').strip()
            if not parent_name:
                return
            
            # Check if parent already exists
            self.cursor.execute(
                "SELECT id FROM parents WHERE full_name = %s AND is_active = TRUE",
                (parent_name,)
            )
            existing_parent = self.cursor.fetchone()
            
            if existing_parent:
                parent_id = existing_parent[0]
            else:
                # Create new parent
                parent_query = '''
                    INSERT INTO parents (full_name, phone, email, relation, is_active)
                    VALUES (%s, %s, %s, %s, %s)
                '''
                parent_values = (
                    parent_name,
                    student_data.get('parent_phone'),
                    student_data.get('parent_email'),
                    'Guardian',
                    True
                )
                self.cursor.execute(parent_query, parent_values)
                parent_id = self.cursor.lastrowid
            
            # Link student to parent
            link_query = '''
                INSERT INTO student_parent (student_id, parent_id, relation_type, is_primary_contact, is_fee_payer, is_emergency_contact)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE relation_type = VALUES(relation_type)
            '''
            self.cursor.execute(link_query, (student_id, parent_id, 'Guardian', True, True, True))
            
        except Exception as e:
            print(f"Error creating parent from import: {e}")
    
    def show_import_errors(self, errors, processed_data):
        """Show import errors and warnings dialog"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Import Results")
        dialog.resize(800, 600)
        dialog.setModal(True)
        layout = QVBoxLayout(dialog)
    
        # Summary
        error_count = sum(1 for e in errors if e.get('type') != 'warning')
        warning_count = sum(1 for e in errors if e.get('type') == 'warning')
    
        if error_count > 0:
            summary = QLabel(f"<b style='color: red;'>{error_count} error(s) found</b>")
        if warning_count > 0:
            warning_label = QLabel(f"<b style='color: orange;'>{warning_count} duplicate(s) detected - new regNo generated</b>")
            warning_label.setStyleSheet("padding: 5px;")
            layout.addWidget(warning_label)
    
        # Error & Warning Details
        error_text = QTextEdit()
        error_text.setReadOnly(True)
        error_content = []
    
        for error in errors:
            row_num = error.get('row', 'Unknown')
            msg = error.get('error', 'Unknown error')
            student = error.get('student', 'Unknown')
            error_type = error.get('type', 'error')
    
            prefix = "⚠️ Warning" if error_type == 'warning' else "❌ Error"
            color = "orange" if error_type == 'warning' else "red"
    
            error_content.append(f'<span style="color: {color};"><b>{prefix} (Row {row_num}):</b></span>')
            error_content.append(f"  • {msg}")
            error_content.append(f"  • Student: {student}")
            error_content.append("")
    
        error_text.setHtml('<br>'.join(error_content))
        layout.addWidget(error_text)
    
        # Buttons
        btn_layout = QHBoxLayout()
        if processed_data:
            import_btn = QPushButton(f"Import {len(processed_data)} Valid Records")
            import_btn.setStyleSheet("background-color: #28a745; color: white; font-weight: bold;")
            import_btn.clicked.connect(lambda: dialog.accept())
            btn_layout.addWidget(import_btn)
    
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(close_btn)
    
        layout.addLayout(btn_layout)
        dialog.exec_()
    
    def import_valid_records(self, processed_data, dialog):
        """Import only the valid records after showing errors"""
        try:
            progress = QProgressDialog("Importing valid records...", "Cancel", 0, 100, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            success_count, error_count, import_errors = self.import_to_database(processed_data, progress)
            progress.close()
            dialog.accept()
            
            result_msg = f"Import completed!\n"
            result_msg += f"Successfully imported: {success_count} students\n"
            if error_count > 0:
                result_msg += f"Failed to import: {error_count} students"
            
            QMessageBox.information(self, "Import Results", result_msg)
            self.load_students()
            
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import valid records: {str(e)}")

    def show_import_preview(self, processed_data):
        """Show a preview of the data to be imported"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Import Preview")
        dialog.resize(1000, 500)
        dialog.setModal(True)
    
        layout = QVBoxLayout(dialog)
    
        title = QLabel("Preview of Data to be Imported (First 20 Rows)")
        title.setStyleSheet("font-weight: bold; font-size: 14px; padding: 10px;")
        layout.addWidget(title)
    
        # Table
        table = QTableWidget()
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels([
            "Full Name", "Reg No", "Grade", "Class Year", "Email", "DOB", "Parent", "Phone"
        ])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
    
        # Take first 20 rows
        preview_data = processed_data[:20]
        table.setRowCount(len(preview_data))
    
        for row_idx, data in enumerate(preview_data):
            full_name = data.get('full_name', '')
            reg_no = data.get('regNo', '')
            grade = data.get('grade_applied_for', '')
            class_year = data.get('class_year', '')
            email = data.get('email', '')
            dob = data.get('date_of_birth', '')
            parent = data.get('parent_name', '')
            phone = data.get('parent_phone', '')
    
            table.setItem(row_idx, 0, QTableWidgetItem(full_name))
            table.setItem(row_idx, 1, QTableWidgetItem(reg_no))
            table.setItem(row_idx, 2, QTableWidgetItem(grade))
            table.setItem(row_idx, 3, QTableWidgetItem(class_year))
            table.setItem(row_idx, 4, QTableWidgetItem(email))
            table.setItem(row_idx, 5, QTableWidgetItem(str(dob) if dob else ""))
            table.setItem(row_idx, 6, QTableWidgetItem(parent))
            table.setItem(row_idx, 7, QTableWidgetItem(phone))
    
        layout.addWidget(table)
    
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
    
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dialog.reject)
    
        import_btn = QPushButton("Start Import")
        import_btn.setStyleSheet("background-color: #28a745; color: white; font-weight: bold;")
        import_btn.clicked.connect(lambda: dialog.accept())
    
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(import_btn)
        layout.addLayout(btn_layout)
    
        # Show dialog
        if dialog.exec() == QDialog.Accepted:
            return True  # User confirmed import
        return False  # User canceled


def main():
    """Main function to run the application"""
    app = QApplication(sys.argv)
    
    # Set application properties
    app.setApplicationName("Students Management System")
    app.setApplicationVersion("2.0")
    app.setOrganizationName("Winspire Learning Hub")
    
    # Create and show the main window
    window = StudentsForm()
    window.show()
    
    # Run the application
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()


    def closeEvent(self, event):
        """Handle close event to clean up database connection"""
        try:
            if hasattr(self, 'cursor'):
                self.cursor.close()
            if hasattr(self, 'db_connection'):
                self.db_connection.close()
        except:
            pass
        event.accept()

    def __del__(self):
        """Close database connection when object is destroyed"""
        try:
            if hasattr(self, 'cursor'):
                self.cursor.close()
            if hasattr(self, 'db_connection'):
                self.db_connection.close()
        except:
            pass