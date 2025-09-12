# ui/audit_base_form.py
from PySide6.QtWidgets import QWidget, QLineEdit, QMessageBox, QFileDialog, QPushButton
from PySide6.QtGui import QFont, QCursor
from PySide6.QtCore import Qt
from typing import Optional, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell  # For checking merged cells
from datetime import datetime
import os
import platform
import subprocess
from models.models import get_db_connection


class AuditBaseForm(QWidget):
    """
    Base form class with shared styling, user session, and audit logging.
    All forms should inherit from this.
    """

    def __init__(self, parent=None, user_session: Optional[Dict[str, Any]] = None):
        super().__init__(parent)
        self.user_session = user_session
        self.colors = {}
        self.fonts = {}
        self.db_connection = None
        self.cursor = None
        self.setup_styling()

    # Inside ui/audit_base_form.py
    def _ensure_connection(self):
        """Ensure database connection is active and reconnected if needed."""
        try:
            # Check if connection exists and is actually open/usable
            if (not hasattr(self, 'db_connection') or self.db_connection is None or
                not self.db_connection.is_connected()): # Use is_connected()
                
                # Need to (re)connect
                print("AuditBaseForm: Establishing new database connection...")
                self.db_connection = get_db_connection()
                if self.db_connection is None:
                     raise Exception("get_db_connection returned None")
                print("AuditBaseForm: Database connection established.")
    
            # Ensure cursor exists and is usable
            # Simple check: if cursor is None or assumed closed, recreate it.
            # Note: mysql-connector cursors don't typically have a 'closed' attribute easily checked.
            # It's safer to recreate if db_connection was just established or if cursor is None.
            if (not hasattr(self, 'cursor') or self.cursor is None or
                not hasattr(self, 'db_connection') or self.db_connection is None or
                not self.db_connection.is_connected()): # Check connection again before creating cursor
                
                 if self.db_connection and self.db_connection.is_connected():
                     print("AuditBaseForm: Creating new cursor...")
                     self.cursor = self.db_connection.cursor(buffered=True, dictionary=True) # Add dictionary=True if needed
                     print("AuditBaseForm: Cursor created.")
                 else:
                     raise Exception("Cannot create cursor: Database connection is not available or not connected.")
    
        except mysql.connector.Error as db_err: # Catch specific DB errors
            print(f"AuditBaseForm._ensure_connection: MySQL Error: {db_err}")
            # Close connection if it exists but is faulty
            if hasattr(self, 'db_connection') and self.db_connection:
                try:
                    self.db_connection.close()
                except:
                    pass
                self.db_connection = None
                self.cursor = None
            raise Exception(f"Database connection error: {db_err}")
    
        except Exception as e:
            print(f"AuditBaseForm._ensure_connection: General Error: {e}")
            # Ensure cleanup on failure
            if hasattr(self, 'db_connection') and self.db_connection:
                try:
                    self.db_connection.close()
                except:
                    pass
            self.db_connection = None
            self.cursor = None
            raise Exception(f"Failed to ensure database connection: {e}")

    def setup_hand_cursor(self, widget):
        """Set hand cursor for interactive elements like buttons"""
        if isinstance(widget, QPushButton):
            widget.setCursor(QCursor(Qt.PointingHandCursor))
        
    def apply_hand_cursor_to_buttons(self):
        """Apply hand cursor to all buttons in the form"""
        for button in self.findChildren(QPushButton):
            self.setup_hand_cursor(button)

    def setup_styling(self):
        """Set up shared colors, fonts, and QSS styling"""
        # === Color Palette ===
        self.colors = {
            'primary': '#0056b3',           # Blue
            'primary_dark': '#007BFF',
            'success': '#28a745',          # Green
            'danger': '#dc3545',           # Red
            'warning': '#ffc107',          # Amber
            'info': '#17a2b8',             # Cyan
            'secondary': '#6c757d',        # Gray (changed from dark slate)
            'secondary_hover': '#5a6268',  # Darker gray on hover
            'text_primary': '#212529',     # Dark text
            'text_secondary': '#6c757d',
            'background': '#ffffff',       # White
            'surface': '#f8f9fa',          # Light gray surface
            'border': '#dee2e6',           # Light border
            'input_border': '#adb5bd',     #ced4da
            'input_focus': '#80bdff',
            'input_background': '#ffffff',
            'table_header': '#455A64',     # gray #e6edf3 
            'table_header_dark': '#263238', # Darker
            'light': '#f1f5f9',             # Light background (for scrollbars)
            'main_tab_gradient_start': '#0066cc',  # Main tab gradient start
            'main_tab_gradient_end': '#004499',    # Main tab gradient end
            'main_tab_border': '#003366',          # Main tab border
            'tab_active': '#ffffff',               # Active tab background
            'tab_hover': 'rgba(255,255,255,0.15)', # Tab hover effect
            'tab_checked': 'rgba(255,255,255,0.2)', # Checked tab background
            'ribbon_background': '#ffffff'         # White ribbon background
        }
    
        # === Fonts ===
        self.fonts = {
            'label': QFont("Arial", 14, QFont.Weight.Bold),
            'entry': QFont("Arial", 14),
            'button': QFont("Arial", 12, QFont.Weight.Bold),
            'table': QFont("Tahoma", 11),  # Slightly smaller for better fit
            'table_header': QFont("Tahoma", 12, QFont.Weight.Bold),  # Reduced from 13
            'section': QFont("Arial", 16, QFont.Weight.Bold),
            'title': QFont("Arial", 18, QFont.Weight.Bold),
            'tab': QFont("Segoe UI", 13, QFont.Weight.Medium),      # Tab font
            'tab_active': QFont("Segoe UI", 13, QFont.Weight.Bold)  # Active tab font
        }
    
        # === QSS Styling ===
        self.setStyleSheet(self.get_global_stylesheet())

    def get_global_stylesheet(self):
        """Get the complete global stylesheet"""
        return f"""
            /* === GLOBAL BASE STYLES === */
            QWidget {{
                background-color: #f5f5f5;
                color: {self.colors['text_primary']};
                font-family: 'Segoe UI', Arial, sans-serif;
                font-size: 13px;
            }}
    
            /* === MAIN NAVIGATION TABS (ENHANCED WITH ROUNDED CORNERS) === */
            QToolBar#mainTabs {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 {self.colors['main_tab_gradient_start']}, 
                    stop: 1 {self.colors['main_tab_gradient_end']});
                color: white;
                spacing: 2px;
                padding: 4px 8px;
                border: none;
                border-bottom: 2px solid {self.colors['main_tab_border']};
                border-top: none;
                min-height: 40px;
                max-height: 40px;
            }}
            
            /* Menu Toggle Button (Left Side) */
            QPushButton#menuToggle {{
                background: rgba(255, 255, 255, 0.1);
                color: white;
                padding: 6px 10px;
                border: none;
                border-radius: 8px;
                min-width: 36px;
                max-width: 44px;
                font-size: 16px;
                font-weight: bold;
                margin: 2px;
            }}
            
            QPushButton#menuToggle:hover {{
                background: rgba(255, 255, 255, 0.25);
                border: 1px solid rgba(255, 255, 255, 0.4);
            }}
            
            QPushButton#menuToggle:checked {{
                background: {self.colors['tab_checked']};
                border: 1px solid rgba(255, 255, 255, 0.5);
                font-weight: bold;
            }}
            
            /* Menu Action Buttons (Sidebar) */
            QPushButton#menuAction {{
                background-color: {self.colors['background']};
                color: {self.colors['text_primary']};
                border: 1px solid {self.colors['border']};
                border-radius: 8px;
                padding: 10px 14px;
                font-size: 13px;
                font-weight: 500;
                text-align: left;
                min-height: 40px;
                max-height: 40px;
            }}
            QPushButton#menuAction:hover {{
                background-color: {self.colors['surface']};
                border-color: #b0b0b0;
            }}
            QPushButton#menuAction:pressed {{
                background-color: #e9ecef;
                font-weight: bold;
            }}

            
            /* Main Tab Buttons (Center) */
            QPushButton#tabButton {{
                background: transparent;
                color: white;
                padding: 8px 12px;
                border: none;
                border-radius: 4px;
                min-width: 70px;
                font-size: 13px;
                font-weight: 500;
                margin: 2px -2px;
            }}
            
            QPushButton#tabButton:hover {{
                background: #f8f9fa;
                border: 1px solid #dee2e6;
                color: #212529;
            }}
            
            QPushButton#tabButton:pressed {{
                background: #f1f3f4;
                border: 1px solid #ced4da;
                color: #212529;
            }}
            
            QPushButton#tabButton:checked {{
                background: #f8f9fa;
                border: 1px solid #dee2e6;
                color: #212529;
                font-weight: bold;
                /* Simulate "inset" */
                padding-top: 9px;
                padding-bottom: 7px;
            }}
    
            /* Profile Section Styling (Right Side) */
            QWidget#profileContainer {{
                background: transparent;
                border-radius: 16px;
                padding: 8px 14px;
                margin: 4px;
            }}
            
            QLabel#userName {{
                color: white;
                font-size: 12px;
                font-weight: 600;
                padding-right: 8px;
                background: transparent;
            }}
            
            QLabel#profilePic {{
                border: 2px solid rgba(255, 255, 255, 0.2);
                border-radius: 16px;          /* ✅ Now a perfect circle */
                background: transparent;
            }}
            
            QLabel#profilePic:hover {{
                border: 2px solid rgba(255, 255, 255, 0.6);  /* ✅ Stronger hover */
            }}
            
            QPushButton#ribbonToggle {{
                background: rgba(255, 255, 255, 0.15);
                color: white;
                padding: 2px 4px;
                border: 1px solid rgba(255, 255, 255, 0.2);
                border-radius: 4px;
                font-size: 14px;
                font-weight: bold;
                min-width: 10px;
                margin-left: 6px;
            }}
            
            QPushButton#ribbonToggle:hover {{
                background: rgba(255, 255, 255, 0.25);
                border: 1px solid rgba(255, 255, 255, 0.4);
            }}
    
            /* === RIBBON CONTAINER === */
            #ribbonContainer {{
                background: {self.colors['ribbon_background']};
                border-bottom: 1px solid {self.colors['border']};
                border-top: none;
                padding: 8px;
            }}
            
            #ribbonPanel {{
                background-color: transparent;
                border: none;
            }}
    
            /* === GROUP BOXES === */
            QGroupBox {{
                font-weight: bold;
                font-size: 16px;
                color: #2c3e50;
                border: 2px solid {self.colors['border']};
                border-radius: 12px;
                margin-top: 16px;
                padding-top: 15px;
                background-color: {self.colors['background']};
            }}
            
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px;
                background-color: {self.colors['background']};
                color: {self.colors['primary']};
                font-weight: bold;
            }}
    
            /* === LABELS === */
            QLabel {{
                font-size: 14px;
                font-weight: 500;
                color: {self.colors['text_primary']};
                padding: 2px;
            }}
    
            /* === INPUT FIELDS === */
            QLineEdit, QComboBox, QDateEdit {{
                padding: 10px 14px;
                border: 2px solid {self.colors['input_border']};
                border-radius: 8px;
                font-size: 13px;
                background-color: {self.colors['input_background']};
                color: {self.colors['text_primary']};
                min-height: 18px;
            }}
            
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{
                border-color: {self.colors['input_focus']};
                background-color: #f8f9fa;
            }}
            
            QLineEdit:disabled, QComboBox:disabled {{
                background-color: #f1f5f9;
                color: #64748b;
                border-color: #cbd5e1;
            }}
    
            /* === COMBO BOX DROPDOWN === */
            QComboBox::drop-down {{
                border: none;
                width: 20px;
            }}
            
            QComboBox::down-arrow {{
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid {self.colors['text_secondary']};
            }}
    
            /* === BUTTONS === */
            QPushButton {{
                border: none;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: 600;
                font-size: 12px;
                min-height: 28px;
                max-height: 32px;
                background-color: {self.colors['primary']};
                color: white;
            }}
    
            QPushButton:hover {{
                background-color: {self.colors['primary_dark']};
                border: 1px solid rgba(255, 255, 255, 0.3);
            }}
            
            QPushButton:pressed {{
                padding: 9px 15px 7px 17px;
                background-color: {self.colors['primary_dark']};
            }}
    
            /* === MESSAGE BOX BUTTONS – COMPACT & BALANCED === */
            QMessageBox,
            QDialogButtonBox {{
                spacing: 12px;
                margin: 10px;
            }}
            
            QMessageBox QLabel {{
                padding: 8px 12px;
                color: {self.colors['text_primary']};
            }}
            
            QDialogButtonBox QPushButton {{
                min-height: 24px;
                max-height: 24px;
                min-width: 65px;
                max-width: 95px;
                padding: 6px 10px;
                font-size: 12px;
                font-weight: 600;
                border-radius: 6px;
                background-color: {self.colors['secondary']};
                color: white;
            }}
            
            QDialogButtonBox QPushButton:hover {{
                background-color: {self.colors['secondary_hover']};
            }}
            
            QDialogButtonBox QPushButton[class="primary"] {{
                background-color: {self.colors['primary']};
            }}
            QDialogButtonBox QPushButton[class="primary"]:hover {{
                background-color: {self.colors['primary_dark']};
            }}
            
            QDialogButtonBox QPushButton[class="danger"] {{
                background-color: {self.colors['danger']};
            }}
            QDialogButtonBox QPushButton[class="danger"]:hover {{
                background-color: #c0392b;
            }}
            
            /* Slight spacing fix */
            QMessageBox QLabel {{
                padding: 8px 12px;
            }}
    
            /* === MESSAGE BOX BUTTONS - SPECIFIC STYLING === */
            QMessageBox QPushButton {{
                background-color: {self.colors['secondary']};
                color: white;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 600;
                min-width: 80px;
            }}
            
            QMessageBox QPushButton:hover {{
                background-color: {self.colors['secondary_hover']};
            }}
            
            /* Specific button types */
            QMessageBox QPushButton[text="OK"],
            QMessageBox QPushButton[text="Yes"],
            QMessageBox QPushButton[text="Save"] {{
                background-color: {self.colors['primary']};
            }}
            
            QMessageBox QPushButton[text="OK"]:hover,
            QMessageBox QPushButton[text="Yes"]:hover,
            QMessageBox QPushButton[text="Save"]:hover {{
                background-color: {self.colors['primary_dark']};
            }}
            
            QMessageBox QPushButton[text="Cancel"],
            QMessageBox QPushButton[text="No"],
            QMessageBox QPushButton[text="Discard"] {{
                background-color: {self.colors['danger']};
            }}
            
            QMessageBox QPushButton[text="Cancel"]:hover,
            QMessageBox QPushButton[text="No"]:hover,
            QMessageBox QPushButton[text="Discard"]:hover {{
                background-color: #c0392b;
            }}
    
            /* === BUTTON VARIANTS === */
            QPushButton[class="success"] {{
                background-color: {self.colors['success']};
                color: white;
            }}
            QPushButton[class="success"]:hover {{
                background-color: #229954;
            }}
    
            QPushButton[class="danger"] {{
                background-color: {self.colors['danger']};
                color: white;
            }}
            QPushButton[class="danger"]:hover {{
                background-color: #c0392b;
            }}
    
            QPushButton[class="warning"] {{
                background-color: {self.colors['warning']};
                color: #212529;
            }}
            QPushButton[class="warning"]:hover {{
                background-color: #e67e22;
                color: white;
            }}
    
            QPushButton[class="info"] {{
                background-color: {self.colors['info']};
                color: white;
            }}
            QPushButton[class="info"]:hover {{
                background-color: #7d3c98;
            }}
    
            QPushButton[class="primary"] {{
                background-color: {self.colors['primary']};
                color: white;
            }}
            QPushButton[class="primary"]:hover {{
                background-color: {self.colors['primary_dark']};
            }}
    
            QPushButton[class="secondary"] {{
                background-color: {self.colors['secondary']};
                color: white;
            }}
            QPushButton[class="secondary"]:hover {{
                background-color: {self.colors['secondary_hover']};
            }}
    
            /* === TABLE STYLING === */
            QTableWidget {{
                border: 2px solid {self.colors['border']};
                border-radius: 12px;
                background-color: {self.colors['background']};
                alternate-background-color: #f8fafc;
                gridline-color: {self.colors['border']};
                selection-background-color: rgba(13, 148, 136, 0.15);
                selection-color: {self.colors['text_primary']};
                font-size: 12px;
            }}
            
            QTableWidget::item {{
                padding: 6px 10px;
                border-bottom: 1px solid {self.colors['border']};
                color: {self.colors['text_primary']};
                min-height: 20px;
                max-height: 26px;
            }}
            
            QTableWidget::item:selected {{
                background-color: rgba(13, 148, 136, 0.2);
                color: {self.colors['text_primary']};
                border: 1px solid {self.colors['table_header']};
                font-weight: 600;
            }}
            
            QTableWidget::item:hover {{
                background-color: rgba(13, 148, 136, 0.1);
            }}
    
            /* === TABLE HEADER === */
            QHeaderView::section {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {self.colors['table_header']}, 
                    stop:1 {self.colors['table_header_dark']});
                color: white;
                padding: 6px 8px;
                border: none;
                font-weight: bold;
                font-size: 12px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
                min-height: 16px;
                max-height: 18px;
            }}
    
            /* Add a blue bottom border to tie to brand */
            QHeaderView::section {{
                border-bottom: 3px solid {self.colors['primary']};
            }}
    
            QHeaderView::section:first {{
                border-top-left-radius: 10px;
            }}
            
            QHeaderView::section:last {{
                border-top-right-radius: 10px;
            }}
            
            QHeaderView::section:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2563eb, stop:1 #1e40af );
            }}
    
            /* === VERTICAL HEADER (ROW NUMBERS) === */
            QTableWidget::verticalHeader {{
                background-color: #f1f5f9;
                color: #475569;
                font-size: 12px;
                font-weight: 500;
                border-right: 1px solid #cbd5e1;
                width: 30px;
            }}
            
            QTableWidget::verticalHeader::section {{
                background-color: #f1f5f9;
                color: #1e293b;
                padding: 8px;
                border-bottom: 1px solid #cbd5e1;
                font-weight: 600;
                min-height: 20px;
                max-height: 26px;
            }}
            
            QTableWidget::verticalHeader::section:selected {{
                background-color: #e2e8f0;
                color: #0f172a;
            }}
            
            QTableWidget::verticalHeader::section:hover {{
                background-color: #e2e8f0;
            }}
    
            /* === SCROLLBARS === */
            QScrollBar:vertical {{
                background: {self.colors['light']};
                width: 14px;
                border-radius: 7px;
                margin: 0px;
            }}
            
            QScrollBar::handle:vertical {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {self.colors['table_header']}, 
                    stop:1 {self.colors['table_header_dark']});
                border-radius: 7px;
                min-height: 25px;
                margin: 2px;
            }}
            
            QScrollBar::handle:vertical:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #0f766e, stop:1 {self.colors['table_header']});
            }}
    
            QScrollBar:horizontal {{
                background: {self.colors['light']};
                height: 14px;
                border-radius: 7px;
                margin: 0px;
            }}
            
            QScrollBar::handle:horizontal {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {self.colors['table_header']}, 
                    stop:1 {self.colors['table_header_dark']});
                border-radius: 7px;
                min-width: 25px;
                margin: 2px;
            }}
    
            /* === PHOTO LABEL === */
            QLabel[class="photo"] {{
                border: 2px solid {self.colors['border']};
                border-radius: 12px;
                background-color: {self.colors['surface']};
                color: #7f8c8d;
                font-size: 12px;
                text-align: center;
                padding: 16px;
                min-width: 160px;
                min-height: 160px;
                max-width: 160px;
                max-height: 160px;
            }}
    
            /* === MENU BAR === */
            QMenuBar {{
                background: transparent;
                border: none;
                color: #333;
                spacing: 3px;
            }}
            
            QMenuBar::item {{
                background: transparent;
                padding: 6px 12px;
                border-radius: 6px;
                font-weight: 500;
            }}
            
            QMenuBar::item:selected {{
                background: rgba(0, 102, 204, 0.1);
                color: {self.colors['primary']};
            }}
            
            /* === MENU POPUP === */
            QMenu {{
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 6px;
            }}
            
            QMenu::item {{
                padding: 8px 24px 8px 12px;
                border-radius: 6px;
                background-color: transparent;
            }}
            
            QMenu::item:selected {{
                background-color: #f1f5f9;
                color: {self.colors['primary_dark']};
                font-weight: 500;
            }}
    
            /* === TAB WIDGET (For Content Tabs) === */
            QTabWidget::pane {{
                border: 2px solid {self.colors['border']};
                border-radius: 8px;
                background-color: {self.colors['background']};
                margin-top: -1px;
            }}
    
            QTabBar::tab {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #e6edf3,
                    stop: 1 #d8e2ec
                );
                border: 1px solid #cacedb;
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: 500;
                color: #2c3e50;
            }}
            
            QTabBar::tab:hover:!selected {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #dce6f1,
                    stop: 1 #cacedb
                );
                border: 1px solid #b8c8db;
            }}
            
            QTabBar::tab:selected {{
                background: white;
                border-bottom: 2px solid #007BFF;
                color: #007BFF;
                font-weight: bold;
            }}

            /* Notification badge */
            #notificationBadge {{
                background-color: #e74c3c;
                color: white;
                border-radius: 8px;
                padding: 2px 6px;
                font-size: 10px;
                font-weight: bold;
                min-width: 16px;
                max-width: 16px;
                min-height: 16px;
                max-height: 16px;
            }}
            
            /* Toast notifications */
            #toastNotification {{
                background-color: #2c3e50;
                color: white;
                padding: 10px 15px;
                border-radius: 5px;
                border: 1px solid #34495e;
                font-size: 12px;
            }}
            
            /* Email action buttons */
            #emailAction {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3498db,
                    stop:1 #2980b9
                );
                color: white;
                border: 1px solid #2980b9;
            }}
            
            #emailAction:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2980b9,
                    stop:1 #2471a3
                );
            }}
            
            /* === OVERRIDE: RIBBON BUTTONS - SMALL, FLAT, NO BORDER === */
            QPushButton#ribbonButton {{
                background-color: transparent;
                border: none;
                border-radius: 4px;
                min-width: 50px;
                max-width: 50px;
                min-height: 30px;
                max-height: 30px;
                padding: 4px;
                icon-size: 24px;
                font-size: 9px;
                color: #495057;
            }}
            QPushButton#ribbonButton:hover {{
                background-color: #e0e0e0;
            }}
            QPushButton#ribbonButton:pressed {{
                background-color: #e8e8e8;
            }}
            QPushButton#ribbonButton:focus {{
                outline: none;
            }}
        """

    def create_button(self, text, callback=None, button_class="primary", parent=None):
        """Create a button with hand cursor and optional styling class"""
        button = QPushButton(text, parent)
        
        # Set hand cursor
        button.setCursor(QCursor(Qt.PointingHandCursor))
        
        # Apply CSS class if specified
        if button_class:
            button.setProperty("class", button_class)
        
        # Connect callback if provided
        if callback:
            button.clicked.connect(callback)
            
        return button

    def create_ribbon_button(self, text, callback=None, parent=None):
        """Create a ribbon-style button with hand cursor"""
        button = QPushButton(text, parent)
        button.setObjectName("ribbonButton")
        
        # Set hand cursor
        button.setCursor(QCursor(Qt.PointingHandCursor))
        
        # Connect callback if provided
        if callback:
            button.clicked.connect(callback)
            
        return button

    def showEvent(self, event):
        """Override showEvent to apply hand cursors after the widget is fully constructed"""
        super().showEvent(event)
        # Apply hand cursor to all existing buttons
        self.apply_hand_cursor_to_buttons()
        
        
    def log_audit_action(self, action: str, table_name: str, record_id: int, description: str):
        """Log an audit action to the audit_log table."""
        if not hasattr(self, 'cursor') or not hasattr(self, 'db_connection'):
            print("Error: Database cursor or connection not available.")
            return

        try:
            user_id = self.user_session.get('user_id') if self.user_session else None
            ip_address = self.user_session.get('ip_address', '127.0.0.1') if self.user_session else '127.0.0.1'

            query = """
                INSERT INTO audit_log (user_id, action, table_name, record_id, description, ip_address)
                VALUES (%s, %s, %s, %s, %s, %s)
            """
            self.cursor.execute(query, (user_id, action, table_name, record_id, description, ip_address))
            self.db_connection.commit()
        except Exception as e:
            print(f"Failed to log audit action: {e}")


    def export_with_green_header(self, data, headers, filename_prefix="export", title=None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.cell.cell import MergedCell
            from datetime import datetime
            import os
            import platform
            import subprocess
    
            # Use provided title or fallback
            display_title = title or "CBCENTRA SCHOOL MANAGEMENT SYSTEM"
    
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Export"
    
            # === STYLING ===
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')  # Dark green
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            data_font = Font(name='Arial', size=11)
            data_alignment = Alignment(horizontal='left', vertical='center')
    
            # === METADATA ROWS ===
            ws.insert_rows(0, 3)
            total_cols = len(headers)
            last_col_letter = get_column_letter(total_cols)
    
            # Title
            ws.merge_cells(f'A1:{last_col_letter}1')
            title_cell = ws['A1']
            title_cell.value = display_title.upper()
            title_cell.font = Font(name='Arial', size=16, bold=True, color='2E7D32')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
            # Subtitle
            ws.merge_cells(f'A2:{last_col_letter}2')
            subtitle_cell = ws['A2']
            subtitle_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Exported by: {self.user_session.get('full_name', 'Unknown')}"
            subtitle_cell.font = Font(name='Arial', size=10, italic=True, color='555555')
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
            ws['A3'].value = ""  # Empty row
    
            # === WRITE HEADERS (row 4) ===
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=4, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
    
            # === WRITE DATA ===
            for row_data in data:
                ws.append(row_data)
    
            # === COLUMN WIDTH AUTO-FIT (SAFE) ===
            for column_cells in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column_cells:
                    if isinstance(cell, MergedCell):
                        continue
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
    
            # === DATA FORMATTING ===
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    
            # === FREEZE HEADER ROW ===
            ws.freeze_panes = 'A5'
    
            # === FOOTER ===
            footer_row = ws.max_row + 2
            ws.merge_cells(f'A{footer_row}:{last_col_letter}{footer_row}')
            footer_cell = ws[f'A{footer_row}']
            footer_cell.value = "Generated by CBCentra School Management System"
            footer_cell.font = Font(name='Arial', size=9, italic=True, color='999999')
            footer_cell.alignment = Alignment(horizontal='center', vertical='center')
    
            # === ASK USER FOR SAVE LOCATION ===
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            suggested_filename = f"{filename_prefix}_{timestamp}.xlsx"
            
            # Default to exports folder, but let user choose
            export_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
            os.makedirs(export_dir, exist_ok=True)
            suggested_path = os.path.join(export_dir, suggested_filename)
            
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Save Export File",
                suggested_path,
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            # If user cancels, abort
            if not filename:
                return False
            
            # Ensure .xlsx extension
            if not filename.lower().endswith('.xlsx'):
                filename += '.xlsx'
            
            # === SAVE FILE ===
            try:
                wb.save(filename)
            except Exception as e:
                QMessageBox.critical(self, "Save Error", f"Could not save file:\n{str(e)}")
                return False
    
            # === OPEN FILE ===
            try:
                if platform.system() == "Windows":
                    os.startfile(filename)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.call(["open", filename])
                else:  # Linux
                    subprocess.call(["xdg-open", filename])
            except Exception as e:
                print(f"Could not open file: {e}")
    
            QMessageBox.information(
                self,
                "Success",
                f"Data exported successfully!\nSaved to: {os.path.basename(filename)}"
            )
    
        except ImportError:
            QMessageBox.critical(
                self,
                "Export Error",
                "Required library 'openpyxl' not installed.\nRun: pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export:\n{str(e)}")

    def get_school_info(self, school_id=None):
        """
        Get school info from database
        If no school_id, return default or first school
        """
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
    
            if school_id:
                cursor.execute("SELECT school_name, address, phone, email FROM schools WHERE id = %s", (school_id,))
            else:
                cursor.execute("SELECT school_name, address, phone, email FROM schools ORDER by id LIMIT 1")
    
            result = cursor.fetchone()
            conn.close()
    
            if result:
                return {
                    'name': result[0],
                    'address': result[1],
                    'phone': result[2],
                    'email': result[3]
                }
            else:
                return {
                    'name': 'CBCentra School Manangement System',
                    'address': 'N/A',
                    'phone': 'N/A',
                    'email': 'info@cbcentra.edu'
                }
        except Exception as e:
            print(f"Error fetching school info: {e}")
            return {
                'name': 'CBCentra School Manangement System',
                'address': 'N/A',
                'phone': 'N/A',
                'email': 'info@cbcentra.edu'
            }